#!/usr/bin/env python3
"""
update_onboarding_dashboard.py
Generates onboarding.html from:
  - onboarding-data/*.xlsx  (LMS Accelerate item-level data, one row per course item per person)
  - data/*.xlsx             (playbook traffic, filtered for Accelerate URLs only)
"""

import openpyxl
import os
import re
import json
import warnings
warnings.filterwarnings('ignore')

# ── LMS column indices (0-based) ─────────────────────────────────────────────
COL_FIRST      = 2   # First Name
COL_LAST       = 3   # Last Name
COL_EMAIL      = 4   # Email Address
COL_JOBTITLE   = 5   # Job Title
COL_MARKET     = 7   # Market
COL_MGR_FIRST  = 9   # ManagerFirstName
COL_MGR_LAST   = 10  # ManagerLastName
COL_MGR_EMAIL  = 11  # SUPEMAILADDR
COL_MGR_TITLE  = 12  # Manager JobTitle
COL_HIRE_DATE  = 14  # Hire Date
COL_CURRIC_ID  = 17  # Curriculum ID
COL_CURRIC_TTL = 18  # Curriculum Title
COL_CURRIC_CMP = 19  # Curriculum Complete (Yes/No)
COL_ASSIGN_DT  = 20  # Curriculum Assignment Date
COL_DAYS_REM   = 21  # Days Remaining (LMS-computed per curriculum deadline)
COL_ITEM_TTL   = 26  # Item Title
COL_ITEM_TYPE  = 23  # Item Type (ONLINE/VILT)
COL_ITEM_DATE  = 27  # Item Completion Date
COL_ITEM_STS   = 29  # Item Completion Status Description
COL_ITEM_REQ   = 30  # Item Required Date

# ── Playbook traffic columns ──────────────────────────────────────────────────
PB_FIRST = 1
PB_LAST  = 2
PB_DATE  = 9
PB_URL   = 10

# ── Sub-curricula in display order ───────────────────────────────────────────
CURRICULA = [
    ('ACCELERATE_GS', 'Getting Started'),
    ('ACCELERATE_SW', 'Sales Workflow'),
    ('ACCELERATE_CP', 'Core Portfolio'),
    ('ACCELERATE_P',  'Prospecting'),
    ('ACCELERATE_SS', 'Sales Skills'),
    ('ACCELERATE_PM', 'Pipeline Mgmt'),
]

PROGRAM_DAYS = 35

TLG = {
    'jason ackerman', 'bianca davis', 'james parker', 'resmie biba',
    'chris curtis', 'sara thompson', 'jeremy macbean', 'bradley pierce',
    'laura sefcik', 'samantha maresca', 'staci musco', 'cj homer',
    'rich moore', 'dale kinsey',
    'john lechner', 'resmie nesimi', "samantha d'angelo", 'bianca dipasquale', 'doug falk',
}

PAGE_LABELS = {
    'coreportfolio':         'Core Portfolio',
    'salesworkflow':         'Sales Workflow',
    'overview':              'Overview',
    'welcome':               'Welcome',
    'resources':             'Resources',
    'understandingsalesforce': 'Understanding Salesforce',
    'managers':              'Managers',
    'index':                 'Home',
    'prospectingskills':     'Prospecting Skills',
    'salesforceprospecting': 'Salesforce Prospecting',
    'callprep':              'Prep Essentials',
    'workingwithnumbers':    'Working With Numbers',
    'movingdeals':           'Moving Deals Forward',
    'pipelineownership':     'Pipeline Ownership',
}


def pkey(first, last):
    return f"{str(first).strip().lower()} {str(last).strip().lower()}"


def _date(val):
    """Return YYYY-MM-DD string from a datetime object, YYYY-MM-DD string, or LMS M/D/YYYY format."""
    if not val:
        return None
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s:
        return None
    # LMS exports dates as "M/D/YYYY timezone" e.g. "6/11/2026 US/Alaska"
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m:
        return f'{int(m.group(3)):04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}'
    return s


def extract_date(fname):
    m = re.search(r'(\d{4}-\d{2})', os.path.basename(fname))
    return m.group(1) if m else '0000-00'


def _file_date_label(fname):
    """Return a human-readable 'Data through' date from a filename."""
    import datetime as _dt
    _MONTHS = ['January','February','March','April','May','June',
               'July','August','September','October','November','December']
    base = os.path.basename(str(fname))
    m = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', base)
    if m:
        return f'{_MONTHS[int(m.group(1))-1]} {int(m.group(2))}, {m.group(3)}'
    m = re.search(r'(\d{2})\.(\d{2})(?!\.\d)', base)
    if m:
        yr = _dt.datetime.today().year
        return f'{_MONTHS[int(m.group(1))-1]} {int(m.group(2))}, {yr}'
    m = re.search(r'(\d{4})-(\d{2})', base)
    if m:
        return f'{_MONTHS[int(m.group(2))-1]} {m.group(1)}'
    dt = _dt.datetime.fromtimestamp(os.path.getmtime(str(fname)))
    return f'{_MONTHS[dt.month-1]} {dt.day}, {dt.year}'


def page_label(url):
    m = re.search(r'/([^/]+?)(?:\.html)?$', url.rstrip('/'))
    if not m or m.group(1) in ('accelerate_sales_playbook', ''):
        return 'Home'
    slug = m.group(1).lower()
    return PAGE_LABELS.get(slug, slug.replace('-', ' ').title())


def load_lms():
    folder = 'onboarding-data'
    files = sorted(
        [f for f in os.listdir(folder) if f.endswith('.xlsx')],
        key=extract_date
    )
    seen = {}  # email -> record (latest file wins)

    for fname in files:
        wb = openpyxl.load_workbook(f"{folder}/{fname}", read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 2:
            continue
        data = all_rows[1:]

        from collections import defaultdict
        by_email = defaultdict(list)
        for r in data:
            email = str(r[COL_EMAIL] or '').strip().lower()
            if email:
                by_email[email].append(r)

        for email, prows in by_email.items():
            first = str(prows[0][COL_FIRST] or '').strip()
            last  = str(prows[0][COL_LAST]  or '').strip()

            if pkey(first, last) in TLG:
                continue

            # Parent curriculum row (overall program)
            parent = next((r for r in prows if r[COL_CURRIC_ID] == 'ACCELERATE'), None)
            assign_date  = parent[COL_ASSIGN_DT]  if parent else None
            overall_done = (parent[COL_CURRIC_CMP] == 'Yes') if parent else False
            days_rem_lms = None  # computed in JS from assignDate

            # Per sub-curriculum data
            curricula_data = {}
            for cid, cname in CURRICULA:
                crows = [r for r in prows if r[COL_CURRIC_ID] == cid]
                if not crows:
                    continue

                curric_done    = crows[0][COL_CURRIC_CMP] == 'Yes'
                item_rows      = [r for r in crows if r[COL_ITEM_TTL]]

                items = []
                for ir in item_rows:
                    title = str(ir[COL_ITEM_TTL] or '')
                    if 'coming soon' in title.lower():
                        continue
                    items.append({
                        'title': title,
                        'type':  str(ir[COL_ITEM_TYPE] or 'ONLINE'),
                        'done':  bool(ir[COL_ITEM_STS]),
                        'date':  ir[COL_ITEM_DATE].strftime('%Y-%m-%d') if ir[COL_ITEM_DATE] else None,
                        'req':   _date(ir[COL_ITEM_REQ]),
                    })

                total = len(items)
                done  = sum(1 for i in items if i['done'])
                pct   = round(done / total * 100) if total else 0

                curricula_data[cid] = {
                    'title':    cname,
                    'complete': curric_done,
                    'pct':      pct,
                    'done':     done,
                    'total':    total,
                    'items':    items,
                }

            all_done    = sum(v['done']  for v in curricula_data.values())
            all_total   = sum(v['total'] for v in curricula_data.values())
            overall_pct = round(all_done / all_total * 100) if all_total else 0

            seen[email] = {
                'email':       email,
                'key':         pkey(first, last),
                'name':        f"{first} {last}",
                'first':       first,
                'last':        last,
                'jobTitle':    str(prows[0][COL_JOBTITLE]  or ''),
                'market':      str(prows[0][COL_MARKET]    or ''),
                'manager':     f"{prows[0][COL_MGR_FIRST] or ''} {prows[0][COL_MGR_LAST] or ''}".strip(),
                'mgrEmail':    str(prows[0][COL_MGR_EMAIL] or ''),
                'mgrTitle':    str(prows[0][COL_MGR_TITLE] or ''),
                'hireDate':    prows[0][COL_HIRE_DATE].strftime('%Y-%m-%d') if prows[0][COL_HIRE_DATE] else None,
                'assignDate':  assign_date.strftime('%Y-%m-%d') if assign_date else None,
                'daysRem':     days_rem_lms,
                'overallDone': overall_done,
                'overallPct':  overall_pct,
                'curricula':   curricula_data,
                'playbook':    [],
            }

    return seen


PB_EMAIL = 3  # playbook traffic email column

def attach_playbook(records):
    folder = 'data'
    if not os.path.exists(folder):
        return records
    files = sorted([f for f in os.listdir(folder) if f.endswith('.xlsx')], key=extract_date)

    for fname in files:
        wb = openpyxl.load_workbook(f"{folder}/{fname}", read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))[1:]

        for r in rows:
            url = str(r[PB_URL] or '')
            if 'accelerate' not in url.lower():
                continue
            pb_email = str(r[PB_EMAIL] or '').strip().lower()
            if pb_email and pb_email in records:
                dt = r[PB_DATE]
                records[pb_email]['playbook'].append({
                    'page': page_label(url),
                    'url':  url,
                    'date': dt.strftime('%Y-%m-%d') if dt else '',
                })

    return records


def _parse_html_xls(path):
    from html.parser import HTMLParser
    class _TP(HTMLParser):
        def __init__(self):
            super().__init__()
            self.rows, self._row, self._cell, self._in = [], [], '', False
        def handle_starttag(self, tag, attrs):
            if tag in ('td', 'th'):
                self._in = True; self._cell = ''
            elif tag == 'tr':
                self._row = []
        def handle_endtag(self, tag):
            if tag in ('td', 'th'):
                self._row.append(self._cell.strip()); self._in = False
            elif tag == 'tr':
                if self._row: self.rows.append(self._row)
        def handle_data(self, data):
            if self._in: self._cell += data
    with open(path, encoding='utf-8', errors='ignore') as f:
        content = f.read()
    p = _TP()
    p.feed(content)
    if not p.rows:
        return []
    headers = p.rows[0]
    return [dict(zip(headers, row)) for row in p.rows[1:] if row]


def _load_salesforce(records):
    """Read Closed Won Salesforce export from leaderboard-data/, return first CW deal per cohort email."""
    import glob
    xls_files = glob.glob(os.path.join('leaderboard-data', '*.xls'))
    cw_path = None
    for p in xls_files:
        with open(p, encoding='utf-8', errors='ignore') as f:
            chunk = f.read(2000)
        if 'Opportunity Owner Email' in chunk:
            cw_path = p
            break
    if not cw_path:
        return {}

    rows = _parse_html_xls(cw_path)
    sales = {}
    for row in rows:
        if row.get('Stage', '').strip() != 'Closed Won':
            continue
        email = row.get('Opportunity Owner Email', '').strip().lower()
        if email not in records:
            continue
        try:
            amount = float(str(row.get('Amount', '0') or '0').replace(',', '').strip())
        except ValueError:
            amount = 0.0
        raw_date = str(row.get('Close Date', ''))
        m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', raw_date)
        close_date = f'{int(m.group(3)):04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}' if m else raw_date[:10]
        # Only count deals closed within the rep's first 45 days of the program
        assign_date = records[email].get('assignDate')
        days_to_close = None
        if assign_date and close_date:
            try:
                from datetime import date as _d
                atc = (_d.fromisoformat(close_date) - _d.fromisoformat(assign_date)).days
                if atc < 0 or atc > 45:
                    continue
                days_to_close = atc
            except Exception:
                continue
        account = row.get('Account Name', '').strip().title()
        # Keep the earliest closed deal (true "first sale")
        if email not in sales or close_date < sales[email]['closeDate']:
            sales[email] = {'amount': amount, 'accountName': account, 'closeDate': close_date, 'daysToClose': days_to_close}
    return sales


def generate_html(records, sales_map=None):
    from datetime import date
    people           = sorted(records.values(), key=lambda p: p['name'].lower())
    people_json      = json.dumps(people, default=str)
    sales_map        = sales_map or {}
    sales_map_json   = json.dumps(sales_map, ensure_ascii=False)
    cohort_total     = len(people)
    today_str        = date.today().isoformat()
    try:
        _lms_dir   = 'onboarding-data'
        _lms_files = [f for f in os.listdir(_lms_dir) if f.endswith('.xlsx')]
        _lms_files.sort(key=lambda f: os.path.getmtime(os.path.join(_lms_dir, f)))
        _lms_label = _file_date_label(os.path.join(_lms_dir, _lms_files[-1])) if _lms_files else date.today().strftime('%B %-d, %Y')
    except Exception:
        _lms_label = date.today().strftime('%B %-d, %Y')
    file_date         = _lms_label
    header_date_label = _lms_label
    total       = len(people)

    curric_ids   = json.dumps([c[0] for c in CURRICULA])
    curric_names = json.dumps({c[0]: c[1] for c in CURRICULA})

    # Page-level course map: exact LMS title → playbook page slug (confirmed 2026-07-09)
    _course_page = {
        'The KM Sales Experience':                                                          'understandingsalesforce',
        'Salesforce Live Workshop':                                                         'understandingsalesforce',
        'Mastering your Daily Sales Workflow':                                              'salesworkflow',
        'Why Konica Minolta: Selling the Value of Our Technology, Services and Solutions':  'coreportfolio',
        'bizhub One i-Series: Functionality, Ease of Use, Security':                        'coreportfolio',
        'Layered Security: Introducing the Model':                                          'coreportfolio',
        'User Authentication and Access Control':                                           'coreportfolio',
        'Document Security':                                                                'coreportfolio',
        'Introduction to Production Print':                                                 'coreportfolio',
        'Introducing BlueIrisIQ: Our Story':                                               'coreportfolio',
        'Managed IT (MIT) New Hire Kickstart':                                              'coreportfolio',
        'Identifying Real Sales Opportunities and Using the Lease Upgrade Sheet Part 1':    'prospectingskills',
        'Identifying Real Sales Opportunities and Using the Lease Upgrade Sheet Part 2':    'prospectingskills',
        'How to Focus on the Work That Matters Most: Highest Payoff Activities':            'prospectingskills',
        'Introduction to Microsoft Copilot':                                                'prospectingskills',
        'Microsoft Copilot: The Art of Prompting':                                          'prospectingskills',
        'How to Research and Write Messages That Win Meetings: Foundations of a Targeted Message':    'prospectingskills',
        'How to Research and Write Messages That Win Meetings: Leveraging the Persona Worksheet':     'prospectingskills',
        'Design Better Prospecting Calls with Disarm, Purpose, Question':                   'prospectingskills',
        'Build Better Voicemails: Structure, Clarity, Impact':                              'prospectingskills',
        'Prospecting Live Workshop':                                                        'prospectingskills',
        'Prospecting Foundations: Turning Skills into Action':                              'salesforceprospecting',
        'From Prep to Performance: Running Calls with Confidence':                          'callprep',
        'Discovery That Delivers: A Structured Approach to Better Sales Conversations':     'callprep',
        'Presentation Best Practices':                                                      'callprep',
        "Sales Math: Numbers Don't Lie":                                                    'workingwithnumbers',
        'Konica Minolta Premier Finance Leasing Live Workshop':                             'workingwithnumbers',
        'Managing and Moving Your Deals Forward':                                           'movingdeals',
        'How to Build Accurate Forecasts and Strong Pipelines':                             'pipelineownership',
    }
    _page_order = [
        'index', 'overview', 'welcome', 'understandingsalesforce',
        'salesworkflow',
        'coreportfolio',
        'prospectingskills', 'salesforceprospecting',
        'callprep', 'workingwithnumbers',
        'movingdeals', 'pipelineownership',
        'resources', 'managers',
    ]
    course_page_json = json.dumps(_course_page, ensure_ascii=False)
    page_order_json  = json.dumps(_page_order)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Accelerate Onboarding</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<style>
  :root {{
    --bg:#0d1117; --surface:#161b22; --surface2:#1e2530; --border:#30363d;
    --accent:#3b82f6; --accent2:#6d28d9; --accent3:#d97706;
    --text:#e6edf3; --muted:#8b949e; --green:#22c55e; --red:#ef4444;
    --amber:#f59e0b; --teal:#14b8a6;
    --green-subtle:#22c55e22; --red-subtle:#ef444422; --amber-subtle:#f59e0b22;
    --font:'Segoe UI',system-ui,sans-serif;
  }}
  body.light-mode {{
    --bg:#f4f6fb; --surface:#ffffff; --surface2:#eef1f7; --border:#d0d7e8;
    --accent:#2563eb; --text:#1a1d27; --muted:#475569; --green:#059669;
    --red:#dc2626; --amber:#d97706; --teal:#0f766e;
    --green-subtle:#05966922; --red-subtle:#dc262622; --amber-subtle:#d9770622;
  }}
  body.light-mode select,body.light-mode input{{color-scheme:light;}}
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{background:var(--bg);color:var(--text);font-family:var(--font);min-height:100vh;transition:background .2s,color .2s;}}

  /* ── Header ── */
  .header{{padding:20px 28px 16px;border-bottom:1px solid var(--border);display:grid;grid-template-columns:1fr auto 1fr;align-items:center;gap:12px;}}
  .header-left{{display:flex;align-items:center;gap:16px;}}
  .header-center{{display:flex;justify-content:center;align-items:center;}}
  .header-right{{display:flex;justify-content:flex-end;align-items:center;gap:8px;}}
  .kma-logo{{height:38px;width:auto;display:block;}}
  .kma-logo-light{{display:none;}}
  .light-mode .kma-logo-dark{{display:none;}}
  .light-mode .kma-logo-light{{display:block;}}
  .header h1{{font-size:18px;font-weight:700;letter-spacing:.3px;}}
  .header h1 span{{color:var(--muted);font-weight:400;}}
  .header-date{{font-size:11px;color:var(--muted);margin-top:2px;}}
  .data-badge{{font-size:11px;color:var(--muted);background:var(--surface2);border:1px solid var(--border);border-radius:20px;padding:3px 10px;white-space:nowrap;}}

  /* ── Buttons ── */
  .btn-theme{{background:transparent;border:1px solid var(--border);color:var(--muted);border-radius:6px;padding:5px 12px;font-size:12px;cursor:pointer;transition:all .15s;}}
  .btn-theme:hover{{border-color:var(--accent);color:var(--text);}}
  .btn-export{{background:var(--accent);border:1px solid var(--accent);color:#fff;border-radius:6px;padding:5px 12px;font-size:12px;cursor:pointer;font-weight:600;}}
  .btn-export:hover{{opacity:.88;}}
  .export-drop{{position:relative;}}
  .export-menu{{position:absolute;top:calc(100% + 6px);right:0;background:var(--surface);border:1px solid var(--border);border-radius:8px;min-width:210px;box-shadow:0 4px 24px rgba(0,0,0,.28);display:none;z-index:200;overflow:visible;}}
  .export-menu.open{{display:block;}}
  .export-item{{display:block;width:100%;text-align:left;padding:10px 14px;font-size:13px;color:var(--text);background:transparent;border:none;cursor:pointer;transition:background .1s;font-family:inherit;}}
  .export-item:hover{{background:var(--surface2);}}
  .export-parent{{position:relative;display:flex;justify-content:space-between;align-items:center;padding:10px 14px;font-size:13px;color:var(--text);cursor:default;transition:background .1s;}}
  .export-parent:hover{{background:var(--surface2);}}
  .export-chevron{{font-size:11px;color:var(--muted);margin-left:10px;}}
  .export-submenu{{position:absolute;right:100%;top:0;background:var(--surface);border:1px solid var(--border);border-radius:8px;min-width:90px;box-shadow:0 4px 24px rgba(0,0,0,0.28);display:none;z-index:201;overflow:hidden;margin-right:4px;}}
  .export-parent:hover .export-submenu{{display:block;}}

  /* ── Info tooltip ── */
  .info-btn{{display:inline-flex;align-items:center;justify-content:center;width:15px;height:15px;border-radius:50%;background:var(--surface2);border:1px solid var(--border);color:var(--muted);font-size:9px;font-weight:700;cursor:pointer;margin-left:5px;vertical-align:middle;flex-shrink:0;line-height:1;transition:border-color .15s,color .15s;}}
  .info-btn:hover{{border-color:var(--accent);color:var(--accent);}}
  .info-popover{{position:fixed;z-index:9999;background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:12px 14px;font-size:12px;color:var(--text);line-height:1.6;max-width:260px;box-shadow:0 4px 24px rgba(0,0,0,.5);display:none;}}
  .info-popover.visible{{display:block;}}

  /* ── Filters ── */
  .filters{{padding:14px 28px;display:flex;gap:10px;flex-wrap:wrap;align-items:center;border-bottom:1px solid var(--border);background:var(--surface);}}
  .filter-label{{font-size:12px;color:var(--muted);text-transform:uppercase;letter-spacing:.6px;margin-right:4px;}}
  select,input[type=text]{{background:var(--surface2);border:1px solid var(--border);color:var(--text);border-radius:6px;padding:6px 10px;font-size:13px;cursor:pointer;outline:none;color-scheme:dark;}}
  select:focus,input:focus{{border-color:var(--accent);}}
  .btn-reset{{background:transparent;border:1px solid var(--border);color:var(--muted);border-radius:6px;padding:6px 14px;font-size:12px;cursor:pointer;transition:border-color .15s,color .15s;}}
  .btn-reset:hover{{border-color:var(--accent);color:var(--text);}}
  .btn-group{{display:inline-flex;border:1px solid var(--border);border-radius:6px;overflow:hidden;}}
  .btn-group button{{background:transparent;border:none;border-right:1px solid var(--border);color:var(--muted);padding:6px 12px;font-size:12px;cursor:pointer;transition:background .15s,color .15s;white-space:nowrap;}}
  .btn-group button:last-child{{border-right:none;}}
  .btn-group button:hover{{background:var(--surface2);color:var(--text);}}
  .btn-group button.active{{background:var(--accent);color:#fff;font-weight:600;}}

  .sort-btn{{background:transparent;border:1px solid var(--border);color:var(--muted);border-radius:6px;padding:4px 10px;font-size:11px;cursor:pointer;transition:all .15s;white-space:nowrap;}}
  .sort-btn:hover{{border-color:var(--accent);color:var(--text);}}
  .sort-btn.active{{border-color:var(--accent);color:var(--accent);background:var(--accent)11;}}
  .mgr-group-hdr-row td{{padding:0;}}
  .mgr-group-hdr{{display:flex;align-items:center;justify-content:space-between;padding:8px 14px;cursor:pointer;background:var(--surface2);border-bottom:1px solid var(--border);user-select:none;}}
  .mgr-group-hdr:hover{{background:var(--border);}}
  .mgr-group-hdr.open .mgr-chevron{{transform:rotate(180deg);}}
  .mgr-chevron{{font-size:10px;color:var(--muted);transition:transform .15s;}}
  .result-count{{margin-left:auto;font-size:12px;color:var(--muted);}}

  /* ── Stats ── */
  .stats{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;padding:20px 28px;}}
  .stat{{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:16px 18px;}}
  .stat-label{{font-size:11px;text-transform:uppercase;letter-spacing:.7px;color:var(--muted);margin-bottom:6px;}}
  .stat-value{{font-size:28px;font-weight:700;line-height:1;}}
  .stat-value.green{{color:var(--green);}}
  .stat-value.red{{color:var(--red);}}
  .stat-value.amber{{color:var(--amber);}}
  .stat-value.blue{{color:var(--accent);}}
  .stat-sub{{font-size:11px;color:var(--muted);margin-top:4px;}}

  /* ── Section / Table ── */
  .section{{padding:0 28px 32px;}}
  .section-header{{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;flex-wrap:wrap;gap:8px;padding-top:20px;}}
  .section-title{{font-size:13px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;}}
  body.light-mode .section-title{{color:var(--text);}}
  .section-hint{{font-size:11px;color:var(--muted);}}
  .table-wrap{{overflow-x:auto;border:1px solid var(--border);border-radius:10px;background:var(--surface);}}
  table{{border-collapse:collapse;width:100%;min-width:680px;}}
  thead th{{font-size:10px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.05em;padding:10px 12px;text-align:left;border-bottom:1px solid var(--border);background:var(--surface2);white-space:nowrap;}}
  thead th.curric-col{{text-align:center;min-width:96px;white-space:normal;line-height:1.3;font-size:10px;}}
  thead th.overall-col{{text-align:center;min-width:80px;}}
  thead th.sortable-th{{cursor:pointer;user-select:none;}}
  thead th.sortable-th:hover{{color:var(--text);}}
  thead th.sort-active{{color:var(--accent) !important;}}
  tbody tr{{border-bottom:1px solid var(--border);cursor:pointer;transition:background .1s;}}
  tbody tr:last-child{{border-bottom:none;}}
  tbody tr:hover td{{background:var(--surface2) !important;}}
  tbody tr.selected td{{background:#3b82f611 !important;}}
  td{{padding:8px 12px;vertical-align:middle;font-size:12px;}}
  td.name-cell{{font-weight:500;white-space:nowrap;display:flex;align-items:center;gap:8px;}}
  td.market-cell{{font-size:11px;color:var(--muted);white-space:nowrap;}}
  td.pct-cell{{text-align:center;padding:5px 4px;}}
  tfoot tr{{border-top:2px solid var(--border);background:var(--surface2);}}
  tfoot td{{padding:7px 12px;font-size:11px;font-weight:600;color:var(--muted);}}
  tfoot td.pct-cell{{text-align:center;}}

  /* ── Pills & Badges ── */
  .pct-pill{{display:inline-block;border-radius:5px;padding:3px 9px;font-size:11px;font-weight:600;min-width:44px;text-align:center;line-height:1.4;}}
  .status-badge{{display:inline-block;font-size:10px;font-weight:700;padding:2px 9px;border-radius:20px;white-space:nowrap;}}
  .sb-completed{{background:var(--green-subtle);color:var(--green);}}
  .sb-ontrack{{background:#3b82f611;color:var(--accent);}}
  .sb-overdue{{background:var(--red-subtle);color:var(--red);}}
  .days-badge{{display:inline-block;font-size:10px;font-weight:600;padding:2px 7px;border-radius:4px;}}

  /* ── Modal overlay ── */
  .modal-overlay{{position:fixed;inset:0;background:rgba(0,0,0,.6);z-index:500;display:none;align-items:flex-start;justify-content:center;padding:40px 16px;overflow-y:auto;}}
  .modal-overlay.open{{display:flex;}}
  .modal-card{{background:var(--surface);border:1px solid var(--border);border-radius:12px;width:100%;max-width:760px;position:relative;}}
  .modal-close{{position:absolute;top:14px;right:16px;background:transparent;border:none;color:var(--muted);font-size:18px;cursor:pointer;line-height:1;padding:4px 8px;border-radius:4px;}}
  .modal-close:hover{{color:var(--text);background:var(--surface2);}}
  .modal-header{{padding:20px 24px 16px;border-bottom:1px solid var(--border);}}
  .modal-name{{font-size:18px;font-weight:700;margin-bottom:4px;}}
  .modal-meta{{font-size:12px;color:var(--muted);display:flex;flex-wrap:wrap;gap:12px;}}
  .modal-body{{padding:20px 24px;}}

  /* ── 45-day progress ── */
  .progress-section{{margin-bottom:20px;}}
  .progress-label{{font-size:12px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px;display:flex;align-items:center;justify-content:space-between;}}
  .progress-bar-wrap{{background:var(--surface2);border-radius:6px;height:10px;overflow:hidden;}}
  .progress-bar-fill{{height:10px;border-radius:6px;transition:width .3s;}}
  .progress-sublabel{{font-size:11px;color:var(--muted);margin-top:5px;}}

  /* ── Curriculum sections ── */
  .curric-section{{margin-bottom:16px;border:1px solid var(--border);border-radius:8px;overflow:hidden;}}
  .curric-header{{display:flex;align-items:center;gap:10px;padding:10px 14px;background:var(--surface2);cursor:pointer;user-select:none;}}
  .curric-header:hover{{background:var(--border);}}
  .curric-title{{flex:1;font-size:13px;font-weight:600;}}
  .curric-chevron{{font-size:10px;color:var(--muted);transition:transform .2s;}}
  .curric-header.open .curric-chevron{{transform:rotate(180deg);}}
  .curric-items{{display:none;padding:8px 0;}}
  .curric-items.open{{display:block;}}
  .item-row{{display:flex;align-items:center;gap:10px;padding:7px 14px;border-bottom:1px solid var(--border);}}
  .item-row:last-child{{border-bottom:none;}}
  .item-check{{width:16px;height:16px;border-radius:3px;display:flex;align-items:center;justify-content:center;font-size:10px;flex-shrink:0;}}
  .item-check.done{{background:var(--green-subtle);color:var(--green);border:1px solid var(--green);}}
  .item-check.not{{background:var(--surface2);color:var(--muted);border:1px solid var(--border);}}
  .item-title{{flex:1;font-size:12px;}}
  .item-type{{font-size:9px;font-weight:700;padding:1px 6px;border-radius:3px;background:var(--surface2);color:var(--muted);text-transform:uppercase;}}
  .item-date{{font-size:10px;color:var(--muted);white-space:nowrap;}}

  /* ── Playbook section ── */
  .playbook-section{{margin-top:20px;}}
  .playbook-header{{font-size:12px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:10px;padding-bottom:8px;border-bottom:1px solid var(--border);}}
  .playbook-empty{{font-size:12px;color:var(--muted);padding:12px 0;}}
  .pb-visit-row{{display:flex;align-items:center;gap:10px;padding:6px 0;border-bottom:1px solid var(--border);}}
  .pb-visit-row:last-child{{border-bottom:none;}}
  .pb-page{{flex:1;font-size:12px;}}
  .pb-date{{font-size:11px;color:var(--muted);white-space:nowrap;}}

  /* ── Legend ── */
  .legend{{display:flex;gap:14px;flex-wrap:wrap;align-items:center;margin-bottom:10px;font-size:11px;color:var(--muted);}}
  .leg{{display:flex;align-items:center;gap:5px;}}
  .leg-dot{{width:10px;height:10px;border-radius:2px;}}

  /* ── Manager info ── */
  .mgr-block{{background:var(--surface2);border-radius:8px;padding:12px 14px;margin-bottom:16px;display:flex;flex-direction:column;gap:3px;}}
  .mgr-label{{font-size:10px;text-transform:uppercase;letter-spacing:.5px;color:var(--muted);margin-bottom:2px;}}
  .mgr-name{{font-size:13px;font-weight:600;}}
  .mgr-detail{{font-size:11px;color:var(--muted);}}

  /* ── Charts ── */
  .charts{{display:grid;grid-template-columns:1fr 1fr;gap:16px;padding:0 28px 16px;}}
  @media(max-width:900px){{.charts{{grid-template-columns:1fr;}}}}
  @media(max-width:480px){{.chart-wrap{{height:160px;}}}}
  .chart-card{{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:18px;}}
  .chart-title{{font-size:13px;font-weight:600;margin-bottom:14px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;}}
  body.light-mode .chart-title{{color:var(--text);}}
  .chart-wrap{{position:relative;height:220px;}}

  /* ── Print ── */
  @media print {{
    .header,.filters,.stats,.charts,.section,.print-hide,.modal-overlay{{display:none!important;}}
    #print-header,#print-roster-wrap{{display:block!important;}}
  }}
  #print-header,#print-roster-wrap{{display:none;}}
  #print-header{{padding:16px 0;margin-bottom:12px;border-bottom:2px solid #333;}}
  .ph-title{{font-size:18px;font-weight:700;}}
  .ph-sub{{font-size:12px;color:#555;margin-top:4px;}}
  #print-roster-wrap table{{width:100%;border-collapse:collapse;font-size:11px;}}
  #print-roster-wrap th{{background:#eee;padding:6px 8px;text-align:left;border:1px solid #ccc;font-size:10px;}}
  #print-roster-wrap td{{padding:5px 8px;border:1px solid #ddd;}}
</style>
</head>
<body>

<!-- Info popover -->
<div class="info-popover" id="info-popover"></div>

<!-- Header -->
<div class="header">
  <div class="header-left">
    <div>
      <h1>Accelerate Onboarding</h1>
      <div class="header-date">Data through {header_date_label}</div>
    </div>
    <span class="data-badge" id="data-badge">{total} learners &middot; updated {file_date}</span>
  </div>
  <div class="header-center">
    <img src="https://jasonackerman1.github.io/playbook-dashboard/KMA-wht.svg" class="kma-logo kma-logo-dark" alt="KM Academy">
    <img src="https://jasonackerman1.github.io/playbook-dashboard/KMA-drk.svg" class="kma-logo kma-logo-light" alt="KM Academy">
  </div>
  <div class="header-right">
    <div class="export-drop print-hide" id="export-drop">
      <button class="btn-export" onclick="toggleExportDrop()">&#128438; Export &#9660;</button>
      <div class="export-menu" id="export-menu">
        <div class="export-parent">Full Report<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('full')">PDF</button><button class="export-item" onclick="runExportXLSX('full')">Excel</button></div></div>
        <div class="export-parent">Overdue Only<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('overdue')">PDF</button><button class="export-item" onclick="runExportXLSX('overdue')">Excel</button></div></div>
        <div class="export-parent">Manager Summary<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('manager-summary')">PDF</button><button class="export-item" onclick="runExportXLSX('manager-summary')">Excel</button></div></div>
      </div>
    </div><span class="info-btn print-hide" onclick="showInfo(event,'export')">?</span>
    <button class="btn-theme" id="btn-theme" onclick="toggleTheme()">&#9728; Light</button>
  </div>
</div>

<!-- Filters -->
<div class="filters">
  <span class="filter-label">Market</span>
  <select id="f-market" onchange="applyFilters()"><option value="">All Markets</option></select>
  <span class="filter-label">Status</span>
  <select id="f-status" onchange="applyFilters()">
    <option value="">All</option>
    <option value="Completed">Completed</option>
    <option value="On Track">On Track</option>
    <option value="Overdue">Overdue</option>
  </select>
  <span class="filter-label">Sort</span>
  <select id="f-sort" onchange="applyFilters()">
    <option value="name">Name A→Z</option>
    <option value="pct-desc">Completion High→Low</option>
    <option value="pct-asc">Completion Low→High</option>
    <option value="days-asc">Most Urgent First</option>
  </select>
  <span class="filter-label">Group</span>
  <div class="btn-group" id="test-group-btns">
    <button class="active" onclick="setTestGroup('all',this)">All</button>
    <button onclick="setTestGroup('hide',this)">Hide Test Group</button>
  </div>
  <button class="btn-reset" onclick="resetFilters()">Reset</button>
  <span class="info-btn" onclick="showInfo(event,'filters-info')" style="margin-left:4px;">?</span>
  <span class="result-count" id="result-count"></span>
</div>

<!-- Stat cards -->
<div class="stats">
  <div class="stat">
    <div class="stat-label">Total Enrolled <span class="info-btn" onclick="showInfo(event,'total-enrolled')">?</span></div>
    <div class="stat-value blue" id="s-total">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">Overdue <span class="info-btn" onclick="showInfo(event,'overdue')">?</span></div>
    <div class="stat-value red" id="s-overdue">&#8212;</div>
    <div class="stat-sub" id="s-overdue-sub"></div>
  </div>
  <div class="stat">
    <div class="stat-label">On Track <span class="info-btn" onclick="showInfo(event,'ontrack')">?</span></div>
    <div class="stat-value" id="s-ontrack">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">Completed <span class="info-btn" onclick="showInfo(event,'completed')">?</span></div>
    <div class="stat-value green" id="s-completed">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">Days to First Sale <span class="info-btn" onclick="showInfo(event,'first-sale')">?</span></div>
    <div class="stat-value amber" id="s-sales">TBD</div>
    <div class="stat-sub">avg from program start to close</div>
  </div>
</div>

<!-- Charts -->
<div class="charts">
  <div class="chart-card">
    <div class="chart-title">Completion by Market <span class="info-btn" onclick="showInfo(event,'market-chart')">?</span></div>
    <div class="chart-wrap"><canvas id="marketChart"></canvas></div>
  </div>
  <div class="chart-card">
    <div class="chart-title">Curriculum Progress <span class="info-btn" onclick="showInfo(event,'curric-chart')">?</span></div>
    <div class="chart-wrap"><canvas id="curricChart"></canvas></div>
  </div>
</div>

<!-- Heatmap section -->
<div class="section">
  <div class="section-header">
    <div>
      <div class="section-title">Progress Report <span class="info-btn" onclick="showInfo(event,'heatmap')">?</span></div>
      <div class="section-hint">Click any row to see full detail &mdash; curriculum breakdown, course checklist &amp; playbook activity</div>
    </div>
    <div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;">
      <div style="display:flex;gap:6px;align-items:center;">
        <button class="sort-btn active" id="view-individual" onclick="setTableView('individual')">Individual</button>
        <button class="sort-btn" id="view-manager" onclick="setTableView('manager')">By Manager</button>
        <span class="info-btn" onclick="showInfo(event,'view-toggle')" style="margin-left:2px;">?</span>
      </div>
      <input type="text" id="table-search" oninput="filterTableRows()" placeholder="Search name..." style="font-size:12px;padding:4px 10px;width:180px;background:var(--surface2);border:1px solid var(--border);color:var(--text);border-radius:6px;outline:none;">
    </div>
  </div>
  <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-bottom:12px;padding:10px 16px;background:var(--surface);border:1px solid var(--border);border-radius:8px;font-size:11px;color:var(--muted);">
    <div>
      <div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-bottom:5px;">Playbook &#9679;</div>
      <div style="display:flex;flex-wrap:wrap;gap:8px;">
        <span class="leg"><span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:#22c55e;margin-right:4px;flex-shrink:0;"></span>Using playbook</span>
        <span class="leg"><span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:#b91c1c;margin-right:4px;flex-shrink:0;"></span>Completing without playbook</span>
        <span class="leg"><span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:#6b7280;margin-right:4px;flex-shrink:0;"></span>No activity yet</span>
      </div>
    </div>
    <div>
      <div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-bottom:5px;">Gap</div>
      <div style="display:flex;flex-wrap:wrap;gap:8px;">
        <span class="leg"><span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:#15803d;margin-right:4px;flex-shrink:0;"></span>On pace or ahead (0% or less)</span>
        <span class="leg"><span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:#b45309;margin-right:4px;flex-shrink:0;"></span>Check-in (1&ndash;40%)</span>
        <span class="leg"><span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:#b91c1c;margin-right:4px;flex-shrink:0;"></span>Coaching needed (41%+)</span>
      </div>
    </div>
    <div>
      <div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-bottom:5px;">Curricula &#9632;</div>
      <div style="display:flex;flex-wrap:wrap;gap:8px;">
        <span class="leg"><span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:#15803d;margin-right:4px;flex-shrink:0;"></span>Done</span>
        <span class="leg"><span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:#1d4ed8;margin-right:4px;flex-shrink:0;"></span>In Progress</span>
        <span class="leg"><span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:#b91c1c;margin-right:4px;flex-shrink:0;"></span>Past Due</span>
        <span class="leg"><span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:#6b7280;margin-right:4px;flex-shrink:0;"></span>Not Started</span>
      </div>
    </div>
  </div>
  <div class="table-wrap">
    <table>
      <thead id="heatmap-head"></thead>
      <tbody id="heatmap-body"></tbody>
      <tfoot id="heatmap-foot"></tfoot>
    </table>
  </div>
</div>

<!-- User detail modal -->
<div class="modal-overlay" id="modal-overlay" onclick="closeModalBg(event)">
  <div class="modal-card" id="modal-card">
    <button class="modal-close" onclick="closeModal()">&#10005;</button>
    <div id="modal-content"></div>
  </div>
</div>

<!-- Print-only elements -->
<div id="print-header">
  <div class="ph-title">Accelerate Onboarding — <span id="ph-report-type">Full Report</span></div>
  <div class="ph-sub" id="ph-sub"></div>
</div>
<div id="print-roster-wrap"></div>

<script>
const PEOPLE = {people_json};
const SALES_MAP = {sales_map_json};
const COHORT_TOTAL = {cohort_total};
const COURSE_PAGE = {course_page_json};
const PAGE_ORDER = {page_order_json};
const CURRIC_IDS = {curric_ids};
const CURRIC_NAMES = {curric_names};
/* The designed pacing schedule, straight from each curriculum's LMS title
   (e.g. "Sales Skills (Weeks 2 & 3)"). [startDay, endDay] in program days,
   1-indexed, inclusive -- used to say which curriculum someone SHOULD be
   focused on right now, independent of both deadlines and raw % complete. */
const CURRIC_WEEK_RANGE = {{
  "ACCELERATE_GS": [1, 7],
  "ACCELERATE_SW": [1, 7],
  "ACCELERATE_CP": [1, 7],
  "ACCELERATE_P":  [1, 14],
  "ACCELERATE_SS": [8, 21],
  "ACCELERATE_PM": [22, 35],
}};
const PROGRAM_DAYS = {PROGRAM_DAYS};
const PLAYBOOK_CURRIC = {{
  'welcome':                 'ACCELERATE_GS',
  'understandingsalesforce': 'ACCELERATE_GS',
  'salesworkflow':           'ACCELERATE_SW',
  'coreportfolio':           'ACCELERATE_CP',
  'prospectingskills':       'ACCELERATE_P',
  'salesforceprospecting':   'ACCELERATE_P',
  'callprep':                'ACCELERATE_SS',
  'workingwithnumbers':      'ACCELERATE_SS',
  'movingdeals':             'ACCELERATE_PM',
  'pipelineownership':       'ACCELERATE_PM',
}};
const TODAY = new Date();
TODAY.setHours(0,0,0,0);

const TLG_SET = new Set([
  'jason ackerman','bianca davis','james parker','resmie biba',
  'chris curtis','sara thompson','jeremy macbean','bradley pierce',
  'laura sefcik','samantha maresca','staci musco','cj homer','rich moore','dale kinsey',
  'john lechner','resmie nesimi',"samantha d'angelo",'bianca dipasquale','doug falk'
]);

let hideTLG = true;
let testGroupFilter = 'all'; // 'all' | 'only' | 'hide'
const TEST_GROUP_DATE = '2026-06-04';

function setTestGroup(val, btn) {{
  testGroupFilter = val;
  document.querySelectorAll('#test-group-btns button').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  applyFilters();
}}
let filtered = [];
let marketChartObj = null;
let curricChartObj = null;
let tableSort = {{col: 'overall', dir: 'asc'}};
let tableView = 'individual';

/* ── Utilities ── */
function pct2color(p) {{
  if (p === 0) return {{bg:'#f1efe8', fg:'#888780'}};
  if (p < 40)  return {{bg:'#FCEBEB', fg:'#791F1F'}};
  if (p < 75)  return {{bg:'#FAEEDA', fg:'#633806'}};
  return {{bg:'#EAF3DE', fg:'#27500A'}};
}}

/* ── Deadline engine ──
   Each item carries its own LMS-assigned due date (item.req, or null if the
   item has no fixed deadline). A curriculum's days-remaining is the soonest
   due date among its own INCOMPLETE items. */
function itemDaysLeft(item) {{
  if (!item.req) return null;
  return Math.round((new Date(item.req) - TODAY) / 86400000);
}}
function curricDaysLeft(c) {{
  if (!c || c.complete) return null;
  const vals = c.items.filter(i => !i.done && i.req).map(itemDaysLeft);
  return vals.length ? Math.min(...vals) : null;
}}
function curricTrackStatus(c) {{
  if (!c) return 'Unknown';
  if (c.complete) return 'Completed';
  const dl = curricDaysLeft(c);
  if (dl === null) return 'On Track';
  return dl <= 0 ? 'Overdue' : 'On Track';
}}

function curricPillStyle(c, p) {{
  if (!c) return 'background:#6b7280;color:#fff';
  if (c.complete) return 'background:#15803d;color:#fff';
  const dl = curricDaysLeft(c);
  if (dl !== null && dl <= 0) return 'background:#b91c1c;color:#fff';
  if (c.pct > 0)  return 'background:#1d4ed8;color:#fff';
  return 'background:#6b7280;color:#fff';
}}
function curricDotBg(c, p) {{
  if (!c) return '#6b7280';
  if (c.complete) return '#15803d';
  const dl = curricDaysLeft(c);
  if (dl !== null && dl <= 0) return '#b91c1c';
  if (c.pct > 0) return '#1d4ed8';
  return '#6b7280';
}}

function overallPillStyle(p) {{
  if (p.overallDone) return 'background:#15803d;color:#fff;font-weight:700';
  if (computeStatus(p) === 'Overdue') return 'background:#b91c1c;color:#fff;font-weight:700';
  if (p.overallPct > 0) return 'background:#1d4ed8;color:#fff;font-weight:700';
  return 'background:#6b7280;color:#fff;font-weight:700';
}}

function computeStatus(p) {{
  if (p.overallDone) return 'Completed';
  const ids = Object.keys(p.curricula);
  if (!ids.length) return 'Unknown';
  const anyOverdue = ids.some(cid => curricTrackStatus(p.curricula[cid]) === 'Overdue');
  return anyOverdue ? 'Overdue' : 'On Track';
}}

function overdueCount(p) {{
  return Object.values(p.curricula).filter(c => curricTrackStatus(c) === 'Overdue').length;
}}

function soonestDaysLeft(p) {{
  const vals = Object.values(p.curricula).map(curricDaysLeft).filter(v => v !== null);
  if (vals.length) return Math.min(...vals);
  return computedDaysLeft(p);
}}

function daysLeft(p) {{
  return soonestDaysLeft(p);
}}

/* Kept for the Expected%/Gap pacing metric only (not for Overdue/Days Left,
   which now come from each curriculum's own LMS due dates above). */
function daysElapsed(p) {{
  if (!p.assignDate) return 0;
  const start = new Date(p.assignDate);
  return Math.max(0, Math.round((TODAY - start) / 86400000));
}}
function computedDaysLeft(p) {{
  if (!p.assignDate) return null;
  return PROGRAM_DAYS - daysElapsed(p);
}}
function expectedPct(p) {{
  return Math.min(100, Math.round(daysElapsed(p) / PROGRAM_DAYS * 100));
}}
function gapPct(p) {{
  return expectedPct(p) - p.overallPct;
}}
/* Where they should be, curriculum-by-curriculum, per the designed weekly
   schedule -- independent of both LMS due dates (Days Left) and raw %
   complete (Biggest Gap). */
function curricScheduleStatus(cid, d) {{
  const range = CURRIC_WEEK_RANGE[cid];
  if (!range) return 'active';
  if (d < range[0]) return 'upcoming';
  if (d > range[1]) return 'due';
  return 'active';
}}
function expectedFocus(p) {{
  const d = daysElapsed(p);
  const active = [];
  const behindSchedule = [];
  CURRIC_IDS.forEach(cid => {{
    const c = p.curricula[cid];
    if (!c || c.complete) return;
    const st = curricScheduleStatus(cid, d);
    if (st === 'active') active.push(cid);
    else if (st === 'due') behindSchedule.push(cid);
  }});
  if (behindSchedule.length) return {{mode: 'behind', ids: behindSchedule}};
  if (active.length) return {{mode: 'active', ids: active}};
  return {{mode: 'ahead', ids: []}};
}}
function gapStyle(g) {{
  if (g > 40) return 'background:#b91c1c;color:#fff;font-weight:700';
  if (g > 0)  return 'background:#b45309;color:#fff;font-weight:700';
  return 'background:#15803d;color:#fff;font-weight:700';
}}
function pctPillStyle(pct) {{
  if (pct === 0)  return 'background:#b91c1c;color:#fff;font-weight:600';
  if (pct < 50)   return 'background:#b45309;color:#fff;font-weight:600';
  return 'background:#6b7280;color:#fff;font-weight:600';
}}
function urlSlug(url) {{
  const m = (url || '').match(/\/([^\/]+?)(?:\.html)?(?:\?.*)?$/);
  if (!m || !m[1] || m[1] === 'accelerate_sales_playbook' || m[1] === '') return 'index';
  return m[1].toLowerCase();
}}
function pbEngagement(p) {{
  const visits = (p.playbook || []).filter(v => !p.assignDate || v.date >= p.assignDate);
  if (!visits.length) {{
    const level = p.overallPct > 0 ? 'alert' : 'none';
    return {{totalVisits:0, uniquePages:0, visitedCurricula:{{}}, mismatches:[], level}};
  }}
  const totalVisits = visits.length;
  const slugs = visits.map(v => urlSlug(v.url));
  const uniquePages = new Set(slugs).size;
  const visitedCurricula = {{}};
  slugs.forEach(slug => {{
    const cid = PLAYBOOK_CURRIC[slug];
    if (cid) visitedCurricula[cid] = (visitedCurricula[cid] || 0) + 1;
  }});
  const mismatches = CURRIC_IDS.filter(cid => {{
    const c = p.curricula[cid];
    return c && c.done > 0 && !visitedCurricula[cid];
  }});
  const level = mismatches.length > 0 ? 'partial' : 'active';
  return {{totalVisits, uniquePages, visitedCurricula, mismatches, level}};
}}
function pbSummaryHtml(p) {{
  const eng = pbEngagement(p);
  let html = '<div style="margin-bottom:16px;padding-bottom:16px;border-bottom:1px solid var(--border);">';
  html += '<div style="font-size:12px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px;">Playbook Engagement</div>';
  if (eng.totalVisits === 0) {{
    const msg = p.overallPct > 0
      ? '<span style="color:var(--red)">&#9888; Completing courses with no playbook visits recorded</span>'
      : '<span style="color:var(--muted)">No playbook visits recorded yet</span>';
    html += '<div style="font-size:13px;margin-bottom:8px;">' + msg + '</div>';
  }} else {{
    html += '<div style="font-size:13px;color:var(--muted);margin-bottom:8px;">' + eng.totalVisits + ' visit' + (eng.totalVisits!==1?'s':'') + ' &middot; ' + eng.uniquePages + ' unique page' + (eng.uniquePages!==1?'s':'') + '</div>';
  }}
  html += '<div style="display:flex;flex-wrap:wrap;gap:5px;margin-bottom:8px;">';
  CURRIC_IDS.forEach(cid => {{
    const c = p.curricula[cid];
    const hasDone = c && c.done > 0;
    const visited = !!eng.visitedCurricula[cid];
    let bg, color, icon;
    if (visited)       {{ bg='var(--green-subtle)'; color='var(--green)';  icon='&#10003;'; }}
    else if (hasDone)  {{ bg='var(--red-subtle)';   color='var(--red)';    icon='&#10007;'; }}
    else               {{ bg='var(--surface2)';      color='var(--muted)';  icon='&middot;'; }}
    html += '<span style="display:inline-flex;align-items:center;gap:4px;padding:3px 9px;border-radius:99px;background:'+bg+';color:'+color+';font-size:11px;font-weight:600;">'+icon+' '+CURRIC_NAMES[cid]+'</span>';
  }});
  html += '</div>';
  if (eng.mismatches.length > 0) {{
    html += '<div style="background:var(--red-subtle);color:var(--red);border-radius:6px;padding:8px 12px;font-size:12px;">&#9888; Courses completed without visiting: ' + eng.mismatches.map(cid=>CURRIC_NAMES[cid]).join(', ') + '</div>';
  }}
  html += '</div>';
  return html;
}}

function fmtDate(d) {{
  if (!d) return '';
  const dt = new Date(d + 'T00:00:00');
  return dt.toLocaleDateString('en-US', {{month:'short', day:'numeric', year:'numeric'}});
}}

function cv() {{
  return getComputedStyle(document.body).getPropertyValue('--' + arguments[0]).trim();
}}

/* ── Theme ── */
function toggleTheme() {{
  const light = document.body.classList.toggle('light-mode');
  document.getElementById('btn-theme').innerHTML = light ? '&#9790; Dark' : '&#9728; Light';
  localStorage.setItem('pb-theme', light ? 'light' : 'dark');
  renderCharts();
}}
(function() {{
  const t = localStorage.getItem('pb-theme');
  if (t !== 'dark') {{
    document.body.classList.add('light-mode');
    document.getElementById('btn-theme').innerHTML = '&#9790; Dark';
  }}
}})();

/* ── Info popover ── */
const INFO = {{
  "total-enrolled": "The total number of people currently enrolled in the Accelerate Onboarding program. If filters are applied above, this reflects only the people currently shown.",
  "overdue": "People who have missed the LMS deadline for at least one required course. Deadlines are set per course relative to each person's program start date — for example, Getting Started items may be due within the first week. Click any row to see exactly which courses are overdue.",
  "ontrack": "People who have not missed any course deadlines yet. They may or may not be keeping pace — check the Gap column to see who is falling behind even if they are technically still on time. Click any row to see which courses still need attention.",
  "completed": "People who have finished every required course across all six curricula in the Accelerate program.",
  "first-sale": "Average number of days from each rep's Accelerate program start date to their first Closed Won deal. Only deals closed within the first 45 days of the program qualify. Open pipeline is not included. Data pulled from Salesforce.",
  "filters-info": "Market: narrow the table to one market. Status: show only people who are Completed, On Track, or Overdue. Sort: change the default row order. All filters work together — for example, filter to a market and sort by Most Urgent to quickly find who needs attention there. Use Reset to clear everything.",
  "view-toggle": "Individual view shows one row per learner. By Manager groups learners under their manager so you can see how each team is tracking. The search box updates automatically — in Individual view it searches by learner name, in By Manager view it searches by manager name.",
  "market-chart": "Average overall completion percentage by market. Hover over a bar to see how many reps in that market are done, on track, or overdue. A short bar is a signal that market may need extra support.",
  "curric-chart": "Average completion percentage per curriculum across the whole cohort. Hover over a bar to see how many reps have finished, are in progress, have not started, or are past the deadline for that curriculum. A short bar is where reps are getting stuck.",
  "heatmap": "One row per learner. Click any row to open their full detail — every individual lesson, completion dates, curriculum breakdown, and playbook activity timeline.",
  "col-learner": "The colored dot to the left of the name shows whether this person is using the Accelerate Playbook alongside their LMS courses. Green = visiting the playbook. Red = completing courses with no playbook visits recorded. Gray = no activity yet. Click any row to see their full playbook and course history.",
  "col-status": "Each person's current standing in the program. Completed = finished all required courses. Overdue = past the LMS deadline for at least one course. On Track = within all deadlines so far. Note: On Track does not mean ahead of pace — check the Gap column to see if someone is falling behind.",
  "col-days": "Days until the nearest upcoming course deadline, based on LMS-assigned due dates. Negative numbers mean a deadline has already passed — the number shows how many days ago. Each curriculum has its own deadline schedule tied to the person's program start date.",
  "col-actual": "Weighted overall completion percentage. Calculated as total lessons completed divided by total lessons assigned across all six curricula. Each lesson counts equally regardless of which curriculum it belongs to.",
  "col-expected": "How far along this person should be today based on the program's designed pacing schedule. Accounts for how many days they have had the program and what the schedule calls for at this point. Compare to Actual % — if Actual is lower, they are behind pace.",
  "col-gap": "The difference between Actual % and Expected %. A negative number means behind schedule by that amount. A positive number means ahead. The color scales with severity: green when close, red when the gap is significant enough to warrant a check-in.",
  "expected-focus-info": "Which curriculum this person should be actively working on right now based on the program's pacing schedule — Getting Started, Sales Workflow, and Core Portfolio in Week 1; Prospecting and Sales Skills in Weeks 2–3; Pipeline Management in Weeks 4–5. If a curriculum is shown in red, it should already be complete but is not.",
  "biggest-gap-info": "The curriculum this person has fallen furthest behind on, measured by completion percentage. This is not the same as what is overdue — a curriculum can have a big gap without having passed its deadline yet. Use this to identify where to focus a coaching conversation.",
  "col-curricula": "Six colored squares, one per curriculum in program order: Getting Started, Sales Workflow, Core Portfolio, Prospecting, Sales Skills, Pipeline Management. Green = complete. Blue = in progress. Red = past its deadline and not done. Gray = not started yet. Click the row to see individual lesson detail.",
  "col-days-to-close": "How many days it took this rep to close their first Salesforce deal after starting the Accelerate program. Only deals closed within the first 45 days of the program qualify. A dash means no qualifying deal has been recorded yet.",
  "export": "Downloads a report based on whoever is currently visible — apply filters first, then export. Full Report: all visible learners with status and progress. Overdue Only: people past a course deadline with their manager's contact info for easy follow-up. Manager Summary: one row per manager with their team's headcount and progress. Example: filter to a market, then export Overdue Only to get a targeted outreach list.",
}};
function showInfo(e, key) {{
  e.stopPropagation();
  const pop = document.getElementById('info-popover');
  if (pop.dataset.key === key && pop.classList.contains('visible')) {{
    pop.classList.remove('visible'); pop.dataset.key = ''; return;
  }}
  pop.dataset.key = key;
  pop.textContent = INFO[key] || '';
  pop.classList.add('visible');
  const r = e.target.getBoundingClientRect();
  pop.style.left = Math.min(r.left, window.innerWidth - 280) + 'px';
  pop.style.top  = (r.bottom + 6) + 'px';
}}

/* ── Export dropdown ── */
function toggleExportDrop() {{
  document.getElementById('export-menu').classList.toggle('open');
}}

/* ── Table search ── */
function filterTableRows() {{
  const q = (document.getElementById('table-search').value || '').toLowerCase();
  if (tableView === 'manager') {{
    let showGroup = true;
    document.querySelectorAll('#heatmap-body tr').forEach(row => {{
      if (row.classList.contains('mgr-group-hdr-row')) {{
        showGroup = !q || (row.dataset.manager || '').includes(q);
        row.style.display = showGroup ? '' : 'none';
      }} else {{
        row.style.display = showGroup ? '' : 'none';
      }}
    }});
  }} else {{
    document.querySelectorAll('#heatmap-body tr').forEach(row => {{
      row.style.display = (!q || (row.dataset.name || '').includes(q)) ? '' : 'none';
    }});
  }}
}}

function setTableView(v) {{
  tableView = v;
  document.getElementById('view-individual').classList.toggle('active', v === 'individual');
  document.getElementById('view-manager').classList.toggle('active', v === 'manager');
  const searchEl = document.getElementById('table-search');
  searchEl.placeholder = v === 'manager' ? 'Search manager...' : 'Search name...';
  searchEl.value = '';
  renderTable();
}}

function toggleMgrGroup(el) {{
  el.classList.toggle('open');
  const isOpen = el.classList.contains('open');
  let row = el.closest('tr').nextElementSibling;
  while (row && !row.classList.contains('mgr-group-hdr-row')) {{
    row.style.display = isOpen ? '' : 'none';
    row = row.nextElementSibling;
  }}
}}

/* ── Column sort ── */
function sortByCol(col) {{
  if (tableSort.col === col) {{
    tableSort.dir = tableSort.dir === 'asc' ? 'desc' : 'asc';
  }} else {{
    tableSort.col = col;
    tableSort.dir = 'asc';
  }}
  renderTable();
}}


/* ── Filters ── */
function resetFilters() {{
  document.getElementById('f-market').value = '';
  document.getElementById('f-status').value = '';
  document.getElementById('f-sort').value = 'name';
  document.getElementById('table-search').value = '';
  testGroupFilter = 'all';
  document.querySelectorAll('#test-group-btns button').forEach((b,i) => b.classList.toggle('active', i===0));
  filterTableRows();
  applyFilters();
}}

function applyFilters() {{
  const mkt    = document.getElementById('f-market').value;
  const status = document.getElementById('f-status').value;
  const sort   = document.getElementById('f-sort').value;

  filtered = PEOPLE.filter(p => {{
    if (hideTLG && TLG_SET.has(p.name.toLowerCase())) return false;
    if (testGroupFilter === 'hide' && p.assignDate === TEST_GROUP_DATE) return false;
    if (mkt && p.market !== mkt) return false;
    const st = computeStatus(p);
    if (status && st !== status) return false;
    return true;
  }});

  if (sort === 'pct-desc') {{ filtered.sort((a,b) => b.overallPct - a.overallPct); tableSort = {{col:'overall', dir:'desc'}}; }}
  else if (sort === 'pct-asc') {{ filtered.sort((a,b) => a.overallPct - b.overallPct); tableSort = {{col:'overall', dir:'asc'}}; }}
  else if (sort === 'days-asc') {{ filtered.sort((a,b) => {{ const da = daysLeft(a) ?? 999, db = daysLeft(b) ?? 999; return da - db; }}); tableSort = {{col:'days', dir:'asc'}}; }}
  else {{ filtered.sort((a,b) => a.name.localeCompare(b.name)); tableSort = {{col:'name', dir:'asc'}}; }}

  renderStats();
  renderTable();
  renderCharts();
  document.getElementById('result-count').textContent = filtered.length + ' shown';
}}

/* ── Stats ── */
function renderStats() {{
  const total = filtered.length;
  const overdue = filtered.filter(p => computeStatus(p) === 'Overdue').length;
  const ontrack = filtered.filter(p => computeStatus(p) === 'On Track').length;
  const completed = filtered.filter(p => computeStatus(p) === 'Completed').length;

  document.getElementById('s-total').textContent = total;
  document.getElementById('s-overdue').textContent = overdue;
  document.getElementById('s-overdue-sub').textContent = overdue ? overdue + ' past a curriculum deadline' : '';
  document.getElementById('s-ontrack').textContent = ontrack;
  document.getElementById('s-completed').textContent = completed;
  const salesVals = Object.values(SALES_MAP).filter(v => v.daysToClose != null);
  const avgDays = salesVals.length ? Math.round(salesVals.reduce((s,v) => s + v.daysToClose, 0) / salesVals.length) : null;
  document.getElementById('s-sales').textContent = avgDays !== null ? avgDays + 'd' : '—';
}}

/* ── Heatmap table ── */
function renderTable() {{
  // Header
  const thead = document.getElementById('heatmap-head');
  function thS(label, col, extraCls, infoKey) {{
    const active = tableSort.col === col;
    const arrow = active ? (tableSort.dir === 'asc' ? ' ↑' : ' ↓') : '';
    const cls = ['sortable-th', extraCls, active ? 'sort-active' : ''].filter(Boolean).join(' ');
    const btn = infoKey ? '<span class="info-btn print-hide" onclick="event.stopPropagation();showInfo(event,\\'' + infoKey + '\\')">?</span>' : '';
    return '<th class="' + cls + '" data-col="' + col + '" onclick="sortByCol(this.dataset.col)">' + label + arrow + btn + '</th>';
  }}
  let hRow = '<tr>' + thS('Learner','name','','col-learner') + thS('Status','status','','col-status') + thS('Days Left','days','','col-days');
  hRow += thS('Actual %','overall','overall-col','col-actual');
  hRow += thS('Expected %','expected','overall-col','col-expected');
  hRow += thS('Gap','gap','overall-col','col-gap');
  hRow += '<th class="curric-col" style="font-size:10px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.05em;padding:10px 12px;text-align:center;background:var(--surface2);border-bottom:1px solid var(--border);white-space:normal;line-height:1.3;">Expected Focus<span class="info-btn print-hide" onclick="showInfo(event,\\'expected-focus-info\\')">?</span></th>';
  hRow += '<th class="curric-col" style="font-size:10px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.05em;padding:10px 12px;text-align:center;background:var(--surface2);border-bottom:1px solid var(--border);white-space:normal;line-height:1.3;">Biggest Gap<span class="info-btn print-hide" onclick="showInfo(event,\\'biggest-gap-info\\')">?</span></th>';
  hRow += '<th style="font-size:10px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.05em;padding:10px 12px;text-align:center;background:var(--surface2);border-bottom:1px solid var(--border);white-space:nowrap;">Curricula<span class="info-btn print-hide" onclick="showInfo(event,\\'col-curricula\\')">?</span></th>';
  hRow += thS('Days to Close','daysToClose','overall-col','col-days-to-close');
  hRow += '</tr>';
  thead.innerHTML = hRow;

  // Body
  const tbody = document.getElementById('heatmap-body');
  if (!filtered.length) {{
    tbody.innerHTML = '<tr><td colspan="10" style="text-align:center;color:var(--muted);padding:32px;">No learners match the current filters.</td></tr>';
    document.getElementById('heatmap-foot').innerHTML = '';
    return;
  }}

  let display = filtered.slice();
  if (tableSort.col) {{
    display.sort((a, b) => {{
      let va, vb;
      const col = tableSort.col;
      if (col === 'name')         {{ va = a.name; vb = b.name; }}
      else if (col === 'market')  {{ va = a.market; vb = b.market; }}
      else if (col === 'status')  {{ va = computeStatus(a); vb = computeStatus(b); }}
      else if (col === 'days')    {{ va = daysLeft(a) ?? 9999; vb = daysLeft(b) ?? 9999; }}
      else if (col === 'overall')   {{ va = a.overallPct; vb = b.overallPct; }}
      else if (col === 'expected')  {{ va = expectedPct(a); vb = expectedPct(b); }}
      else if (col === 'gap')       {{ va = gapPct(a); vb = gapPct(b); }}
      else if (col === 'daysToClose') {{ va = SALES_MAP[a.email] ? (SALES_MAP[a.email].daysToClose ?? 999) : 999; vb = SALES_MAP[b.email] ? (SALES_MAP[b.email].daysToClose ?? 999) : 999; }}
      else {{ va = a.curricula[col] ? a.curricula[col].pct : 0; vb = b.curricula[col] ? b.curricula[col].pct : 0; }}
      if (va < vb) return tableSort.dir === 'asc' ? -1 : 1;
      if (va > vb) return tableSort.dir === 'asc' ? 1 : -1;
      return 0;
    }});
  }}
  function worstCurric(p) {{
    let worst = null, worstDl = null;
    CURRIC_IDS.forEach(cid => {{
      const c = p.curricula[cid];
      const dl = curricDaysLeft(c);
      if (dl === null) return;
      if (worstDl === null || dl < worstDl) {{ worstDl = dl; worst = cid; }}
    }});
    return worst ? {{cid: worst, name: CURRIC_NAMES[worst], dl: worstDl}} : null;
  }}

  function expectedFocusCell(p) {{
    if (p.overallDone) return '<td class="pct-cell" style="text-align:center;color:var(--muted);">&mdash;</td>';
    const ef = expectedFocus(p);
    const names = ef.ids.map(id => CURRIC_NAMES[id]);
    const label = names.length <= 2 ? names.join(', ') : names[0] + ' +' + (names.length - 1) + ' more';
    if (ef.mode === 'behind') {{
      return '<td class="pct-cell" style="text-align:center;"><span class="pct-pill" style="background:#b91c1c;color:#fff;font-weight:600;" title="Should already be done: ' + escHtml(names.join(', ')) + '">Behind: ' + escHtml(label) + '</span></td>';
    }}
    if (ef.mode === 'active') {{
      return '<td class="pct-cell" style="text-align:center;"><span class="pct-pill" style="background:#1d4ed8;color:#fff;font-weight:600;" title="Should be working on: ' + escHtml(names.join(', ')) + '">' + escHtml(label) + '</span></td>';
    }}
    return '<td class="pct-cell" style="text-align:center;"><span class="pct-pill" style="background:#15803d;color:#fff;font-weight:600;">Ahead of schedule</span></td>';
  }}

  function biggestGapCurric(p) {{
    let worst = null, worstPct = null;
    CURRIC_IDS.forEach(cid => {{
      const c = p.curricula[cid];
      if (!c || c.complete) return;
      if (worstPct === null || c.pct < worstPct) {{ worstPct = c.pct; worst = cid; }}
    }});
    return worst ? {{cid: worst, name: CURRIC_NAMES[worst], pct: worstPct}} : null;
  }}

  function personRow(p) {{
    const status = computeStatus(p);
    const statusClass = status === 'Completed' ? 'sb-completed' : status === 'On Track' ? 'sb-ontrack' : 'sb-overdue';
    const od = overdueCount(p);
    const sdl = soonestDaysLeft(p);
    const w = worstCurric(p);
    const wName = w ? ' <span style="color:var(--muted);font-weight:400;">(' + escHtml(w.name) + ')</span>' : '';
    const daysStr = p.overallDone ? '&mdash;' :
      od > 0 ? '<span style="color:var(--red);font-weight:700">' + od + ' past due</span>' + wName :
      sdl === null ? '&mdash;' :
      '<span style="color:var(--green)">' + sdl + 'd left</span>' + wName;
    let dotsCell = '<td style="text-align:center;padding:5px 12px;"><div style="display:inline-flex;gap:3px;align-items:center;">';
    CURRIC_IDS.forEach(cid => {{
      const c = p.curricula[cid];
      const bg = curricDotBg(c, p);
      const cdl = curricDaysLeft(c);
      const statusTip = !c ? 'N/A' : c.complete ? 'Completed' : cdl === null ? 'On Track (no deadline yet)' :
        cdl <= 0 ? Math.abs(cdl) + 'd overdue' : cdl + 'd left';
      const lbl = CURRIC_NAMES[cid] + ': ' + (c ? c.pct + '%' : 'N/A') + ' · ' + statusTip;
      dotsCell += '<span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:' + bg + ';" title="' + lbl + '"></span>';
    }});
    dotsCell += '</div></td>';
    const eng = pbEngagement(p);
    const dotColor = (eng.level === 'alert') ? '#b91c1c' : (eng.level === 'none') ? '#6b7280' : '#22c55e';
    const dotTip  = (eng.level === 'alert') ? 'Completing courses -- no playbook visits recorded' : (eng.level === 'none') ? 'No activity yet' : 'Using the playbook';
    const bg = biggestGapCurric(p);
    const bgCell = bg ?
      '<td class="pct-cell" style="text-align:center;"><span class="pct-pill" style="' + pctPillStyle(bg.pct) + '">' + escHtml(bg.name) + ' &middot; ' + bg.pct + '%</span></td>' :
      '<td class="pct-cell" style="text-align:center;color:var(--muted);">&mdash;</td>';
    const pSaleRow = SALES_MAP[p.email];
    const saleCell = pSaleRow
      ? '<td class="pct-cell" style="font-weight:700;font-size:11px;color:var(--accent);white-space:nowrap;">' + (pSaleRow.daysToClose != null ? pSaleRow.daysToClose + 'd' : '—') + '</td>'
      : '<td class="pct-cell" style="text-align:center;color:var(--muted);opacity:.4;">&#8212;</td>';
    return '<tr data-email="' + escHtml(p.email) + '" data-name="' + escHtml(p.name.toLowerCase()) + '" onclick="openModal(this.dataset.email)" title="Click to see full detail">' +
      '<td class="name-cell"><span style="width:8px;height:8px;border-radius:50%;background:' + dotColor + ';display:inline-block;flex-shrink:0;" title="' + dotTip + '"></span>' + escHtml(p.name) + '</td>' +
      '<td><span class="status-badge ' + statusClass + '">' + status + '</span></td>' +
      '<td style="font-size:11px;">' + daysStr + '</td>' +
      '<td class="pct-cell" style="font-weight:600;font-size:12px;">' + p.overallPct + '%</td>' +
      '<td class="pct-cell" style="font-weight:600;font-size:12px;color:var(--muted);">' + expectedPct(p) + '%</td>' +
      (function(){{ const g=gapPct(p); return '<td class="pct-cell"><span class="pct-pill" style="' + gapStyle(g) + '">' + g + '%</span></td>'; }})() +
      expectedFocusCell(p) +
      bgCell +
      dotsCell +
      saleCell +
    '</tr>';
  }}

  if (tableView === 'manager') {{
    const groups = {{}};
    filtered.forEach(p => {{
      const mgr = p.manager || '(No Manager)';
      if (!groups[mgr]) groups[mgr] = [];
      groups[mgr].push(p);
    }});
    const colCount = 10;
    let html = '';
    Object.keys(groups).sort().forEach(mgr => {{
      const team = groups[mgr];
      const avg = Math.round(team.reduce((s, p) => s + p.overallPct, 0) / team.length);
      const od = team.filter(p => computeStatus(p) === 'Overdue').length;
      const odStr = od > 0 ? ' &nbsp;·&nbsp; <span style="color:var(--red);font-weight:600">' + od + ' overdue</span>' : '';
      html += '<tr class="mgr-group-hdr-row" data-manager="' + escHtml(mgr.toLowerCase()) + '"><td colspan="' + colCount + '">' +
        '<div class="mgr-group-hdr open" onclick="toggleMgrGroup(this)">' +
        '<span style="font-weight:700;font-size:13px;">' + escHtml(mgr) + '</span>' +
        '<span style="font-size:11px;color:var(--muted);">' + team.length + ' learner' + (team.length !== 1 ? 's' : '') + ' &nbsp;·&nbsp; avg ' + avg + '%' + odStr + ' &nbsp;<span class="mgr-chevron">&#9660;</span></span>' +
        '</div></td></tr>';
      team.forEach(p => {{ html += personRow(p); }});
    }});
    tbody.innerHTML = html;
  }} else {{
    tbody.innerHTML = display.map(personRow).join('');
  }}

  // Footer (averages)
  const tfoot = document.getElementById('heatmap-foot');
  let fRow = '<tr><td colspan="3" style="font-weight:700;font-size:11px;">Averages (' + filtered.length + ' learners)</td>';
  const oAvg = filtered.length ? Math.round(filtered.reduce((s,p) => s+p.overallPct,0)/filtered.length) : 0;
  fRow += '<td class="pct-cell" style="font-weight:700;font-size:11px;">' + oAvg + '%</td>';
  const eAvg = filtered.length ? Math.round(filtered.reduce((s,p) => s+expectedPct(p),0)/filtered.length) : 0;
  fRow += '<td class="pct-cell" style="font-weight:700;font-size:11px;color:var(--muted);">' + eAvg + '%</td>';
  const gAvg = eAvg - oAvg;
  fRow += '<td class="pct-cell"><span class="pct-pill" style="' + gapStyle(gAvg) + '">' + gAvg + '%</span></td>';
  fRow += '<td></td>';
  fRow += '<td></td>';
  fRow += '<td></td>';
  fRow += '<td></td>';
  fRow += '</tr>';
  tfoot.innerHTML = fRow;
}}

function escHtml(s) {{
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}}

/* ── Charts ── */
function renderCharts() {{
  renderMarketChart();
  renderCurricChart();
}}

function renderMarketChart() {{
  const markets = [...new Set(filtered.map(p => p.market))].sort();
  const avgs = markets.map(m => {{
    const mp = filtered.filter(p => p.market === m);
    return Math.round(mp.reduce((s,p) => s+p.overallPct,0)/mp.length);
  }});
  const marketStats = markets.map(m => {{
    const mp = filtered.filter(p => p.market === m);
    const total    = mp.length;
    const complete = mp.filter(p => p.overallDone).length;
    const overdue  = mp.filter(p => computeStatus(p) === 'Overdue').length;
    const onTrack  = total - complete - overdue;
    return {{total, complete, onTrack, overdue}};
  }});
  const ctx = document.getElementById('marketChart').getContext('2d');
  if (marketChartObj) marketChartObj.destroy();
  marketChartObj = new Chart(ctx, {{
    type: 'bar',
    data: {{
      labels: markets,
      datasets: [{{
        label: 'Avg Completion %',
        data: avgs,
        backgroundColor: '#3b82f6aa',
        borderColor: '#3b82f6',
        borderWidth: 1,
        borderRadius: 4,
      }}]
    }},
    options: {{
      indexAxis: 'y',
      responsive: true,
      maintainAspectRatio: false,
      plugins: {{
        legend: {{display:false}},
        tooltip: {{
          callbacks: {{
            label: c => '  Avg progress: ' + c.parsed.x + '%',
            afterLabel: c => {{
              const s = marketStats[c.dataIndex];
              const lines = [
                '  ' + s.total + ' learners in this market',
                '  ✓ ' + s.complete + ' fully complete',
                '  → ' + s.onTrack + ' on track',
              ];
              if (s.overdue > 0) lines.push('  ⚠ ' + s.overdue + ' past their deadline');
              return lines;
            }}
          }}
        }}
      }},
      scales: {{
        x: {{
          min:0, max:100,
          grid: {{color: 'rgba(255,255,255,0.06)'}},
          ticks: {{color: cv('muted'), font:{{size:11}}, callback: v => v+'%'}},
          border: {{color:'transparent'}},
        }},
        y: {{
          grid: {{display:false}},
          ticks: {{color: cv('muted'), font:{{size:11}}}},
          border: {{color:'transparent'}},
        }}
      }}
    }}
  }});
}}

function renderCurricChart() {{
  const labels = CURRIC_IDS.map(cid => CURRIC_NAMES[cid]);
  const avgs = CURRIC_IDS.map(cid => {{
    const vals = filtered.map(p => p.curricula[cid] ? p.curricula[cid].pct : 0);
    return vals.length ? Math.round(vals.reduce((s,v) => s+v,0)/vals.length) : 0;
  }});
  const curricStats = CURRIC_IDS.map(cid => {{
    const people = filtered.map(p => p.curricula[cid]).filter(Boolean);
    const total      = filtered.length;
    const complete   = people.filter(c => c.complete).length;
    const inProgress = people.filter(c => !c.complete && c.pct > 0).length;
    const notStarted = people.filter(c => !c.complete && c.pct === 0).length;
    const pastDue    = filtered.filter(p => {{ const c = p.curricula[cid]; return c && !c.complete && curricDaysLeft(c) !== null && curricDaysLeft(c) <= 0; }}).length;
    return {{total, complete, inProgress, notStarted, pastDue}};
  }});
  const ctx = document.getElementById('curricChart').getContext('2d');
  if (curricChartObj) curricChartObj.destroy();
  const colors = ['#3b82f6','#22c55e','#f59e0b','#a855f7','#ef4444','#14b8a6'];
  curricChartObj = new Chart(ctx, {{
    type: 'bar',
    data: {{
      labels,
      datasets: [{{
        label: 'Avg %',
        data: avgs,
        backgroundColor: colors.map(c => c+'99'),
        borderColor: colors,
        borderWidth: 1,
        borderRadius: 4,
      }}]
    }},
    options: {{
      responsive: true,
      maintainAspectRatio: false,
      plugins: {{
        legend: {{display:false}},
        tooltip: {{
          callbacks: {{
            label: c => '  Avg progress: ' + c.parsed.y + '%',
            afterLabel: c => {{
              const s = curricStats[c.dataIndex];
              const lines = [
                '  ✓ ' + s.complete   + ' of ' + s.total + ' fully complete',
                '  ◑ ' + s.inProgress + ' in progress',
                '  ○ ' + s.notStarted + ' not started',
              ];
              if (s.pastDue > 0) lines.push('  ⚠ ' + s.pastDue + ' past their deadline');
              return lines;
            }}
          }}
        }}
      }},
      scales: {{
        y: {{
          min:0, max:100,
          grid: {{color: 'rgba(255,255,255,0.06)'}},
          ticks: {{color: cv('muted'), font:{{size:11}}, callback: v => v+'%'}},
          border: {{color:'transparent'}},
        }},
        x: {{
          grid: {{display:false}},
          ticks: {{color: cv('muted'), font:{{size:10}}, maxRotation:30}},
          border: {{color:'transparent'}},
        }}
      }}
    }}
  }});
}}

/* ── User detail modal ── */
function openModal(email) {{
  const p = PEOPLE.find(x => x.email === email);
  if (!p) return;
  const status = computeStatus(p);
  const dl = daysLeft(p);
  const elapsed = daysElapsed(p);
  const pct45 = Math.min(100, Math.round(elapsed / PROGRAM_DAYS * 100));
  const barColor = p.overallDone ? 'var(--green)' : dl <= 0 ? 'var(--red)' : 'var(--accent)';

  let daysLabel = '';
  if (p.overallDone) daysLabel = 'Program complete';
  else if (dl === null) daysLabel = 'No assignment date on file';
  else if (dl < 0) daysLabel = Math.abs(dl) + ' days past deadline';
  else if (dl === 0) daysLabel = 'Due today';
  else daysLabel = dl + ' days remaining';

  const statusBadge = '<span class="status-badge ' +
    (status==='Completed'?'sb-completed':status==='On Track'?'sb-ontrack':'sb-overdue') +
    '">' + status + '</span>';

  // Curriculum sections
  let curricHtml = '';
  CURRIC_IDS.forEach(cid => {{
    const c = p.curricula[cid];
    if (!c) return;
    const dl = curricDaysLeft(c);
    const drBg = dl !== null && dl <= 0 ? 'var(--red-subtle)' : 'var(--green-subtle)';
    const drColor = dl !== null && dl <= 0 ? 'var(--red)' : 'var(--green)';
    const drLabel = dl === null ? '' :
      '<span class="days-badge" style="background:' + drBg + ';color:' + drColor + '">' +
      (dl < 0 ? Math.abs(dl) + 'd overdue' : dl === 0 ? 'Due today' : dl + 'd left') + '</span>';
    const doneBadge = c.complete ?
      '<span class="status-badge sb-completed" style="font-size:10px">Done</span>' : '';

    let missingLabel = '';
    if (!c.complete) {{
      const missingCourses   = c.items.filter(i => !i.done && i.type !== 'VILT').length;
      const missingWorkshops = c.items.filter(i => !i.done && i.type === 'VILT').length;
      if (missingCourses > 0 || missingWorkshops > 0) {{
        const parts = [];
        if (missingCourses   > 0) parts.push(missingCourses   + ' ' + (missingCourses   === 1 ? 'course'   : 'courses'));
        if (missingWorkshops > 0) parts.push(missingWorkshops + ' ' + (missingWorkshops === 1 ? 'workshop' : 'workshops'));
        missingLabel = '<span style="font-size:11px;color:var(--muted);white-space:nowrap;">Missing: ' + parts.join(', ') + '</span>';
      }}
    }}

    const itemsHtml = c.items.map(it => {{
      const checkClass = it.done ? 'done' : 'not';
      const checkMark  = it.done ? '&#10003;' : '&nbsp;';
      const typeLabel  = it.type === 'VILT' ?
        '<span class="item-type" style="background:#a855f711;color:#a855f7;">Workshop</span>' :
        '<span class="item-type">Online</span>';
      const dateStr = it.done && it.date ?
        '<span class="item-date">' + fmtDate(it.date) + '</span>' : '';
      return '<div class="item-row">' +
        '<div class="item-check ' + checkClass + '">' + checkMark + '</div>' +
        '<div class="item-title">' + escHtml(it.title) + '</div>' +
        typeLabel + dateStr + '</div>';
    }}).join('');

    curricHtml += '<div class="curric-section">' +
      '<div class="curric-header" onclick="toggleCurric(this)">' +
        '<span class="curric-title">' + escHtml(c.title) + '</span>' +
        '<span class="pct-pill" style="' + curricPillStyle(c, null) + ';font-size:11px">' + c.pct + '%</span>' +
        drLabel + doneBadge + missingLabel +
        '<span class="curric-chevron">&#9660;</span>' +
      '</div>' +
      '<div class="curric-items">' + itemsHtml + '</div>' +
    '</div>';
  }});

  // Activity timeline — deduped playbook visits + LMS completions, sorted by date
  const tlEvents = [];
  const tlSeen = new Set();
  (p.playbook || []).filter(v => !p.assignDate || v.date >= p.assignDate).forEach(v => {{
    const k = (v.page || '') + '|' + (v.date || '');
    const pbSlug = urlSlug(v.url);
    if (!tlSeen.has(k)) {{ tlSeen.add(k); tlEvents.push({{t:'pb', date:v.date||'', page:v.page||'', slug:pbSlug, curricId:PLAYBOOK_CURRIC[pbSlug]||null}}); }}
  }});
  CURRIC_IDS.forEach(cid => {{
    const c = p.curricula[cid];
    if (!c) return;
    c.items.forEach(it => {{
      if (it.done && it.date && (!p.assignDate || it.date >= p.assignDate)) tlEvents.push({{t:'lms', date:it.date, title:it.title, curric:c.title, curricId:cid}});
    }});
  }});
  const _pgIdx = slug => {{ const i = PAGE_ORDER.indexOf(slug); return i >= 0 ? i : 999; }};
  tlEvents.sort((a,b) => {{
    if (a.date !== b.date) return a.date < b.date ? -1 : 1;
    const aSlug = a.t === 'pb' ? a.slug : (COURSE_PAGE[a.title] || '');
    const bSlug = b.t === 'pb' ? b.slug : (COURSE_PAGE[b.title] || '');
    const ai = _pgIdx(aSlug), bi = _pgIdx(bSlug);
    if (ai !== bi) return ai - bi;
    return a.t === 'pb' && b.t !== 'pb' ? -1 : a.t !== 'pb' && b.t === 'pb' ? 1 : 0;
  }});

  // Build filtered playbook visits (on/after assignDate) and per-curriculum date lookup
  const filteredPlaybook = (p.playbook || []).filter(v => !p.assignDate || v.date >= p.assignDate);
  const pbVisitDates = {{}};
  filteredPlaybook.forEach(v => {{
    const cid = PLAYBOOK_CURRIC[urlSlug(v.url)];
    if (cid) {{ if (!pbVisitDates[cid]) pbVisitDates[cid] = new Set(); pbVisitDates[cid].add(v.date); }}
  }});
  function dayBefore(d) {{
    const dt = new Date(d + 'T00:00:00'); dt.setDate(dt.getDate() - 1);
    return dt.toISOString().slice(0, 10);
  }}
  // Path score: of all individual completions, what fraction had a playbook visit same day or day before
  let ledCount = 0, totalCount = 0;
  CURRIC_IDS.forEach(cid => {{
    const c = p.curricula[cid];
    if (!c) return;
    const visits = pbVisitDates[cid];
    c.items.filter(i => i.done && i.date && (!p.assignDate || i.date >= p.assignDate)).forEach(i => {{
      totalCount++;
      if (visits && (visits.has(i.date) || visits.has(dayBefore(i.date)))) ledCount++;
    }});
  }});
  let pathLabel = '', pathColor = '';
  if (totalCount > 0) {{
    const ratio = ledCount / totalCount;
    const pct = Math.round(ratio * 100) + '%';
    if (ratio >= 0.6)      {{ pathLabel = 'Playbook-led ' + pct;      pathColor = 'var(--green)'; }}
    else if (ratio >= 0.3) {{ pathLabel = 'Mixed ' + pct;              pathColor = 'var(--amber)'; }}
    else                   {{ pathLabel = 'Skipping playbook ' + pct;  pathColor = 'var(--red)'; }}
  }}

  // Sequence check: first completion dates should be non-decreasing across CURRIC_IDS order
  const curricSeq = CURRIC_IDS.map(cid => {{
    const c = p.curricula[cid];
    if (!c) return null;
    const first = c.items.filter(i => i.done && i.date).map(i => i.date).sort()[0];
    return first ? {{cid, first}} : null;
  }}).filter(Boolean);
  let seqLabel = '', seqColor = '';
  if (curricSeq.length >= 2) {{
    const inSeq = curricSeq.every((x, i) => i === 0 || x.first >= curricSeq[i-1].first);
    seqLabel = inSeq ? 'In sequence' : 'Out of sequence';
    seqColor = inSeq ? 'var(--green)' : 'var(--amber)';
  }}

  const tlHtml = tlEvents.length === 0
    ? '<div style="font-size:12px;color:var(--muted);padding:12px 14px;">No activity recorded yet.</div>'
    : tlEvents.map(ev => ev.t === 'pb'
        ? '<div class="item-row"><div style="width:10px;height:10px;border-radius:50%;background:var(--accent);flex-shrink:0;"></div><div style="flex:1;font-size:12px;"><span style="color:var(--accent);font-weight:500;">Playbook</span> &mdash; ' + escHtml(ev.page) + '</div><div class="item-date">' + fmtDate(ev.date) + '</div></div>'
        : '<div class="item-row"><div style="width:10px;height:10px;border-radius:50%;background:var(--green);flex-shrink:0;"></div><div style="flex:1;font-size:12px;"><span style="color:var(--green);font-weight:500;">Completed</span> &mdash; ' + escHtml(ev.title) + ' <span style="color:var(--muted);font-size:11px;">(' + escHtml(ev.curric) + ')</span></div><div class="item-date">' + fmtDate(ev.date) + '</div></div>'
      ).join('');

  const pSale = SALES_MAP[email];
  const saleHtml = pSale
    ? '<span>&middot; First Sale: <strong style="color:var(--accent)">' + (pSale.daysToClose != null ? pSale.daysToClose + 'd' : '') + '</strong> &middot; $' + Math.round(pSale.amount).toLocaleString('en-US') + ' &middot; ' + escHtml(pSale.accountName) + ' &middot; ' + fmtDate(pSale.closeDate) + '</span>'
    : '<span style="color:var(--muted);">&middot; First Sale: None yet</span>';
  const content = `
    <div class="modal-header" style="padding-right:52px;">
      <div style="position:absolute;top:17px;right:54px;">${{statusBadge}}</div>
      <div class="modal-name">${{escHtml(p.name)}}${{p.jobTitle ? ' <span style="font-size:14px;font-weight:400;color:var(--muted);">&middot; ' + escHtml(p.jobTitle) + '</span>' : ''}}</div>
      <div class="modal-meta" style="margin-top:6px;gap:8px;flex-wrap:wrap;">
        ${{p.email ? '<a href="mailto:' + escHtml(p.email) + '" style="color:var(--accent)">' + escHtml(p.email) + '</a>' : ''}}
        ${{p.market ? '<span>&middot; ' + escHtml(p.market) + '</span>' : ''}}
        ${{p.hireDate ? '<span>&middot; Hired ' + fmtDate(p.hireDate) + '</span>' : ''}}
        ${{p.assignDate ? '<span>&middot; Enrolled ' + fmtDate(p.assignDate) + '</span>' : ''}}
        ${{saleHtml}}
      </div>
      ${{p.manager ? '<div class="modal-meta" style="margin-top:3px;gap:6px;"><span style="color:var(--muted);font-size:11px;text-transform:uppercase;letter-spacing:.4px;font-weight:600;">Manager</span><span>' + escHtml(p.manager) + '</span><span>&middot; ' + escHtml(p.mgrTitle) + '</span><span>&middot; <a href="mailto:' + escHtml(p.mgrEmail) + '" style="color:var(--accent)">' + escHtml(p.mgrEmail) + '</a></span></div>' : ''}}
    </div>
    <div class="modal-body">
      <div class="progress-section">
        <div class="progress-label">
          <span>35-Day LMS Window</span>
          <span style="font-size:11px;color:var(--muted)">${{daysLabel}}</span>
        </div>
        <div class="progress-bar-wrap">
          <div class="progress-bar-fill" style="width:${{pct45}}%;background:${{barColor}}"></div>
        </div>
        <div class="progress-sublabel" style="display:flex;gap:16px;flex-wrap:wrap;align-items:center;">
          <span>Actual: <strong>${{p.overallPct}}%</strong></span>
          <span style="color:var(--muted)">Expected: <strong style="color:var(--text)">${{expectedPct(p)}}%</strong></span>
          <span>Gap: <strong style="color:${{gapPct(p) > 40 ? 'var(--red)' : gapPct(p) > 0 ? '#b45309' : 'var(--green)'}}">${{gapPct(p)}}%</strong>
            ${{gapPct(p) > 40 ? '<span style="font-size:11px;color:var(--red)">— coaching recommended</span>' : gapPct(p) > 0 ? '<span style="font-size:11px;color:var(--muted)">— check in</span>' : '<span style="font-size:11px;color:var(--green)">— on pace</span>'}}</span>
        </div>
      </div>

      ${{pbSummaryHtml(p)}}

      <div style="font-size:12px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px;">Curriculum Progress</div>
      <div style="display:flex;gap:12px;flex-wrap:wrap;font-size:11px;color:var(--muted);margin-bottom:10px;align-items:center;">
        <span style="display:inline-flex;align-items:center;gap:4px;"><span style="width:8px;height:8px;background:#15803d;border-radius:2px;display:inline-block;"></span>Done</span>
        <span style="display:inline-flex;align-items:center;gap:4px;"><span style="width:8px;height:8px;background:#1d4ed8;border-radius:2px;display:inline-block;"></span>In Progress</span>
        <span style="display:inline-flex;align-items:center;gap:4px;"><span style="width:8px;height:8px;background:#b91c1c;border-radius:2px;display:inline-block;"></span>Past Due</span>
        <span style="display:inline-flex;align-items:center;gap:4px;"><span style="width:8px;height:8px;background:#6b7280;border-radius:2px;display:inline-block;"></span>Not Started</span>
      </div>
      ${{curricHtml}}

      <div class="curric-section" style="margin-top:20px;margin-bottom:0;">
        <div class="curric-header" onclick="toggleCurric(this)">
          <span class="curric-title">Activity Timeline</span>
          <span style="font-size:11px;color:var(--muted);">${{tlEvents.length}} event${{tlEvents.length !== 1 ? 's' : ''}}</span>
          ${{pathLabel ? '<span style="font-size:11px;font-weight:700;color:' + pathColor + ';">' + pathLabel + '</span>' : ''}}
          ${{seqLabel ? '<span style="font-size:11px;font-weight:700;color:' + seqColor + ';">' + seqLabel + '</span>' : ''}}
          <span class="curric-chevron">&#9660;</span>
        </div>
        <div class="curric-items">${{tlHtml}}</div>
      </div>
    </div>
  `;

  document.getElementById('modal-content').innerHTML = content;
  document.getElementById('modal-overlay').classList.add('open');
}}

function toggleCurric(header) {{
  header.classList.toggle('open');
  header.nextElementSibling.classList.toggle('open');
}}

function closeModal() {{
  document.getElementById('modal-overlay').classList.remove('open');
}}

function closeModalBg(e) {{
  if (e.target === document.getElementById('modal-overlay')) closeModal();
}}

/* ── Export ── */
function runExport(type) {{
  document.getElementById('export-menu').classList.remove('open');
  const now = new Date().toLocaleDateString('en-US', {{month:'long', day:'numeric', year:'numeric'}});
  document.getElementById('ph-sub').textContent = filtered.length + ' learners · ' + now +
    (type !== 'full' ? ' · Filters active' : '');
  document.getElementById('ph-report-type').textContent =
    type === 'full' ? 'Full Report' :
    type === 'overdue' ? 'Overdue Learners' : 'Manager Summary';

  const wrap = document.getElementById('print-roster-wrap');
  if (type === 'full') {{
    const rows = filtered.map(p => {{
      const status = computeStatus(p);
      const dl = daysLeft(p);
      const daysStr = p.overallDone ? 'Complete' : dl === null ? '—' : dl < 0 ? (Math.abs(dl) + 'd overdue') : (dl + 'd left');
      return '<tr><td>' + escHtml(p.name) + '</td><td>' + escHtml(p.market) + '</td><td>' + status + '</td><td>' + daysStr + '</td><td>' + p.overallPct + '%</td><td>' + escHtml(p.manager) + '</td></tr>';
    }}).join('');
    wrap.innerHTML = '<table><thead><tr><th>Name</th><th>Market</th><th>Status</th><th>Days Left</th><th>Overall %</th><th>Manager</th></tr></thead><tbody>' + rows + '</tbody></table>';
  }} else if (type === 'overdue') {{
    const od = filtered.filter(p => computeStatus(p) === 'Overdue');
    const rows = od.map(p => '<tr><td>' + escHtml(p.name) + '</td><td>' + escHtml(p.email) + '</td><td>' + escHtml(p.market) + '</td><td>' + p.overallPct + '%</td><td>' + escHtml(p.manager) + '</td><td>' + escHtml(p.mgrEmail) + '</td></tr>').join('');
    wrap.innerHTML = '<table><thead><tr><th>Name</th><th>Email</th><th>Market</th><th>Completion %</th><th>Manager</th><th>Manager Email</th></tr></thead><tbody>' + rows + '</tbody></table>';
  }} else {{
    const byMgr = {{}};
    filtered.forEach(p => {{
      const m = p.manager || 'Unknown';
      if (!byMgr[m]) byMgr[m] = {{mgrEmail:p.mgrEmail, people:[]}};
      byMgr[m].people.push(p);
    }});
    const rows = Object.entries(byMgr).sort((a,b)=>a[0].localeCompare(b[0])).map(([mgr,data]) => {{
      const avg = Math.round(data.people.reduce((s,p)=>s+p.overallPct,0)/data.people.length);
      const od = data.people.filter(p=>computeStatus(p)==='Overdue').length;
      return '<tr><td>' + escHtml(mgr) + '</td><td>' + escHtml(data.mgrEmail) + '</td><td>' + data.people.length + '</td><td>' + od + '</td><td>' + avg + '%</td></tr>';
    }}).join('');
    wrap.innerHTML = '<table><thead><tr><th>Manager</th><th>Manager Email</th><th>Team Count</th><th>Overdue</th><th>Avg Completion %</th></tr></thead><tbody>' + rows + '</tbody></table>';
  }}
  window.print();
}}

/* ── Init ── */
function init() {{
  // Populate market dropdown
  const markets = [...new Set(PEOPLE.map(p => p.market))].sort();
  const sel = document.getElementById('f-market');
  markets.forEach(m => {{
    const o = document.createElement('option');
    o.value = m; o.textContent = m;
    sel.appendChild(o);
  }});

  // Close dropdowns on outside click
  document.addEventListener('click', e => {{
    if (!document.getElementById('export-drop').contains(e.target))
      document.getElementById('export-menu').classList.remove('open');
    if (!document.getElementById('info-popover').contains(e.target) && !e.target.classList.contains('info-btn'))
      document.getElementById('info-popover').classList.remove('visible');
  }});

  applyFilters();
}}

function runExportXLSX(type){{
  document.getElementById('export-menu').classList.remove('open');
  const now=new Date().toLocaleDateString('en-US',{{year:'numeric',month:'long',day:'numeric'}});
  function makeSheet(rows,colWidths){{
    const ws=XLSX.utils.aoa_to_sheet(rows);
    ws['!cols']=colWidths.map(w=>({{wch:w}}));
    return ws;
  }}
  function dlXLSX(name,wb){{ XLSX.writeFile(wb,name+'.xlsx'); }}
  const wb=XLSX.utils.book_new();
  if(type==='full'){{
    const rows=[['Name','Market','Status','Days Left','Overall %','Manager']];
    filtered.forEach(p=>{{
      const status=computeStatus(p);
      const d=daysLeft(p);
      const daysStr=p.overallDone?'Complete':d===null?'—':d<0?(Math.abs(d)+'d overdue'):(d+'d left');
      rows.push([p.name,p.market,status,daysStr,p.overallPct+'%',p.manager]);
    }});
    XLSX.utils.book_append_sheet(wb,makeSheet(rows,[28,18,14,14,12,28]),'Full Report');
    dlXLSX('onboarding-full-report',wb);
  }} else if(type==='overdue'){{
    const od=filtered.filter(p=>computeStatus(p)==='Overdue');
    const rows=[['Name','Email','Market','Completion %','Manager','Manager Email']];
    od.forEach(p=>rows.push([p.name,p.email,p.market,p.overallPct+'%',p.manager,p.mgrEmail]));
    XLSX.utils.book_append_sheet(wb,makeSheet(rows,[28,32,18,14,28,32]),'Overdue Only');
    dlXLSX('onboarding-overdue',wb);
  }} else if(type==='manager-summary'){{
    const byMgr={{}};
    filtered.forEach(p=>{{
      const m=p.manager||'Unknown';
      if(!byMgr[m]) byMgr[m]={{mgrEmail:p.mgrEmail,people:[]}};
      byMgr[m].people.push(p);
    }});
    const rows=[['Manager','Manager Email','Team Count','Overdue','Avg Completion %']];
    Object.entries(byMgr).sort((a,b)=>a[0].localeCompare(b[0])).forEach(([mgr,data])=>{{
      const avg=Math.round(data.people.reduce((s,p)=>s+p.overallPct,0)/data.people.length);
      const od=data.people.filter(p=>computeStatus(p)==='Overdue').length;
      rows.push([mgr,data.mgrEmail,data.people.length,od,avg+'%']);
    }});
    XLSX.utils.book_append_sheet(wb,makeSheet(rows,[28,32,12,10,16]),'Manager Summary');
    dlXLSX('onboarding-manager-summary',wb);
  }}
}}

init();
(function(){{
  var n=0,t;
  var h=document.querySelector('.header h1');
  if(h) h.addEventListener('click',function(){{
    n++;clearTimeout(t);
    if(n>=3){{n=0;window.location.href='index.html';}}
    else t=setTimeout(function(){{n=0;}},1500);
  }});
}})();
</script>
</body>
</html>"""


def main():
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    records = load_lms()
    print(f"Loaded {len(records)} learners from LMS data")
    records = attach_playbook(records)
    pb_matches = sum(1 for p in records.values() if p['playbook'])
    print(f"Playbook activity matched for {pb_matches} learners")
    sales_map = _load_salesforce(records)
    print(f"Salesforce: {len(sales_map)} cohort members with a Closed Won deal")
    html = generate_html(records, sales_map)
    with open('onboarding.html', 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"Generated onboarding.html")


if __name__ == '__main__':
    main()

