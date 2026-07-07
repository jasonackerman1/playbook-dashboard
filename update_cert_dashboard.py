#!/usr/bin/env python3
"""update_cert_dashboard.py — generates cert-{vertical}.html from LMS Excel exports in cert-data/.
Accumulates all monthly files per vertical; deduplicates by Email so each person appears once."""

import openpyxl
import os
import re
import json
import warnings
warnings.filterwarnings('ignore')

TLG = {
    "Jason Ackerman","Bianca Davis","James Parker","Resmie Biba",
    "Chris Curtis","Sara Thompson","Jeremy MacBean","Bradley Pierce",
    "Laura Sefcik","Samantha Maresca","Staci Musco","CJ Homer","Rich Moore","Dale Kinsey",
    "John Lechner","Resmie Nesimi","Samantha D'Angelo","Bianca DiPasquale","Doug Falk"
}

VERTICAL_MAP = {
    'healthcare':   'Healthcare',
    'publicsector': 'Public Sector',
    'legal':        'Legal',
    'salesforce':   'Salesforce',
    'dx':           'DX Playbook',
    'gcip':         'GC/IP Sales',
    'iq501':        'IQ501',
    'roadtodx':     'Road to DX',
}

# Column indices (0-based) matching Resmie's LMS export format.
# CAUTION: verify against each new file before regenerating — Resmie may adjust columns.
# Last verified: FY26-Healthcare-Certification-Courses_2026-05.xlsx
COL_FIRST        = 2   # First Name
COL_LAST         = 3   # Last Name
COL_EMAIL        = 4   # Email Address
COL_JOBTITLE     = 5   # Job Title
COL_REGION       = 6   # Region (in data; Market used for grouping/filtering)
COL_MARKET       = 7   # Market
COL_MGR_FIRST    = 9   # ManagerFirstName
COL_MGR_LAST     = 10  # ManagerLastName
COL_MGR_EMAIL    = 11  # SUPEMAILADDR (manager email)
COL_MGR_TITLE    = 12  # Manager JobTitle
COL_HIRE_DATE    = 14  # Hire Date
COL_COMPLETE     = 19  # Curriculum Complete — "Yes"/"No" (LMS-driven)
COL_DATE         = 20  # Curriculum Assignment Date (enrollment date; used for date-range filter)
COL_HC_DATE      = 21  # Healthcare Certification Date (compute quarter from this)
# col 22 = HC Qtr Certified — Excel formula string, ignored; quarter computed via km_fiscal_quarter()
COL_HEALTHCARE   = 23  # HC Certified (manager sign-off)
COL_ACUTE_DATE   = 24  # Acute Care Certification Date
# col 25 = Acute Care Qtr Certified — Excel formula string, ignored
COL_ACUTE        = 26  # Acute Care Certified
COL_AMB_DATE     = 27  # Ambulatory Certification Date
# col 28 = Ambulatory Qtr Certified — Excel formula string, ignored
COL_AMBULATORY   = 29  # Ambulatory Certified
COL_EXT_DATE     = 30  # Extended Care Certification Date
# col 31 = Extended Care Qtr Certified — Excel formula string, ignored
COL_EXTENDED     = 32  # Extended Care Certified
# Layered Security: completely removed from new file

# ── Public Sector column indices (cols 0-17 identical to Healthcare) ──────────
PS_COL_COMPLETE  = 19  # Public Sector Curriculum Status (Yes/No) — only cert field
PS_COL_DATE      = 20  # Curriculum Assignment Date (for date-range filter)
PS_COL_CERT_DATE = 21  # Certification Date (for display and trend chart)
# col 22 = Qr Certified — Excel formula; quarter computed via km_fiscal_quarter()

SUB_CERTS = [
    ('Healthcare', 'Healthcare'),
    ('AcuteCare',  'Acute Care'),
    ('Ambulatory', 'Ambulatory'),
    ('Extended',   'Extended Care'),
]


def km_fiscal_quarter(date):
    """KM fiscal year starts April. Q1=Apr-Jun, Q2=Jul-Sep, Q3=Oct-Dec, Q4=Jan-Mar."""
    if date is None:
        return ''
    import math
    m = date.month
    # Shift so April=1 ... March=12, then divide into 4 equal quarters
    q = math.ceil(((m - 4) % 12 + 1) / 3)
    fy = date.year if m >= 4 else date.year - 1
    return f'FY{fy} Q{q}'


def fmt_date_label(yyyy_mm):
    """Convert 'YYYY-MM' to 'Month YYYY' for display."""
    months = ['January','February','March','April','May','June',
              'July','August','September','October','November','December']
    try:
        y, m = yyyy_mm.split('-')
        return f"{months[int(m)-1]} {y}"
    except Exception:
        return yyyy_mm


def extract_file_date(fname):
    """Extract YYYY-MM from filename for chronological ordering of monthly files."""
    m = re.search(r'(\d{4}-\d{2})', os.path.basename(fname))
    return m.group(1) if m else '0000-00'


def person_key(r):
    """Unique identifier for deduplication — Email preferred, name as fallback."""
    return r['Email'].lower() if r['Email'] else f"{r['FirstName']} {r['LastName']}".lower()


def detect_vertical(fname):
    """Extract vertical slug from filename — supports both naming conventions."""
    fn = os.path.basename(fname).lower()
    # cert-healthcare-2026-05.xlsx
    m = re.match(r'cert-([a-z0-9]+)-', fn)
    if m:
        return m.group(1)
    # FY26-Healthcare-Certification-Courses_2026-05.xlsx
    m = re.match(r'fy\d+-([a-z]+)-', fn)
    if m:
        return m.group(1).lower()
    # Healthcare=Certification-Report-07.06.2026.xlsx / Healthcare-Certification-Foundations-...
    for slug in VERTICAL_MAP:
        if fn.startswith(slug):
            return slug
    return None


def load_rows(filepath):
    """Load non-blank rows, return list of dicts."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw[COL_FIRST] is None:
            continue
        def _date(v): return v.strftime('%Y-%m-%d') if v and hasattr(v, 'strftime') else ''
        def _qtr(v):  return str(v).strip() if v and str(v).strip() not in ('None','') else ''
        def _cert(v): return str(v).strip() if v else 'No'
        def _qtr_from_date(v):
            return km_fiscal_quarter(v) if v and hasattr(v, 'month') else ''
        rows.append({
            'FirstName':  str(raw[COL_FIRST]).strip(),
            'LastName':   str(raw[COL_LAST]).strip(),
            'Email':      str(raw[COL_EMAIL]).strip() if raw[COL_EMAIL] else '',
            'JobTitle':   str(raw[COL_JOBTITLE]).strip() if raw[COL_JOBTITLE] else '',
            'Market':     str(raw[COL_MARKET]).strip() if raw[COL_MARKET] else '',
            'Manager':    ((str(raw[COL_MGR_FIRST]).strip() + ' ' + str(raw[COL_MGR_LAST]).strip()).strip()) if raw[COL_MGR_FIRST] else '',
            'MgrEmail':   str(raw[COL_MGR_EMAIL]).strip() if raw[COL_MGR_EMAIL] else '',
            'MgrTitle':   str(raw[COL_MGR_TITLE]).strip() if raw[COL_MGR_TITLE] else '',
            'HireDate':   _date(raw[COL_HIRE_DATE]),
            'Complete':   _cert(raw[COL_COMPLETE]),
            'Date':       _date(raw[COL_DATE]),
            'HCDate':     _date(raw[COL_HC_DATE]),
            'HCQtr':      _qtr_from_date(raw[COL_HC_DATE]),
            'Healthcare': _cert(raw[COL_HEALTHCARE]),
        })
    wb.close()
    return rows


def load_rows_publicsector(filepath):
    """Load Public Sector rows — single curriculum, no sub-certs."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw[COL_FIRST] is None:
            continue
        def _date(v): return v.strftime('%Y-%m-%d') if v and hasattr(v, 'strftime') else ''
        def _cert(v): return str(v).strip() if v else 'No'
        def _qtr_from_date(v):
            return km_fiscal_quarter(v) if v and hasattr(v, 'month') else ''
        rows.append({
            'FirstName':   str(raw[COL_FIRST]).strip(),
            'LastName':    str(raw[COL_LAST]).strip(),
            'Email':       str(raw[COL_EMAIL]).strip() if raw[COL_EMAIL] else '',
            'JobTitle':    str(raw[COL_JOBTITLE]).strip() if raw[COL_JOBTITLE] else '',
            'Market':      str(raw[COL_MARKET]).strip() if raw[COL_MARKET] else '',
            'Manager':     ((str(raw[COL_MGR_FIRST]).strip() + ' ' + str(raw[COL_MGR_LAST]).strip()).strip()) if raw[COL_MGR_FIRST] else '',
            'MgrEmail':    str(raw[COL_MGR_EMAIL]).strip() if raw[COL_MGR_EMAIL] else '',
            'MgrTitle':    str(raw[COL_MGR_TITLE]).strip() if raw[COL_MGR_TITLE] else '',
            'HireDate':    _date(raw[COL_HIRE_DATE]),
            'PublicSector': _cert(raw[PS_COL_COMPLETE]),
            'Date':        _date(raw[PS_COL_DATE]),
            'CertDate':    _date(raw[PS_COL_CERT_DATE]),
            'CertQtr':     _qtr_from_date(raw[PS_COL_CERT_DATE]),
        })
    wb.close()
    return rows


def generate_html(slug, name, rows):
    raw_json = json.dumps(rows)
    tlg_json = json.dumps(sorted(TLG))

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{name} Certification Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<style>
  :root {{
    --bg:#0f1117; --surface:#1a1d27; --surface2:#22263a; --border:#2e3350;
    --accent:#4f8ef7; --accent2:#7c5cfc; --accent3:#f7c94f;
    --text:#e8ecf4; --muted:#7b82a0; --green:#3ecf8e; --red:#f76f6f;
    --teal:#2dd4bf; --green-subtle:#3ecf8e22; --red-subtle:#f76f6f22;
    --font:'Segoe UI',system-ui,sans-serif;
  }}
  body.light-mode {{
    --bg:#f4f6fb; --surface:#ffffff; --surface2:#eef1f7; --border:#d0d7e8;
    --accent:#2563eb; --accent2:#6d28d9; --accent3:#d97706;
    --text:#1a1d27; --muted:#475569; --green:#059669; --red:#dc2626;
    --teal:#0f766e; --green-subtle:#05966922; --red-subtle:#dc262622;
  }}
  body.light-mode select,body.light-mode input[type=date]{{color-scheme:light;}}
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{background:var(--bg);color:var(--text);font-family:var(--font);min-height:100vh;transition:background .2s,color .2s;}}

  .header{{padding:20px 28px 16px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px;}}
  .header-left{{display:flex;align-items:center;gap:16px;}}
  .info-btn{{display:inline-flex;align-items:center;justify-content:center;width:15px;height:15px;border-radius:50%;background:var(--surface2);border:1px solid var(--border);color:var(--muted);font-size:9px;font-weight:700;cursor:pointer;margin-left:5px;vertical-align:middle;flex-shrink:0;line-height:1;transition:border-color .15s,color .15s;}}
  .info-btn:hover{{border-color:var(--accent);color:var(--accent);}}
  .info-popover{{position:fixed;z-index:9999;background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:12px 14px;font-size:12px;color:var(--text);line-height:1.6;max-width:260px;box-shadow:0 4px 24px rgba(0,0,0,0.5);display:none;}}
  .info-popover.visible{{display:block;}}
  .header h1{{font-size:18px;font-weight:700;letter-spacing:.3px;}}
  .header h1 span{{color:var(--muted);font-weight:400;}}
  .header-date{{font-size:11px;color:var(--muted);margin-top:2px;}}
  .btn-theme{{background:transparent;border:1px solid var(--border);color:var(--muted);border-radius:6px;padding:5px 12px;font-size:12px;cursor:pointer;transition:all .15s;}}
  .btn-theme:hover{{border-color:var(--accent);color:var(--text);}}
  .btn-export{{background:var(--accent);border:1px solid var(--accent);color:#fff;border-radius:6px;padding:5px 12px;font-size:12px;cursor:pointer;transition:all .15s;font-weight:600;}}
  .btn-export:hover{{opacity:0.88;}}
  .export-drop{{position:relative;}}
  .export-menu{{position:absolute;top:calc(100% + 6px);right:0;background:var(--surface);border:1px solid var(--border);border-radius:8px;min-width:210px;box-shadow:0 4px 24px rgba(0,0,0,.28);display:none;z-index:200;overflow:visible;}}
  .export-menu.open{{display:block;}}
  .export-item{{display:block;width:100%;text-align:left;padding:10px 14px;font-size:13px;color:var(--text);background:transparent;border:none;cursor:pointer;transition:background .1s;font-family:inherit;}}
  .export-item:hover{{background:var(--surface2);}}
  .export-parent{{position:relative;display:flex;justify-content:space-between;align-items:center;padding:10px 14px;font-size:13px;color:var(--text);cursor:default;transition:background .1s;}}
  .export-parent:hover{{background:var(--surface2);}}
  .export-chevron{{font-size:11px;color:var(--muted);margin-left:10px;}}
  .export-submenu{{position:absolute;right:100%;top:0;background:var(--surface);border:1px solid var(--border);border-radius:8px;min-width:90px;box-shadow:0 4px 24px rgba(0,0,0,0.28);display:none;z-index:201;overflow:hidden;margin-right:4px;}}
  .export-parent:hover .export-submenu{{display:block;}}

  .filters{{padding:14px 28px;display:flex;gap:10px;flex-wrap:wrap;align-items:center;border-bottom:1px solid var(--border);background:var(--surface);}}
  .filter-label{{font-size:12px;color:var(--muted);text-transform:uppercase;letter-spacing:.6px;margin-right:4px;}}
  select,input[type=date],input[type=text]{{background:var(--surface2);border:1px solid var(--border);color:var(--text);border-radius:6px;padding:6px 10px;font-size:13px;cursor:pointer;outline:none;color-scheme:dark;}}
  select:focus,input:focus{{border-color:var(--accent);}}
  .btn-reset{{background:transparent;border:1px solid var(--border);color:var(--muted);border-radius:6px;padding:6px 14px;font-size:12px;cursor:pointer;transition:border-color .15s,color .15s;}}
  .btn-reset:hover{{border-color:var(--accent);color:var(--text);}}
  .result-count{{margin-left:auto;font-size:12px;color:var(--muted);}}

  .stats{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;padding:20px 28px;}}
  .stat{{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:16px 18px;}}
  .stat-label{{font-size:11px;text-transform:uppercase;letter-spacing:.7px;color:var(--muted);margin-bottom:6px;}}
  .stat-value{{font-size:28px;font-weight:700;line-height:1;}}
  .stat-value.green{{color:var(--green);}}
  .stat-value.red{{color:var(--red);}}
  .stat-value.teal{{color:var(--teal);}}
  .stat-value.blue{{color:var(--accent);}}
  .stat-sub{{font-size:11px;color:var(--muted);margin-top:4px;}}

  .charts{{display:grid;grid-template-columns:1fr 1fr;gap:16px;padding:0 28px 16px;}}
  @media(max-width:900px){{.charts{{grid-template-columns:1fr;}}}}
  @media(max-width:480px){{.chart-wrap{{height:180px;}}}}
  .chart-card{{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:18px;}}
  .chart-title{{font-size:13px;font-weight:600;margin-bottom:14px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;}}
  body.light-mode .chart-title{{color:var(--text);}}
  .chart-wrap{{position:relative;height:260px;}}

  .section{{padding:0 28px 32px;}}
  .section-header{{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;flex-wrap:wrap;gap:8px;}}
  .section-title{{font-size:13px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;}}
  body.light-mode .section-title{{color:var(--text);}}
  .section-hint{{font-size:11px;color:var(--muted);margin-top:3px;}}

  .roster-search{{width:180px;}}
  .roster-wrap{{display:flex;border:1px solid var(--border);border-radius:10px;overflow:hidden;}}
  .roster-left{{width:300px;flex-shrink:0;overflow-y:auto;max-height:820px;border-right:1px solid var(--border);}}
  .roster-person{{display:flex;align-items:center;gap:10px;padding:10px 14px;cursor:pointer;border-bottom:1px solid var(--border);transition:background .1s;}}
  .roster-person:last-child{{border-bottom:none;}}
  .roster-person:hover{{background:var(--surface2);}}
  .roster-person.active{{background:#4f8ef711;border-left:3px solid var(--accent);padding-left:11px;}}
  .roster-name{{flex:1;font-size:13px;}}
  .cert-dot{{width:8px;height:8px;border-radius:50%;flex-shrink:0;}}
  .cert-dot.yes{{background:var(--green);}}
  .cert-dot.no{{background:var(--red);}}
  .cert-badge{{font-size:11px;font-weight:700;border-radius:10px;padding:2px 8px;white-space:nowrap;}}
  .cert-badge.yes{{color:var(--green);background:var(--green-subtle);}}
  .cert-badge.no{{color:var(--red);background:var(--red-subtle);}}
  .roster-right{{flex:1;overflow-y:auto;max-height:820px;padding:16px 20px;}}
  .roster-right-header{{margin-bottom:14px;padding-bottom:12px;border-bottom:1px solid var(--border);}}
  .no-data{{text-align:center;color:var(--muted);padding:40px;font-size:13px;}}
  @media(max-width:680px){{.roster-wrap{{flex-direction:column;}}.roster-left{{width:100%;max-height:220px;border-right:none;border-bottom:1px solid var(--border);}}}}

  .sort-btn{{background:transparent;border:1px solid var(--border);color:var(--muted);border-radius:6px;padding:4px 10px;font-size:11px;cursor:pointer;transition:all .15s;white-space:nowrap;}}
  .sort-btn:hover{{border-color:var(--accent);color:var(--text);}}
  .sort-btn.active{{border-color:var(--accent);color:var(--accent);background:var(--accent)11;}}
  .mgr-group-hdr{{display:flex;align-items:center;justify-content:space-between;padding:10px 14px;cursor:pointer;background:var(--surface2);border-bottom:1px solid var(--border);user-select:none;}}
  .mgr-group-hdr:hover{{background:var(--border);}}
  .mgr-group-hdr.open .mgr-chevron{{transform:rotate(180deg);}}
  .mgr-chevron{{transition:transform .15s;font-size:10px;color:var(--muted);flex-shrink:0;}}
  .mgr-team{{display:none;}}
  .mgr-team.open{{display:block;}}
  .detail-grid{{display:grid;grid-template-columns:1fr 1fr;gap:12px 28px;margin-top:14px;}}
  .detail-label{{font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:3px;}}
  .detail-value{{font-size:14px;font-weight:500;}}
  .badge-status{{display:inline-block;padding:3px 12px;border-radius:20px;font-size:12px;font-weight:700;}}
  .badge-status.certified{{background:var(--green-subtle);color:var(--green);}}
  .badge-status.not-certified{{background:var(--red-subtle);color:var(--red);}}
  @page{{size:landscape;margin:.65in;}}
  @media print{{
    body{{background:#fff!important;color:#111!important;}}
    .header,.filters,.stats,.charts,.section,.print-hide{{display:none!important;}}
    #print-header{{display:block!important;}}
    #print-stats{{display:flex!important;gap:40px;flex-wrap:wrap;margin-bottom:18px;padding-bottom:16px;border-bottom:2px solid #dde4f0;}}
    #print-charts{{display:flex!important;gap:48px;flex-wrap:wrap;margin-bottom:18px;padding-bottom:16px;border-bottom:2px solid #dde4f0;}}
    #print-roster-wrap{{display:block!important;}}
    body.print-no-summary #print-stats,body.print-no-summary #print-charts{{display:none!important;}}
    .ptable{{width:100%;border-collapse:collapse;font-size:11px;}}
    .ptable th{{background:#f0f4ff;color:#111;font-weight:700;padding:5px 8px;border:1px solid #ccc;text-align:left;}}
    .ptable td{{padding:5px 8px;border:1px solid #ddd;vertical-align:middle;}}
    .ptable tr:nth-child(even) td{{background:#fafafa;}}
  }}
</style>
</head>
<body>

<div class="header">
  <div class="header-left">
    <h1>{name} <span>Certification Dashboard</span></h1>
  </div>
  <div style="display:flex;gap:8px;align-items:center;">
    <div class="export-drop print-hide" id="export-drop">
      <button class="btn-export" onclick="toggleExportDrop()">&#128438; Export &#9660;</button>
      <div class="export-menu" id="export-menu">
        <div class="export-parent">Full Report<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('full')">PDF</button><button class="export-item" onclick="runExportXLSX('full')">Excel</button></div></div>
        <div class="export-parent">Not Certified<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('not-certified')">PDF</button><button class="export-item" onclick="runExportXLSX('not-certified')">Excel</button></div></div>
        <div class="export-parent">Manager Summary<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('manager-summary')">PDF</button><button class="export-item" onclick="runExportXLSX('manager-summary')">Excel</button></div></div>
        <div class="export-parent">Action Required<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('action-required')">PDF</button><button class="export-item" onclick="runExportXLSX('action-required')">Excel</button></div></div>
      </div>
    </div><span class="info-btn print-hide" onclick="showInfo(event,'export')">?</span>
    <button class="btn-theme" id="btn-theme" onclick="toggleTheme()">&#9728; Light</button>
  </div>
</div>

<div class="filters">
  <span class="filter-label">Status</span>
  <select id="f-status">
    <option value="">All</option>
    <option value="Yes">Certified</option>
    <option value="No">Not Certified</option>
  </select>
  <span class="filter-label">Market</span>
  <select id="f-market"><option value="">All Markets</option></select>
  <span class="filter-label" style="margin-right:2px">From</span>
  <input type="date" id="f-date-from">
  <span class="filter-label" style="margin:0 2px">To</span>
  <input type="date" id="f-date-to">
  <button class="btn-reset" onclick="resetFilters()">Reset</button>
  <span class="result-count" id="result-count"></span>
  <span id="in-progress-box" style="display:none;align-items:center;gap:6px;background:var(--accent)11;border:1px solid var(--accent)44;border-radius:6px;padding:4px 12px;font-size:12px;font-weight:600;color:var(--accent);margin-left:4px;">
    <span id="in-progress-count">0</span> In Progress
  </span>
</div>

<div class="stats">
  <div class="stat">
    <div class="stat-label">Total Assigned <span class="info-btn" onclick="showInfo(event,'total-assigned')">?</span></div>
    <div class="stat-value" id="s-total">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">Curriculum Complete <span class="info-btn" onclick="showInfo(event,'curriculum-complete')">?</span></div>
    <div class="stat-value blue" id="s-curriculum">&#8212;</div>
    <div class="stat-sub" id="s-curriculum-sub"></div>
  </div>
  <div class="stat">
    <div class="stat-label">Certified <span class="info-btn" onclick="showInfo(event,'certified')">?</span></div>
    <div class="stat-value green" id="s-certified">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">Not Yet Certified <span class="info-btn" onclick="showInfo(event,'not-certified')">?</span></div>
    <div class="stat-value red" id="s-not">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">Completion Rate <span class="info-btn" onclick="showInfo(event,'completion-rate')">?</span></div>
    <div class="stat-value teal" id="s-rate">&#8212;</div>
    <div class="stat-sub" id="s-rate-sub"></div>
  </div>
</div>

<div class="charts">
  <div class="chart-card">
    <div class="chart-title">Certification Pipeline <span class="info-btn" onclick="showInfo(event,'pipeline')">?</span></div>
    <div class="chart-wrap"><canvas id="regionChart"></canvas></div>
  </div>
  <div class="chart-card">
    <div class="chart-title">Certifications Over Time <span class="info-btn" onclick="showInfo(event,'over-time')">?</span></div>
    <div class="chart-wrap"><canvas id="trendChart"></canvas></div>
  </div>
  <div class="chart-card" style="grid-column:1/-1;">
    <div class="chart-title">Learners by Market <span class="info-btn" onclick="showInfo(event,'market')">?</span></div>
    <div class="chart-wrap" style="height:220px;"><canvas id="marketChart"></canvas></div>
  </div>
</div>

<div class="section">
  <div class="section-header">
    <div>
      <div class="section-title">Certification Roster <span class="info-btn" onclick="showInfo(event,'roster')">?</span></div>
      <div class="section-hint">Click a person to see their details &mdash; sorted by status then name</div>
    </div>
    <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;">
      <span style="font-size:11px;color:var(--muted);margin-right:2px;">View:</span>
      <button class="sort-btn active" id="view-individual" onclick="setRosterView('individual')">Individual</button>
      <button class="sort-btn" id="view-manager" onclick="setRosterView('manager')">By Manager</button>
      <span id="sort-controls" style="display:flex;align-items:center;gap:6px;margin-left:6px;">
        <span style="font-size:11px;color:var(--muted);margin-right:2px;">Sort:</span>
        <button class="sort-btn active" data-sort="status" data-label="Status" onclick="setRosterSort('status')">Status ↓</button>
        <button class="sort-btn" data-sort="name"   data-label="Name"   onclick="setRosterSort('name')">Name</button>
        <button class="sort-btn" data-sort="market" data-label="Market" onclick="setRosterSort('market')">Market</button>
      </span>
    </div>
    <input type="text" class="roster-search" id="roster-search" placeholder="Search by name&hellip;" oninput="filterRoster()">
  </div>
  <div class="roster-wrap">
    <div class="roster-left" id="roster-left"></div>
    <div class="roster-right" id="roster-right">
      <div class="no-data">Select a person to see their details</div>
    </div>
  </div>
</div>

<script>
const RAW = {raw_json};
const TLG_SET = new Set({tlg_json});

let filtered = [];

let regionChart, trendChart, marketChart;
let rosterSortField = 'status';
let rosterSortDir   = 'desc';
let rosterView      = 'individual';

function sel(id){{ return document.getElementById(id); }}
function cv(v){{ return getComputedStyle(document.body).getPropertyValue(v).trim(); }}

// Theme
(function(){{
  if(localStorage.getItem('pb-theme')==='light') document.body.classList.add('light-mode');
  sel('btn-theme').textContent = document.body.classList.contains('light-mode') ? '🌙 Dark' : '☀ Light';
}})();
function toggleTheme(){{
  const light = document.body.classList.toggle('light-mode');
  localStorage.setItem('pb-theme', light ? 'light' : 'dark');
  sel('btn-theme').textContent = light ? '🌙 Dark' : '☀ Light';
  applyFilters();
}}
function toggleExportDrop() {{
  sel('export-menu').classList.toggle('open');
}}
document.addEventListener('click', function(e) {{
  const d=sel('export-drop'); if(d&&!d.contains(e.target)) sel('export-menu').classList.remove('open');
}});
function setupPrintHeader(title,subtitle) {{
  sel('ph-title').textContent=title;
  sel('ph-date').textContent=subtitle;
  const statusLabel=sel('f-status').options[sel('f-status').selectedIndex].text;
  const market=sel('f-market').value||'All Markets';
  const from=sel('f-date-from').value||'',to=sel('f-date-to').value||'';
  const parts=[`Status: ${{statusLabel}}`,`Market: ${{market}}`];
  if(from||to) parts.push(`Dates: ${{from||'-'}} to ${{to||'-'}}`);
  if(hideTLG) parts.push('TLG hidden');
  sel('ph-filters').textContent=parts.join('  |  ');
}}
function doPrint(noSummary) {{
  if(noSummary) document.body.classList.add('print-no-summary');
  window.onafterprint=function() {{
    document.body.classList.remove('print-no-summary');
    sel('print-s1-label').style.display='none';
    sel('print-s2').style.display='none';
    sel('ph-desc').style.display='none';
    window.onafterprint=null;
  }};
  window.print();
}}
function tds(cells) {{ return '<tr>'+cells.map(c=>`<td>${{c}}</td>`).join('')+'</tr>'; }}
function thRow(labels) {{ return '<tr>'+labels.map(l=>`<th>${{l}}</th>`).join('')+'</tr>'; }}
function pBox(n,l) {{
  return '<div style="min-width:90px"><div style="font-size:30px;font-weight:700;color:#1a3a5c;line-height:1">'+n+'</div>'
        +'<div style="font-size:10px;color:#666;text-transform:uppercase;letter-spacing:.06em;margin-top:5px">'+l+'</div></div>';
}}
function pSection(title,rows) {{
  var h='<div style="min-width:150px"><div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#666;margin-bottom:10px">'+title+'</div>';
  rows.forEach(function(r){{h+='<div style="display:flex;justify-content:space-between;gap:32px;padding:4px 0;border-bottom:1px solid #eee;font-size:12px"><span>'+r[0]+'</span><strong>'+r[1]+'</strong></div>';}});
  return h+'</div>';
}}
function pFiscalQtr(d) {{
  if(!d) return null;
  var pts=d.split('-'),yr=+pts[0],mo=+pts[1];
  var fy=mo>=4?yr+1:yr,q=mo>=10?3:mo>=7?2:mo>=4?1:4;
  return 'Q'+q+' FY'+String(fy).slice(2);
}}
function runExport(type) {{
  sel('export-menu').classList.remove('open');
  const certField='Healthcare';
  const now=new Date().toLocaleDateString('en-US',{{year:'numeric',month:'long',day:'numeric'}});
  sel('print-s1-label').style.display='none';
  sel('print-s2').style.display='none';

  if(type==='full') {{
    setupPrintHeader('{name} Certification Report',`Generated: ${{now}}  |  ${{filtered.length}} People`);
    var pTotal=filtered.length,pCert=filtered.filter(p=>p.Healthcare==='Yes').length;
    var pLMS=filtered.filter(p=>p.Complete==='Yes'&&p.Healthcare!=='Yes').length;
    var pIP=filtered.filter(p=>p.Complete!=='Yes'&&p.Healthcare!=='Yes').length;
    var pRate=pTotal>0?Math.round(pCert/pTotal*100):0;
    sel('print-stats').innerHTML=pBox(pTotal,'Total Assigned')+pBox(pCert,'Certified')+pBox(pLMS,'LMS Complete')+pBox(pIP,'In Progress')+pBox(pRate+'%','Completion Rate');
    var qMap={{}};
    filtered.filter(p=>p.Healthcare==='Yes'&&p.HCDate).forEach(p=>{{var q=pFiscalQtr(p.HCDate);if(q)qMap[q]=(qMap[q]||0)+1;}});
    var qRows=Object.entries(qMap).sort((a,b)=>a[0].localeCompare(b[0]));
    if(!qRows.length) qRows=[['No data','-']];
    sel('print-charts').innerHTML=pSection('Certification Pipeline',[['In Progress',pIP],['LMS Complete',pLMS],['Certified',pCert]])+pSection('Certifications by Quarter',qRows);
    sel('print-roster-head').innerHTML=thRow(['#','Name','Market','Job Title','Status','Cert Date','Manager','Email']);
    sel('print-roster-body').innerHTML=filtered.map((p,i)=>{{
      const status=p.Healthcare==='Yes'?'Certified':p.Complete==='Yes'?'LMS Complete':'In Progress';
      return tds([i+1,`<b>${{p.FirstName}} ${{p.LastName}}</b>`,p.Market||'-',p.JobTitle||'-',status,p.HCDate||'-',p.Manager||'-',`<small>${{p.Email||'-'}}</small>`]);
    }}).join('');
    doPrint(false);

  }} else if(type==='not-certified') {{
    const notCert=filtered.filter(p=>p.Healthcare!=='Yes');
    setupPrintHeader('Not Certified: {name}',`Generated: ${{now}}  |  ${{notCert.length}} Employees`);
    sel('print-roster-head').innerHTML=thRow(['#','Name','Job Title','Market','Email','Manager','Manager Email']);
    sel('print-roster-body').innerHTML=notCert.length
      ? notCert.sort((a,b)=>(a.Manager||'').localeCompare(b.Manager||'')||(a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName))
          .map((p,i)=>tds([i+1,`<b>${{p.FirstName}} ${{p.LastName}}</b>`,p.JobTitle||'-',p.Market||'-',p.Email||'-',p.Manager||'-',p.MgrEmail||'-'])).join('')
      : '<tr><td colspan="7" style="color:#999;font-style:italic;padding:10px 8px">All employees are certified</td></tr>';
    sel('ph-desc').textContent='Employees who have not yet earned Healthcare foundational certification, sorted by manager. Use this list to contact employees and their managers to drive completion.';
    sel('ph-desc').style.display='block';
    doPrint(true);

  }} else if(type==='manager-summary') {{
    const mgrMap={{}};
    filtered.forEach(p=>{{
      const k=p.Manager||'(No Manager)';
      if(!mgrMap[k]) mgrMap[k]={{name:k,email:p.MgrEmail||'-',total:0,cert:0,lms:0,ip:0}};
      mgrMap[k].total++;
      if(p.Healthcare==='Yes') mgrMap[k].cert++;
      else if(p.Complete==='Yes') mgrMap[k].lms++;
      else mgrMap[k].ip++;
    }});
    const mgrs=Object.values(mgrMap).sort((a,b)=>(b.cert/b.total)-(a.cert/a.total));
    setupPrintHeader('Manager Summary: {name}',`Generated: ${{now}}  |  ${{mgrs.length}} Managers`);
    sel('print-roster-head').innerHTML=thRow(['Manager','Manager Email','Team Size','Certified','LMS Complete','In Progress','Completion %']);
    sel('print-roster-body').innerHTML=mgrs.map(m=>tds([
      `<b>${{m.name}}</b>`,m.email,m.total,m.cert,m.lms,m.ip,`<b>${{Math.round(m.cert/m.total*100)}}%</b>`
    ])).join('');
    sel('ph-desc').textContent='Certification completion rates by manager, sorted from lowest to highest. Managers at the top of the list have the most employees still in progress and need the most follow-up.';
    sel('ph-desc').style.display='block';
    doPrint(true);

  }} else if(type==='action-required') {{
    const almostThere=filtered.filter(p=>p.Complete==='Yes'&&p.Healthcare!=='Yes');
    setupPrintHeader('Action Required: {name}',`Generated: ${{now}}  |  ${{almostThere.length}} Employees Need Follow-Up`);
    sel('ph-desc').textContent='Employees who have completed the LMS curriculum but have not yet received manager-confirmed certification. Follow up to confirm status or escalate to completion.';
    sel('ph-desc').style.display='block';
    sel('print-s1-label').innerHTML=`LMS COMPLETE: AWAITING CERTIFICATION (${{almostThere.length}})<div style="font-weight:400;text-transform:none;letter-spacing:0;font-size:10px;color:#555;font-style:italic;margin-top:4px">These employees have completed the LMS curriculum but have not yet been certified. Follow up to confirm status or escalate to completion.</div>`;
    sel('print-s1-label').style.display='block';
    sel('print-roster-head').innerHTML=thRow(['#','Name','Email','Market','Manager','Manager Email']);
    sel('print-roster-body').innerHTML=almostThere.length
      ? almostThere.sort((a,b)=>(a.Manager||'').localeCompare(b.Manager||''))
          .map((p,i)=>tds([i+1,`<b>${{p.FirstName}} ${{p.LastName}}</b>`,p.Email||'-',p.Market||'-',p.Manager||'-',p.MgrEmail||'-'])).join('')
      : '<tr><td colspan="6" style="color:#999;font-style:italic;padding:10px 8px">No employees in this category</td></tr>';
    doPrint(true);
  }}
}}

// Populate filter dropdowns from data
const allMarkets = [...new Set(RAW.map(r=>r.Market).filter(Boolean))].sort();
allMarkets.forEach(m => sel('f-market').innerHTML += `<option value="${{m}}">${{m}}</option>`);

['f-status','f-market','f-date-from','f-date-to'].forEach(id => {{
  sel(id).addEventListener('change', applyFilters);
}});

document.addEventListener('click', function(e){{
  if(!e.target.classList.contains('info-btn')){{
    sel('info-popover').classList.remove('visible');
  }}
}});

const INFO_MSGS = {{
  'total-assigned':      'Total number of people currently assigned this certification track, after any active filters.',
  'curriculum-complete': 'People who have finished all required LMS coursework (Healthcare foundational curriculum). Completing the LMS is the first step; manager-confirmed certification is a separate process that follows.',
  'certified':           'People who have received full certification, confirmed by their manager after completing the LMS curriculum. This is the final step in the primary certification track.',
  'not-certified':       'People assigned the curriculum who have not yet received manager-confirmed certification. Includes both people still working through the LMS and those who have finished the LMS but are awaiting manager sign-off.',
  'completion-rate':     'Percentage of assigned people who have earned full manager-confirmed certification. Calculated as Certified ÷ Total Assigned.',
  'pipeline':            'The three stages of the certification journey: In Progress (enrolled but LMS coursework not yet complete), Curriculum Complete (LMS done, awaiting manager sign-off), and Certified (fully certified by manager).',
  'over-time':           'Certifications earned per KM fiscal quarter, stacked by curriculum type. KM quarters: Q1 = Apr–Jun, Q2 = Jul–Sep, Q3 = Oct–Dec, Q4 = Jan–Mar.',
  'market':              'Distribution of assigned learners by sales market. Hover a segment to see the exact count and percentage. Reflects current filters.',
  'roster':              'Full list of assigned people with their certification status. Click a name to see job title, market, LMS status, cert date, and manager info. Use the View toggle to group by manager.',
  'export':              'Export a printable report of the data currently on screen. Apply filters first; the report only includes what is currently shown. Full Report lists everyone with full detail. Not Certified is a contact list for outreach sorted by manager. Manager Summary rolls up team count and completion % per manager. Action Required lists people who have finished the LMS but are still awaiting certification.',
}};
function showInfo(e, key){{
  const pop = sel('info-popover');
  pop.textContent = INFO_MSGS[key] || '';
  pop.classList.add('visible');
  const r = e.target.getBoundingClientRect();
  pop.style.top  = (r.bottom + 6) + 'px';
  pop.style.left = Math.min(r.left, window.innerWidth - 280) + 'px';
  e.stopPropagation();
}}

function toggleTLG(){{
  hideTLG = !hideTLG;
  sel('btn-tlg').classList.toggle('active', hideTLG);
  sel('btn-tlg').textContent = hideTLG ? 'Show TLG' : 'Hide TLG';
  applyFilters();
}}

function resetFilters(){{
  ['f-status','f-market','f-date-from','f-date-to'].forEach(id => sel(id).value = '');
  if(!hideTLG){{ hideTLG=true; sel('btn-tlg').classList.add('active'); sel('btn-tlg').textContent='Show TLG'; }}
  applyFilters();
}}

function applyFilters(){{
  const status = sel('f-status').value;
  const market = sel('f-market').value;
  const from   = sel('f-date-from').value;
  const to     = sel('f-date-to').value;
  filtered = RAW.filter(r => {{
    if(hideTLG && TLG_SET.has(r.FirstName+' '+r.LastName)) return false;
    if(status && r.Healthcare !== status) return false;
    if(market && r.Market !== market) return false;
    if(from && r.Date && r.Date < from) return false;
    if(to   && r.Date && r.Date > to)   return false;
    return true;
  }});
  sel('result-count').textContent = `${{filtered.length}} People`;
  render();
}}

function render(){{
  const isLight    = document.body.classList.contains('light-mode');
  const chartLabel = isLight ? cv('--text') : cv('--muted');

  // Stat cards
  const total          = filtered.length;
  const curriculumDone = filtered.filter(r=>r.Complete==='Yes').length;
  const certified      = filtered.filter(r=>r.Healthcare==='Yes').length;
  const notCert        = total - certified;
  const rate           = total > 0 ? Math.round(certified/total*100) : 0;

  sel('s-total').textContent          = total;
  sel('s-curriculum').textContent     = curriculumDone;
  sel('s-curriculum-sub').textContent = total > 0 ? `${{curriculumDone}} of ${{total}} assigned` : '';
  sel('s-certified').textContent      = certified;
  sel('s-not').textContent            = notCert;
  sel('s-rate').textContent           = rate + '%';
  sel('s-rate-sub').textContent       = total > 0 ? `${{certified}} of ${{total}} assigned` : '';

  // Pipeline chart — three stages of the certification journey
  const pipelineNotStarted   = filtered.filter(r=>r.Complete!=='Yes'&&r.Healthcare!=='Yes').length;
  const pipelineCurrComplete = filtered.filter(r=>r.Complete==='Yes'&&r.Healthcare!=='Yes').length;
  const pipelineHCCertified  = filtered.filter(r=>r.Healthcare==='Yes').length;

  const ipBox = sel('in-progress-box');
  sel('in-progress-count').textContent = pipelineNotStarted;
  ipBox.style.display = pipelineNotStarted > 0 ? 'inline-flex' : 'none';

  if(regionChart) regionChart.destroy();
  regionChart = new Chart(sel('regionChart'), {{
    type: 'bar',
    data: {{
      labels: ['In Progress','Curriculum Complete','Certified'],
      datasets: [{{
        data: [pipelineNotStarted, pipelineCurrComplete, pipelineHCCertified],
        backgroundColor: [cv('--red')+'cc', cv('--accent')+'cc', cv('--green')+'cc'],
        borderRadius: 4,
        borderSkipped: false,
      }}]
    }},
    options: {{
      indexAxis: 'y',
      responsive:true, maintainAspectRatio:false,
      plugins:{{
        legend:{{display:false}},
        tooltip:{{
          callbacks:{{
            label: ctx => ` ${{ctx.raw}} People`
          }}
        }},
        datalabels:{{display:false}}
      }},
      scales:{{
        x:{{grid:{{color:cv('--border')}},ticks:{{color:chartLabel,font:{{size:11}},stepSize:1}}}},
        y:{{grid:{{display:false}},ticks:{{color:chartLabel,font:{{size:12}}}}}}
      }}
    }}
  }});

  // Trend chart — certifications per fiscal quarter
  const parseQtr = s => {{ const m=s.match(/FY(\d+)\s+Q(\d)/); return m ? parseInt(m[1])*10+parseInt(m[2]) : 0; }};
  const allTrendQtrs = new Set();
  filtered.forEach(r => {{ if(r.Healthcare==='Yes' && r.HCQtr) allTrendQtrs.add(r.HCQtr); }});
  const trendQtrs = [...allTrendQtrs].sort((a,b)=>parseQtr(a)-parseQtr(b));

  if(trendChart) trendChart.destroy();
  trendChart = new Chart(sel('trendChart'), {{
    type: 'bar',
    data: {{
      labels: trendQtrs,
      datasets: [{{
        label: 'Healthcare foundational',
        data:  trendQtrs.map(q => filtered.filter(r=>r.Healthcare==='Yes'&&r.HCQtr===q).length),
        backgroundColor: cv('--green')+'bb',
        borderRadius: 3,
        borderSkipped: false,
      }}]
    }},
    options: {{
      responsive:true, maintainAspectRatio:false,
      plugins:{{
        legend:{{display:false}},
        tooltip:{{callbacks:{{label: ctx => ` ${{ctx.raw}} Certified`}}}},
        datalabels:{{display:false}}
      }},
      scales:{{
        x:{{grid:{{color:cv('--border')}},ticks:{{color:chartLabel,font:{{size:10}},maxRotation:45}}}},
        y:{{grid:{{color:cv('--border')}},ticks:{{color:chartLabel,font:{{size:11}},stepSize:1}}}}
      }}
    }}
  }});

  // Market donut chart
  const marketCounts = {{}};
  filtered.forEach(r => {{
    const m = r.Market || 'Unknown';
    marketCounts[m] = (marketCounts[m] || 0) + 1;
  }});
  const marketLabels = Object.keys(marketCounts).sort((a,b) => marketCounts[b] - marketCounts[a]);
  const marketData   = marketLabels.map(m => marketCounts[m]);
  const marketColors = ['#4a7cf7','#22c55e','#f59e0b','#a855f7','#ef4444','#06b6d4','#f97316','#64748b'];

  if(marketChart) marketChart.destroy();
  marketChart = new Chart(sel('marketChart'), {{
    type: 'doughnut',
    data: {{
      labels: marketLabels,
      datasets: [{{
        data: marketData,
        backgroundColor: marketColors.slice(0, marketLabels.length),
        borderWidth: 0,
        hoverOffset: 6,
      }}]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      layout: {{ padding: {{ top: 8, bottom: 8 }} }},
      plugins: {{
        legend: {{
          display: true,
          position: 'right',
          labels: {{ color: chartLabel, font: {{ size: 12 }}, padding: 14, boxWidth: 14 }}
        }},
        tooltip: {{
          callbacks: {{
            label: ctx => ` ${{ctx.raw}} (${{Math.round(ctx.raw / filtered.length * 100)}}%)`
          }}
        }},
        datalabels: {{ display: false }}
      }}
    }}
  }});

  renderRosterList();
}}

function setRosterSort(field){{
  if(rosterSortField === field) {{
    rosterSortDir = rosterSortDir === 'desc' ? 'asc' : 'desc';
  }} else {{
    rosterSortField = field;
    rosterSortDir   = 'desc';
  }}
  renderRosterList();
}}

function pipelineSteps(p) {{
  const amber = '#f59e0b';
  function dot(color, tip, check) {{
    const bg     = check ? color+'33' : 'var(--surface2)';
    const border = check ? color : 'var(--border)';
    const txt    = check ? color : 'var(--muted)';
    return `<div title="${{tip}}" style="width:16px;height:16px;border-radius:50%;flex-shrink:0;display:flex;align-items:center;justify-content:center;font-size:9px;font-weight:700;background:${{bg}};border:1.5px solid ${{border}};color:${{txt}}">${{check?'&#10003;':''}}</div>`;
  }}
  const certified = p.Healthcare==='Yes';
  const lmsColor  = certified ? cv('--green') : p.Complete==='Yes' ? amber : 'var(--border)';
  return `<div style="display:flex;align-items:center;gap:3px;flex-shrink:0;">
    ${{dot(lmsColor,'LMS Complete',p.Complete==='Yes')}}
    <div style="width:8px;height:1px;background:var(--border);flex-shrink:0;"></div>
    ${{dot(cv('--green'),'Certified',certified)}}
  </div>`;
}}

function tierStyle(p, extraIndent) {{
  extraIndent = extraIndent || 0;
  const base = 14 + extraIndent;
  if(p.Healthcare==='Yes') return `border-left:3px solid ${{cv('--green')}};padding-left:${{base-3}}px;`;
  if(p.Complete==='Yes')   return `border-left:3px solid #f59e0b;padding-left:${{base-3}}px;`;
  return extraIndent ? `padding-left:${{base}}px;` : '';
}}

function setRosterView(v) {{
  rosterView = v;
  sel('view-individual').classList.toggle('active', v==='individual');
  sel('view-manager').classList.toggle('active', v==='manager');
  sel('sort-controls').style.display = v==='individual' ? 'flex' : 'none';
  sel('roster-right').innerHTML = '<div class="no-data">Select a person to see their details</div>';
  renderRosterList();
}}

function toggleMgrGroup(el) {{
  el.classList.toggle('open');
  el.nextElementSibling.classList.toggle('open');
}}

function renderRosterList(){{
  const certField = 'Healthcare';

  if(rosterView === 'manager') {{
    const groups = {{}};
    filtered.forEach(p => {{
      const mgr = p.Manager || 'No Manager';
      if(!groups[mgr]) groups[mgr] = [];
      groups[mgr].push(p);
    }});
    const mgrsSorted = Object.keys(groups).sort((a,b) => groups[b].length - groups[a].length);
    sel('roster-left').innerHTML = mgrsSorted.map(mgr => {{
      const team = groups[mgr];
      const cert = team.filter(p=>p[certField]==='Yes').length;
      const lms  = team.filter(p=>p.Complete==='Yes').length;
      const peopleHtml = team
        .sort((a,b)=>(a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName))
        .map(p => {{
          const fullName = `${{p.FirstName}} ${{p.LastName}}`;
          return `<div class="roster-person" onclick="rosterSelect(this)" data-name="${{fullName}}" style="${{tierStyle(p,10)}}">
            <span class="roster-name">${{fullName}}</span>
            ${{pipelineSteps(p)}}
          </div>`;
        }}).join('');
      return `<div>
        <div class="mgr-group-hdr open" onclick="toggleMgrGroup(this)">
          <div>
            <div style="font-size:12px;font-weight:600;color:var(--text)">${{mgr}}</div>
            <div style="font-size:11px;color:var(--muted);margin-top:2px;">${{team.length}} rep${{team.length!==1?'s':''}} &middot; ${{lms}} LMS &middot; ${{cert}} certified</div>
          </div>
          <span class="mgr-chevron">&#9660;</span>
        </div>
        <div class="mgr-team open">${{peopleHtml}}</div>
      </div>`;
    }}).join('') || `<div class="no-data">No people match filters</div>`;
    filterRoster();
    return;
  }}

  // Individual view
  const d = rosterSortDir === 'desc' ? -1 : 1;
  const sorted = [...filtered].sort((a,b) => {{
    if(rosterSortField === 'name') {{
      return d * (a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName);
    }} else if(rosterSortField === 'market') {{
      const rc = d * (a.Market||'').localeCompare(b.Market||'');
      return rc !== 0 ? rc : (a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName);
    }} else {{
      function score(p) {{ return p.Healthcare==='Yes'?2:p.Complete==='Yes'?1:0; }}
      const diff = score(a) - score(b);
      if(diff !== 0) return d * diff;
      return (a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName);
    }}
  }});

  document.querySelectorAll('.sort-btn').forEach(btn => {{
    if(!btn.dataset.sort) return;
    const active = btn.dataset.sort === rosterSortField;
    btn.classList.toggle('active', active);
    const arrow = active ? (rosterSortDir === 'desc' ? ' ↓' : ' ↑') : '';
    btn.textContent = btn.dataset.label + arrow;
  }});

  sel('roster-left').innerHTML = sorted.map(p => {{
    const fullName = `${{p.FirstName}} ${{p.LastName}}`;
    return `<div class="roster-person" onclick="rosterSelect(this)" data-name="${{fullName}}" style="${{tierStyle(p)}}">
      <span class="roster-name">${{fullName}}</span>
      ${{pipelineSteps(p)}}
    </div>`;
  }}).join('') || `<div class="no-data">No people match filters</div>`;

  filterRoster();
  const first = sel('roster-left').querySelector('.roster-person:not([style*="none"])') ||
                sel('roster-left').querySelector('.roster-person');
  if(first) rosterSelect(first);
}}

function rosterSelect(el){{
  document.querySelectorAll('.roster-person').forEach(p=>p.classList.remove('active'));
  el.classList.add('active');
  const name = el.dataset.name;
  const p = filtered.find(r=>`${{r.FirstName}} ${{r.LastName}}`===name);
  if(!p) return;
  const isCert = p.Healthcare==='Yes';
  function fmtDate(d) {{
    if (!d) return '&#8212;';
    const [y,mo,dy] = d.split('-');
    const months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
    return months[parseInt(mo)-1] + ' ' + parseInt(dy) + ', ' + y;
  }}
  sel('roster-right').innerHTML = `
    <div class="roster-right-header">
      <div style="font-size:15px;font-weight:700;margin-bottom:6px">${{p.FirstName}} ${{p.LastName}}</div>
      <span class="badge-status ${{isCert?'certified':'not-certified'}}">${{isCert?'Certified':'Not Yet Certified'}}</span>
      ${{isCert && p.HCDate ? `<span style="font-size:12px;color:var(--muted);margin-left:8px">${{fmtDate(p.HCDate)}}</span>` : ''}}
    </div>
    <div class="detail-grid">
      <div><div class="detail-label">Job Title</div><div class="detail-value">${{p.JobTitle||'&#8212;'}}</div></div>
      <div><div class="detail-label">Market</div><div class="detail-value">${{p.Market||'&#8212;'}}</div></div>
      <div><div class="detail-label">Healthcare Foundational</div><div class="detail-value">${{p.Complete==='Yes' ? '&#10003; LMS Complete' : 'Not Complete'}}</div></div>
      <div><div class="detail-label">Hire Date</div><div class="detail-value">${{fmtDate(p.HireDate)}}</div></div>
      <div style="grid-column:1/-1"><div class="detail-label">Email</div><div class="detail-value"><a href="mailto:${{p.Email}}" style="color:var(--accent);text-decoration:none">${{p.Email||'&#8212;'}}</a></div></div>
    </div>
    ${{p.Manager ? `<div style="margin-top:16px;padding-top:14px;border-top:1px solid var(--border)">
      <div style="font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:10px">Manager</div>
      <div class="detail-grid">
        <div><div class="detail-label">Name</div><div class="detail-value">${{p.Manager}}</div></div>
        <div><div class="detail-label">Title</div><div class="detail-value">${{p.MgrTitle||'&#8212;'}}</div></div>
        <div><div class="detail-label">Email</div><div class="detail-value"><a href="mailto:${{p.MgrEmail}}" style="color:var(--accent);text-decoration:none">${{p.MgrEmail||'&#8212;'}}</a></div></div>
      </div>
    </div>` : ''}}
  `;
}}

function filterRoster(){{
  const q = (sel('roster-search')?.value||'').toLowerCase();
  document.querySelectorAll('.roster-person').forEach(el=>{{
    el.style.display = (!q||el.dataset.name.toLowerCase().includes(q)) ? '' : 'none';
  }});
}}

function runExportXLSX(type){{
  sel('export-menu').classList.remove('open');
  const now=new Date().toLocaleDateString('en-US',{{year:'numeric',month:'long',day:'numeric'}});
  function makeSheet(rows,colWidths){{
    const ws=XLSX.utils.aoa_to_sheet(rows);
    ws['!cols']=colWidths.map(w=>({{wch:w}}));
    return ws;
  }}
  function dlXLSX(name,wb){{ XLSX.writeFile(wb,name+'.xlsx'); }}
  const wb=XLSX.utils.book_new();
  if(type==='full'){{
    const pTotal=filtered.length;
    const pCert=filtered.filter(p=>p.Healthcare==='Yes').length;
    const pLMS=filtered.filter(p=>p.Complete==='Yes'&&p.Healthcare!=='Yes').length;
    const pIP=pTotal-pCert-pLMS;
    const pRate=pTotal>0?Math.round(pCert/pTotal*100):0;
    const sumRows=[
      ['Healthcare Certification Report'],
      ['Generated:',now],
      [],
      ['SUMMARY'],
      ['Total Assigned',pTotal],
      ['Certified',pCert],
      ['LMS Complete',pLMS],
      ['In Progress',pIP],
      ['Completion Rate',pRate+'%'],
    ];
    XLSX.utils.book_append_sheet(wb,makeSheet(sumRows,[30,18]),'Summary');
    const rRows=[['Name','Market','Job Title','Status','Cert Date','Manager','Email']];
    filtered.forEach(p=>{{
      const status=p.Healthcare==='Yes'?'Certified':p.Complete==='Yes'?'LMS Complete':'In Progress';
      rRows.push([p.FirstName+' '+p.LastName,p.Market||'-',p.JobTitle||'-',status,p.HCDate||'-',p.Manager||'-',p.Email||'-']);
    }});
    XLSX.utils.book_append_sheet(wb,makeSheet(rRows,[28,18,28,14,14,28,32]),'Roster');
    dlXLSX('hc-full-report',wb);
  }} else if(type==='not-certified'){{
    const notCert=filtered.filter(p=>p.Healthcare!=='Yes').sort((a,b)=>(a.Manager||'').localeCompare(b.Manager||'')||(a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName));
    const rows=[['Name','Job Title','Market','Email','Manager','Manager Email']];
    notCert.forEach(p=>rows.push([p.FirstName+' '+p.LastName,p.JobTitle||'-',p.Market||'-',p.Email||'-',p.Manager||'-',p.MgrEmail||'-']));
    XLSX.utils.book_append_sheet(wb,makeSheet(rows,[28,28,18,32,28,32]),'Not Certified');
    dlXLSX('hc-not-certified',wb);
  }} else if(type==='manager-summary'){{
    const mgrMap={{}};
    filtered.forEach(p=>{{
      const k=p.Manager||'(No Manager)';
      if(!mgrMap[k]) mgrMap[k]={{name:k,email:p.MgrEmail||'-',total:0,cert:0,lms:0,ip:0}};
      mgrMap[k].total++;
      if(p.Healthcare==='Yes') mgrMap[k].cert++;
      else if(p.Complete==='Yes') mgrMap[k].lms++;
      else mgrMap[k].ip++;
    }});
    const rows=[['Manager','Manager Email','Team Size','Certified','LMS Complete','In Progress','Completion %']];
    Object.values(mgrMap).sort((a,b)=>(b.cert/b.total)-(a.cert/a.total)).forEach(m=>{{
      rows.push([m.name,m.email,m.total,m.cert,m.lms,m.ip,Math.round(m.cert/m.total*100)+'%']);
    }});
    XLSX.utils.book_append_sheet(wb,makeSheet(rows,[28,32,12,12,14,12,14]),'Manager Summary');
    dlXLSX('hc-manager-summary',wb);
  }} else if(type==='action-required'){{
    const almostThere=filtered.filter(p=>p.Complete==='Yes'&&p.Healthcare!=='Yes').sort((a,b)=>(a.Manager||'').localeCompare(b.Manager||''));
    const rows=[['Name','Email','Market','Manager','Manager Email']];
    almostThere.forEach(p=>rows.push([p.FirstName+' '+p.LastName,p.Email||'-',p.Market||'-',p.Manager||'-',p.MgrEmail||'-']));
    XLSX.utils.book_append_sheet(wb,makeSheet(rows,[28,32,18,28,32]),'Action Required');
    dlXLSX('hc-action-required',wb);
  }}
}}

applyFilters();
(function(){{
  var n=0,t;
  var h=document.querySelector('.header h1');
  if(h) h.addEventListener('click',function(){{
    n++;clearTimeout(t);
    if(n>=3){{n=0;window.location.href='index.html';}}
    else t=setTimeout(function(){{n=0;}},1500);
  }});
}})();
</script>
<div class="info-popover" id="info-popover"></div>
<!-- Print-only elements (hidden until runExport()) -->
<div id="print-header" style="display:none;margin-bottom:20px;">
  <div style="font-size:20px;font-weight:700;margin-bottom:4px;" id="ph-title"></div>
  <div style="font-size:12px;color:#555;margin-bottom:2px;" id="ph-date"></div>
  <div style="font-size:12px;color:#555;" id="ph-filters"></div>
  <div id="ph-desc" style="display:none;margin-top:10px;padding-top:10px;border-top:1px solid #dde4f0;font-size:12px;color:#444;font-style:italic;"></div>
</div>
<div id="print-stats" style="display:none;gap:40px;flex-wrap:wrap;margin-bottom:18px;padding-bottom:16px;border-bottom:2px solid #dde4f0;"></div>
<div id="print-charts" style="display:none;gap:48px;flex-wrap:wrap;margin-bottom:18px;padding-bottom:16px;border-bottom:2px solid #dde4f0;"></div>
<div id="print-roster-wrap" style="display:none;">
  <div id="print-s1-label" style="display:none;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#555;margin-bottom:10px;padding-bottom:6px;border-bottom:2px solid #1a73e8;"></div>
  <table class="ptable" id="print-roster-table">
    <thead id="print-roster-head"><tr>
      <th>#</th><th>Name</th><th>Market</th><th>Job Title</th>
      <th>Status</th><th>Cert Date</th><th>Manager</th><th>Email</th>
    </tr></thead>
    <tbody id="print-roster-body"></tbody>
  </table>
</div>
</body>
</html>"""


def generate_html_publicsector(slug, name, rows, date_label=''):
    raw_json = json.dumps(rows)
    tlg_json = json.dumps(sorted(TLG))

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{name} Certification Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<style>
  :root {{
    --bg:#0f1117; --surface:#1a1d27; --surface2:#22263a; --border:#2e3350;
    --accent:#4f8ef7; --accent2:#7c5cfc; --accent3:#f7c94f;
    --text:#e8ecf4; --muted:#7b82a0; --green:#3ecf8e; --red:#f76f6f;
    --teal:#2dd4bf; --green-subtle:#3ecf8e22; --red-subtle:#f76f6f22;
    --font:'Segoe UI',system-ui,sans-serif;
  }}
  body.light-mode {{
    --bg:#f4f6fb; --surface:#ffffff; --surface2:#eef1f7; --border:#d0d7e8;
    --accent:#2563eb; --accent2:#6d28d9; --accent3:#d97706;
    --text:#1a1d27; --muted:#475569; --green:#059669; --red:#dc2626;
    --teal:#0f766e; --green-subtle:#05966922; --red-subtle:#dc262622;
  }}
  body.light-mode select,body.light-mode input[type=date]{{color-scheme:light;}}
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{background:var(--bg);color:var(--text);font-family:var(--font);min-height:100vh;transition:background .2s,color .2s;}}
  .header{{padding:20px 28px 16px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px;}}
  .header-left{{display:flex;align-items:center;gap:16px;}}
  .info-btn{{display:inline-flex;align-items:center;justify-content:center;width:15px;height:15px;border-radius:50%;background:var(--surface2);border:1px solid var(--border);color:var(--muted);font-size:9px;font-weight:700;cursor:pointer;margin-left:5px;vertical-align:middle;flex-shrink:0;line-height:1;transition:border-color .15s,color .15s;}}
  .info-btn:hover{{border-color:var(--accent);color:var(--accent);}}
  .info-popover{{position:fixed;z-index:9999;background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:12px 14px;font-size:12px;color:var(--text);line-height:1.6;max-width:260px;box-shadow:0 4px 24px rgba(0,0,0,0.5);display:none;}}
  .info-popover.visible{{display:block;}}
  .header h1{{font-size:18px;font-weight:700;letter-spacing:.3px;}}
  .header h1 span{{color:var(--muted);font-weight:400;}}
  .header-date{{font-size:11px;color:var(--muted);margin-top:2px;}}
  .btn-theme{{background:transparent;border:1px solid var(--border);color:var(--muted);border-radius:6px;padding:5px 12px;font-size:12px;cursor:pointer;transition:all .15s;}}
  .btn-theme:hover{{border-color:var(--accent);color:var(--text);}}
  .btn-export{{background:var(--accent);border:1px solid var(--accent);color:#fff;border-radius:6px;padding:5px 12px;font-size:12px;cursor:pointer;transition:all .15s;font-weight:600;}}
  .btn-export:hover{{opacity:0.88;}}
  .export-drop{{position:relative;}}
  .export-menu{{position:absolute;top:calc(100% + 6px);right:0;background:var(--surface);border:1px solid var(--border);border-radius:8px;min-width:210px;box-shadow:0 4px 24px rgba(0,0,0,.28);display:none;z-index:200;overflow:visible;}}
  .export-menu.open{{display:block;}}
  .export-item{{display:block;width:100%;text-align:left;padding:10px 14px;font-size:13px;color:var(--text);background:transparent;border:none;cursor:pointer;transition:background .1s;font-family:inherit;}}
  .export-item:hover{{background:var(--surface2);}}
  .export-parent{{position:relative;display:flex;justify-content:space-between;align-items:center;padding:10px 14px;font-size:13px;color:var(--text);cursor:default;transition:background .1s;}}
  .export-parent:hover{{background:var(--surface2);}}
  .export-chevron{{font-size:11px;color:var(--muted);margin-left:10px;}}
  .export-submenu{{position:absolute;right:100%;top:0;background:var(--surface);border:1px solid var(--border);border-radius:8px;min-width:90px;box-shadow:0 4px 24px rgba(0,0,0,0.28);display:none;z-index:201;overflow:hidden;margin-right:4px;}}
  .export-parent:hover .export-submenu{{display:block;}}
  .filters{{padding:14px 28px;display:flex;gap:10px;flex-wrap:wrap;align-items:center;border-bottom:1px solid var(--border);background:var(--surface);}}
  .filter-label{{font-size:12px;color:var(--muted);text-transform:uppercase;letter-spacing:.6px;margin-right:4px;}}
  select,input[type=date],input[type=text]{{background:var(--surface2);border:1px solid var(--border);color:var(--text);border-radius:6px;padding:6px 10px;font-size:13px;cursor:pointer;outline:none;color-scheme:dark;}}
  select:focus,input:focus{{border-color:var(--accent);}}
  .btn-reset{{background:transparent;border:1px solid var(--border);color:var(--muted);border-radius:6px;padding:6px 14px;font-size:12px;cursor:pointer;transition:border-color .15s,color .15s;}}
  .btn-reset:hover{{border-color:var(--accent);color:var(--text);}}
  .result-count{{margin-left:auto;font-size:12px;color:var(--muted);}}
  .stats{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;padding:20px 28px;}}
  .stat{{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:16px 18px;}}
  .stat-label{{font-size:11px;text-transform:uppercase;letter-spacing:.7px;color:var(--muted);margin-bottom:6px;}}
  .stat-value{{font-size:28px;font-weight:700;line-height:1;}}
  .stat-value.green{{color:var(--green);}}
  .stat-value.red{{color:var(--red);}}
  .stat-value.teal{{color:var(--teal);}}
  .stat-sub{{font-size:11px;color:var(--muted);margin-top:4px;}}
  .charts{{display:grid;grid-template-columns:1fr 1fr;gap:16px;padding:0 28px 16px;}}
  @media(max-width:900px){{.charts{{grid-template-columns:1fr;}}}}
  @media(max-width:480px){{.chart-wrap{{height:180px;}}}}
  .chart-card{{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:18px;}}
  .chart-title{{font-size:13px;font-weight:600;margin-bottom:14px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;}}
  body.light-mode .chart-title{{color:var(--text);}}
  .chart-wrap{{position:relative;height:260px;}}
  .section{{padding:0 28px 32px;}}
  .section-header{{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;flex-wrap:wrap;gap:8px;}}
  .section-title{{font-size:13px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;}}
  body.light-mode .section-title{{color:var(--text);}}
  .section-hint{{font-size:11px;color:var(--muted);margin-top:3px;}}
  .roster-search{{width:180px;}}
  .roster-wrap{{display:flex;border:1px solid var(--border);border-radius:10px;overflow:hidden;}}
  .roster-left{{width:300px;flex-shrink:0;overflow-y:auto;max-height:820px;border-right:1px solid var(--border);}}
  .roster-person{{display:flex;align-items:center;gap:10px;padding:10px 14px;cursor:pointer;border-bottom:1px solid var(--border);transition:background .1s;}}
  .roster-person:last-child{{border-bottom:none;}}
  .roster-person:hover{{background:var(--surface2);}}
  .roster-person.active{{background:#4f8ef711;border-left:3px solid var(--accent);padding-left:11px;}}
  .roster-name{{flex:1;font-size:13px;}}
  .cert-badge{{font-size:11px;font-weight:700;border-radius:10px;padding:2px 8px;white-space:nowrap;}}
  .cert-badge.yes{{color:var(--green);background:var(--green-subtle);}}
  .cert-badge.no{{color:var(--red);background:var(--red-subtle);}}
  .roster-right{{flex:1;overflow-y:auto;max-height:820px;padding:16px 20px;}}
  .roster-right-header{{margin-bottom:14px;padding-bottom:12px;border-bottom:1px solid var(--border);}}
  .no-data{{text-align:center;color:var(--muted);padding:40px;font-size:13px;}}
  @media(max-width:680px){{.roster-wrap{{flex-direction:column;}}.roster-left{{width:100%;max-height:220px;border-right:none;border-bottom:1px solid var(--border);}}}}
  .sort-btn{{background:transparent;border:1px solid var(--border);color:var(--muted);border-radius:6px;padding:4px 10px;font-size:11px;cursor:pointer;transition:all .15s;white-space:nowrap;}}
  .sort-btn:hover{{border-color:var(--accent);color:var(--text);}}
  .sort-btn.active{{border-color:var(--accent);color:var(--accent);background:var(--accent)11;}}
  .mgr-group-hdr{{display:flex;align-items:center;justify-content:space-between;padding:10px 14px;cursor:pointer;background:var(--surface2);border-bottom:1px solid var(--border);user-select:none;}}
  .mgr-group-hdr:hover{{background:var(--border);}}
  .mgr-group-hdr.open .mgr-chevron{{transform:rotate(180deg);}}
  .mgr-chevron{{transition:transform .15s;font-size:10px;color:var(--muted);flex-shrink:0;}}
  .mgr-team{{display:none;}}
  .mgr-team.open{{display:block;}}
  .detail-grid{{display:grid;grid-template-columns:1fr 1fr;gap:12px 28px;margin-top:14px;}}
  .detail-label{{font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:3px;}}
  .detail-value{{font-size:14px;font-weight:500;}}
  .badge-status{{display:inline-block;padding:3px 12px;border-radius:20px;font-size:12px;font-weight:700;}}
  .badge-status.certified{{background:var(--green-subtle);color:var(--green);}}
  .badge-status.not-certified{{background:var(--red-subtle);color:var(--red);}}
  @page{{size:landscape;margin:.65in;}}
  @media print{{
    body{{background:#fff!important;color:#111!important;}}
    .header,.filters,.stats,.charts,.section,.print-hide{{display:none!important;}}
    #print-header{{display:block!important;}}
    #print-stats{{display:flex!important;gap:40px;flex-wrap:wrap;margin-bottom:18px;padding-bottom:16px;border-bottom:2px solid #dde4f0;}}
    #print-charts{{display:flex!important;gap:48px;flex-wrap:wrap;margin-bottom:18px;padding-bottom:16px;border-bottom:2px solid #dde4f0;}}
    #print-roster-wrap{{display:block!important;}}
    body.print-no-summary #print-stats,body.print-no-summary #print-charts{{display:none!important;}}
    .ptable{{width:100%;border-collapse:collapse;font-size:11px;}}
    .ptable th{{background:#f0f4ff;color:#111;font-weight:700;padding:5px 8px;border:1px solid #ccc;text-align:left;}}
    .ptable td{{padding:5px 8px;border:1px solid #ddd;vertical-align:middle;}}
    .ptable tr:nth-child(even) td{{background:#fafafa;}}
  }}
</style>
</head>
<body>

<div class="header">
  <div class="header-left">
    <div>
      <h1>{name} <span>Certification Dashboard</span></h1>
      <div class="header-date">Data as of {date_label}</div>
    </div>
  </div>
  <div style="display:flex;gap:8px;align-items:center;">
    <div class="export-drop print-hide" id="export-drop">
      <button class="btn-export" onclick="toggleExportDrop()">&#128438; Export &#9660;</button>
      <div class="export-menu" id="export-menu">
        <div class="export-parent">Full Report<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('full')">PDF</button><button class="export-item" onclick="runExportXLSX('full')">Excel</button></div></div>
        <div class="export-parent">Not Certified<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('not-certified')">PDF</button><button class="export-item" onclick="runExportXLSX('not-certified')">Excel</button></div></div>
        <div class="export-parent">Manager Summary<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('manager-summary')">PDF</button><button class="export-item" onclick="runExportXLSX('manager-summary')">Excel</button></div></div>
      </div>
    </div><span class="info-btn print-hide" onclick="showInfo(event,'export')">?</span>
    <button class="btn-theme" id="btn-theme" onclick="toggleTheme()">&#9728; Light</button>
  </div>
</div>

<div class="filters">
  <span class="filter-label">Status</span>
  <select id="f-status">
    <option value="">All</option>
    <option value="Yes">Certified</option>
    <option value="No">Not Certified</option>
  </select>
  <span class="filter-label">Market</span>
  <select id="f-market"><option value="">All Markets</option></select>
  <span class="filter-label" style="margin-right:2px">From</span>
  <input type="date" id="f-date-from">
  <span class="filter-label" style="margin:0 2px">To</span>
  <input type="date" id="f-date-to">
  <button class="btn-reset" onclick="resetFilters()">Reset</button>
  <span class="result-count" id="result-count"></span>
</div>

<div class="stats">
  <div class="stat">
    <div class="stat-label">Total Assigned <span class="info-btn" onclick="showInfo(event,'total-assigned')">?</span></div>
    <div class="stat-value" id="s-total">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">Certified <span class="info-btn" onclick="showInfo(event,'certified')">?</span></div>
    <div class="stat-value green" id="s-certified">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">Not Yet Certified <span class="info-btn" onclick="showInfo(event,'not-certified')">?</span></div>
    <div class="stat-value red" id="s-not">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">Completion Rate <span class="info-btn" onclick="showInfo(event,'completion-rate')">?</span></div>
    <div class="stat-value teal" id="s-rate">&#8212;</div>
    <div class="stat-sub" id="s-rate-sub"></div>
  </div>
</div>

<div class="charts">
  <div class="chart-card">
    <div class="chart-title">Certification Pipeline <span class="info-btn" onclick="showInfo(event,'pipeline')">?</span></div>
    <div class="chart-wrap"><canvas id="pipelineChart"></canvas></div>
  </div>
  <div class="chart-card">
    <div class="chart-title">Certifications Over Time <span class="info-btn" onclick="showInfo(event,'over-time')">?</span></div>
    <div class="chart-wrap"><canvas id="trendChart"></canvas></div>
  </div>
</div>

<div class="section">
  <div class="section-header">
    <div>
      <div class="section-title">Certification Roster <span class="info-btn" onclick="showInfo(event,'roster')">?</span></div>
      <div class="section-hint">Click a person to see their details &mdash; sorted by status then name</div>
    </div>
    <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;">
      <span style="font-size:11px;color:var(--muted);margin-right:2px;">View:</span>
      <button class="sort-btn active" id="view-individual" onclick="setRosterView('individual')">Individual</button>
      <button class="sort-btn" id="view-manager" onclick="setRosterView('manager')">By Manager</button>
      <span id="sort-controls" style="display:flex;align-items:center;gap:6px;margin-left:6px;">
        <span style="font-size:11px;color:var(--muted);margin-right:2px;">Sort:</span>
        <button class="sort-btn active" data-sort="status" data-label="Status" onclick="setRosterSort('status')">Status ↓</button>
        <button class="sort-btn" data-sort="name"   data-label="Name"   onclick="setRosterSort('name')">Name</button>
        <button class="sort-btn" data-sort="market" data-label="Market" onclick="setRosterSort('market')">Market</button>
      </span>
    </div>
    <input type="text" class="roster-search" id="roster-search" placeholder="Search by name&hellip;" oninput="filterRoster()">
  </div>
  <div style="display:flex;flex-wrap:wrap;gap:8px;align-items:center;padding:8px 18px;border-bottom:1px solid var(--border);font-size:11px;color:var(--muted);">
    <span style="font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:0.06em;margin-right:4px;">Legend</span>
    <span style="display:inline-flex;align-items:center;gap:6px;background:var(--surface2);border-radius:4px;padding:3px 8px;">
      <svg width="14" height="14" viewBox="0 0 14 14" style="flex-shrink:0;">
        <circle cx="7" cy="7" r="6" fill="var(--green-subtle)" stroke="var(--green)" stroke-width="1.5"/>
        <text x="7" y="11" text-anchor="middle" font-size="8" font-weight="700" fill="var(--green)">&#10003;</text>
      </svg>
      Certified
    </span>
  </div>
  <div class="roster-wrap">
    <div class="roster-left" id="roster-left"></div>
    <div class="roster-right" id="roster-right">
      <div class="no-data">Select a person to see their details</div>
    </div>
  </div>
</div>

<script>
const RAW = {raw_json};
const TLG_SET = new Set({tlg_json});

let filtered = [];

let pipelineChart, trendChart;
let rosterSortField = 'status';
let rosterSortDir   = 'desc';
let rosterView      = 'individual';

function sel(id){{ return document.getElementById(id); }}
function cv(v){{ return getComputedStyle(document.body).getPropertyValue(v).trim(); }}

(function(){{
  if(localStorage.getItem('pb-theme')==='light') document.body.classList.add('light-mode');
  sel('btn-theme').textContent = document.body.classList.contains('light-mode') ? '🌙 Dark' : '☀ Light';
}})();
function toggleTheme(){{
  const light = document.body.classList.toggle('light-mode');
  localStorage.setItem('pb-theme', light ? 'light' : 'dark');
  sel('btn-theme').textContent = light ? '🌙 Dark' : '☀ Light';
  applyFilters();
}}
function toggleExportDrop(){{ sel('export-menu').classList.toggle('open'); }}
document.addEventListener('click', function(e){{
  const d=sel('export-drop'); if(d&&!d.contains(e.target)) sel('export-menu').classList.remove('open');
}});

function setupPrintHeader(title, subtitle){{
  sel('ph-title').textContent=title;
  sel('ph-date').textContent=subtitle;
  const statusLabel=sel('f-status').options[sel('f-status').selectedIndex].text;
  const market=sel('f-market').value||'All Markets';
  const from=sel('f-date-from').value||'', to=sel('f-date-to').value||'';
  const parts=[`Status: ${{statusLabel}}`,`Market: ${{market}}`];
  if(from||to) parts.push(`Dates: ${{from||'-'}} to ${{to||'-'}}`);
  if(hideTLG) parts.push('TLG hidden');
  sel('ph-filters').textContent=parts.join('  |  ');
}}
function doPrint(noSummary){{
  if(noSummary) document.body.classList.add('print-no-summary');
  window.onafterprint=function(){{
    document.body.classList.remove('print-no-summary');
    sel('ph-desc').style.display='none';
    window.onafterprint=null;
  }};
  window.print();
}}
function tds(cells){{ return '<tr>'+cells.map(c=>`<td>${{c}}</td>`).join('')+'</tr>'; }}
function thRow(labels){{ return '<tr>'+labels.map(l=>`<th>${{l}}</th>`).join('')+'</tr>'; }}
function pBox(n,l){{
  return '<div style="min-width:90px"><div style="font-size:30px;font-weight:700;color:#1a3a5c;line-height:1">'+n+'</div>'
        +'<div style="font-size:10px;color:#666;text-transform:uppercase;letter-spacing:.06em;margin-top:5px">'+l+'</div></div>';
}}
function pSection(title, rows){{
  var h='<div style="min-width:150px"><div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#666;margin-bottom:10px">'+title+'</div>';
  rows.forEach(function(r){{h+='<div style="display:flex;justify-content:space-between;gap:32px;padding:4px 0;border-bottom:1px solid #eee;font-size:12px"><span>'+r[0]+'</span><strong>'+r[1]+'</strong></div>';}});
  return h+'</div>';
}}
function pFiscalQtr(d){{
  if(!d) return null;
  var pts=d.split('-'),yr=+pts[0],mo=+pts[1];
  var fy=mo>=4?yr+1:yr,q=mo>=10?3:mo>=7?2:mo>=4?1:4;
  return 'Q'+q+' FY'+String(fy).slice(2);
}}

function runExport(type){{
  sel('export-menu').classList.remove('open');
  const now=new Date().toLocaleDateString('en-US',{{year:'numeric',month:'long',day:'numeric'}});

  if(type==='full'){{
    setupPrintHeader('{name} Certification Report',`Generated: ${{now}}  |  ${{filtered.length}} People`);
    var pTotal=filtered.length, pCert=filtered.filter(p=>p.PublicSector==='Yes').length;
    var pNot=pTotal-pCert, pRate=pTotal>0?Math.round(pCert/pTotal*100):0;
    sel('print-stats').innerHTML=pBox(pTotal,'Total Assigned')+pBox(pCert,'Certified')+pBox(pNot,'Not Certified')+pBox(pRate+'%','Completion Rate');
    var qMap={{}};
    filtered.filter(p=>p.PublicSector==='Yes'&&p.CertDate).forEach(p=>{{var q=pFiscalQtr(p.CertDate);if(q)qMap[q]=(qMap[q]||0)+1;}});
    var qRows=Object.entries(qMap).sort((a,b)=>a[0].localeCompare(b[0]));
    if(!qRows.length) qRows=[['No data','-']];
    sel('print-charts').innerHTML=pSection('Certification Pipeline',[['In Progress',pNot],['Certified',pCert]])+pSection('Certifications by Quarter',qRows);
    sel('print-roster-head').innerHTML=thRow(['#','Name','Market','Job Title','Status','Cert Date','Manager','Email']);
    sel('print-roster-body').innerHTML=filtered.map((p,i)=>{{
      const status=p.PublicSector==='Yes'?'Certified':'Not Certified';
      return tds([i+1,`<b>${{p.FirstName}} ${{p.LastName}}</b>`,p.Market||'-',p.JobTitle||'-',status,p.CertDate||'-',p.Manager||'-',`<small>${{p.Email||'-'}}</small>`]);
    }}).join('');
    doPrint(false);

  }} else if(type==='not-certified'){{
    const notCert=filtered.filter(p=>p.PublicSector!=='Yes');
    setupPrintHeader('Not Certified: {name}',`Generated: ${{now}}  |  ${{notCert.length}} Employees`);
    sel('print-roster-head').innerHTML=thRow(['#','Name','Job Title','Market','Email','Manager','Manager Email']);
    sel('print-roster-body').innerHTML=notCert.length
      ? notCert.sort((a,b)=>(a.Manager||'').localeCompare(b.Manager||'')||(a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName))
          .map((p,i)=>tds([i+1,`<b>${{p.FirstName}} ${{p.LastName}}</b>`,p.JobTitle||'-',p.Market||'-',p.Email||'-',p.Manager||'-',p.MgrEmail||'-'])).join('')
      : '<tr><td colspan="7" style="color:#999;font-style:italic;padding:10px 8px">All employees are certified</td></tr>';
    sel('ph-desc').textContent='Employees who have not yet earned {name} certification, sorted by manager. Use this list to contact employees and their managers to drive completion.';
    sel('ph-desc').style.display='block';
    doPrint(true);

  }} else if(type==='manager-summary'){{
    const mgrMap={{}};
    filtered.forEach(p=>{{
      const k=p.Manager||'(No Manager)';
      if(!mgrMap[k]) mgrMap[k]={{name:k,email:p.MgrEmail||'-',total:0,cert:0}};
      mgrMap[k].total++;
      if(p.PublicSector==='Yes') mgrMap[k].cert++;
    }});
    const mgrs=Object.values(mgrMap).sort((a,b)=>(b.cert/b.total)-(a.cert/a.total));
    setupPrintHeader('Manager Summary: {name}',`Generated: ${{now}}  |  ${{mgrs.length}} Managers`);
    sel('print-roster-head').innerHTML=thRow(['Manager','Manager Email','Team Size','Certified','Not Certified','Completion %']);
    sel('print-roster-body').innerHTML=mgrs.map(m=>tds([
      `<b>${{m.name}}</b>`,m.email,m.total,m.cert,m.total-m.cert,`<b>${{Math.round(m.cert/m.total*100)}}%</b>`
    ])).join('');
    sel('ph-desc').textContent='Certification completion rates by manager, sorted highest to lowest. Teams at the bottom of the list have the most employees still working toward certification and need the most follow-up.';
    sel('ph-desc').style.display='block';
    doPrint(true);
  }}
}}

const allMarkets = [...new Set(RAW.map(r=>r.Market).filter(Boolean))].sort();
allMarkets.forEach(m => sel('f-market').innerHTML += `<option value="${{m}}">${{m}}</option>`);
['f-status','f-market','f-date-from','f-date-to'].forEach(id => {{
  sel(id).addEventListener('change', applyFilters);
}});

document.addEventListener('click', function(e){{
  if(!e.target.classList.contains('info-btn')) sel('info-popover').classList.remove('visible');
}});

const INFO_MSGS={{
  'total-assigned':  'Total number of people currently assigned the {name} certification, after any active filters.',
  'certified':       'People who have earned the {name} certification.',
  'not-certified':   'People assigned the certification who have not yet earned it.',
  'completion-rate': 'Percentage of assigned people who have earned certification. Calculated as Certified ÷ Total Assigned.',
  'pipeline':        'Two stages of the certification journey: In Progress (assigned but not yet certified) and Certified (fully certified).',
  'over-time':       'Certifications earned per KM fiscal quarter. KM quarters: Q1 = Apr–Jun, Q2 = Jul–Sep, Q3 = Oct–Dec, Q4 = Jan–Mar.',
  'roster':          'Full list of assigned people with their certification status. Click a name to see job title, market, cert date, and manager info. Use the View toggle to group by manager.',
  'export':          'Export a printable report of the data currently on screen. Apply filters first; the report only includes what is currently shown. Examples: filter to a specific region then export for a regional snapshot; hide TLG then export to share with managers; set Status = Not Certified then export for a targeted outreach list. Full Report lists everyone with full detail. Not Certified is a contact list for outreach. Manager Summary rolls up team count and completion % per manager.',
}};
function showInfo(e, key){{
  const pop=sel('info-popover');
  pop.textContent=INFO_MSGS[key]||'';
  pop.classList.add('visible');
  const r=e.target.getBoundingClientRect();
  pop.style.top=(r.bottom+6)+'px';
  pop.style.left=Math.min(r.left, window.innerWidth-280)+'px';
  e.stopPropagation();
}}
function toggleTLG(){{
  hideTLG=!hideTLG;
  sel('btn-tlg').classList.toggle('active', hideTLG);
  sel('btn-tlg').textContent=hideTLG?'Show TLG':'Hide TLG';
  applyFilters();
}}
function resetFilters(){{
  ['f-status','f-market','f-date-from','f-date-to'].forEach(id=>sel(id).value='');
  if(!hideTLG){{ hideTLG=true; sel('btn-tlg').classList.add('active'); sel('btn-tlg').textContent='Show TLG'; }}
  applyFilters();
}}
function applyFilters(){{
  const status=sel('f-status').value;
  const market=sel('f-market').value;
  const from=sel('f-date-from').value;
  const to=sel('f-date-to').value;
  filtered=RAW.filter(r=>{{
    if(hideTLG&&TLG_SET.has(r.FirstName+' '+r.LastName)) return false;
    if(status&&r.PublicSector!==status) return false;
    if(market&&r.Market!==market) return false;
    if(from&&r.Date&&r.Date<from) return false;
    if(to&&r.Date&&r.Date>to) return false;
    return true;
  }});
  sel('result-count').textContent=`${{filtered.length}} People`;
  render();
}}

function render(){{
  const isLight=document.body.classList.contains('light-mode');
  const chartLabel=isLight?cv('--text'):cv('--muted');
  const total=filtered.length;
  const certified=filtered.filter(r=>r.PublicSector==='Yes').length;
  const notCert=total-certified;
  const rate=total>0?Math.round(certified/total*100):0;
  sel('s-total').textContent=total;
  sel('s-certified').textContent=certified;
  sel('s-not').textContent=notCert;
  sel('s-rate').textContent=rate+'%';
  sel('s-rate-sub').textContent=total>0?`${{certified}} of ${{total}} assigned`:'';

  if(pipelineChart) pipelineChart.destroy();
  pipelineChart=new Chart(sel('pipelineChart'),{{
    type:'bar',
    data:{{
      labels:['In Progress','Certified'],
      datasets:[{{
        data:[notCert,certified],
        backgroundColor:[cv('--red')+'cc',cv('--green')+'cc'],
        borderRadius:4, borderSkipped:false,
      }}]
    }},
    options:{{
      indexAxis:'y', responsive:true, maintainAspectRatio:false,
      plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:ctx=>` ${{ctx.raw}} People`}}}},datalabels:{{display:false}}}},
      scales:{{
        x:{{grid:{{color:cv('--border')}},ticks:{{color:chartLabel,font:{{size:11}},stepSize:1}}}},
        y:{{grid:{{display:false}},ticks:{{color:chartLabel,font:{{size:12}}}}}}
      }}
    }}
  }});

  const parseQtr=s=>{{const m=s.match(/FY(\d+)\s+Q(\d)/);return m?parseInt(m[1])*10+parseInt(m[2]):0;}};
  const allQtrs=new Set();
  filtered.forEach(r=>{{if(r.PublicSector==='Yes'&&r.CertQtr) allQtrs.add(r.CertQtr);}});
  const trendQtrs=[...allQtrs].sort((a,b)=>parseQtr(a)-parseQtr(b));
  const trendData=trendQtrs.map(q=>filtered.filter(r=>r.PublicSector==='Yes'&&r.CertQtr===q).length);
  if(trendChart) trendChart.destroy();
  trendChart=new Chart(sel('trendChart'),{{
    type:'bar',
    data:{{
      labels:trendQtrs.length?trendQtrs:['No data yet'],
      datasets:[{{
        label:'Certified',
        data:trendData.length?trendData:[0],
        backgroundColor:cv('--green')+'bb',
        borderRadius:3, borderSkipped:false,
      }}]
    }},
    options:{{
      responsive:true, maintainAspectRatio:false,
      plugins:{{legend:{{display:false}},tooltip:{{mode:'index',intersect:false}},datalabels:{{display:false}}}},
      scales:{{
        x:{{grid:{{color:cv('--border')}},ticks:{{color:chartLabel,font:{{size:10}},maxRotation:45}}}},
        y:{{grid:{{color:cv('--border')}},ticks:{{color:chartLabel,font:{{size:11}},stepSize:1}}}}
      }}
    }}
  }});
  renderRosterList();
}}

function setRosterSort(field){{
  if(rosterSortField===field) rosterSortDir=rosterSortDir==='desc'?'asc':'desc';
  else{{ rosterSortField=field; rosterSortDir='desc'; }}
  renderRosterList();
}}
function setRosterView(v){{
  rosterView=v;
  sel('view-individual').classList.toggle('active',v==='individual');
  sel('view-manager').classList.toggle('active',v==='manager');
  sel('sort-controls').style.display=v==='individual'?'flex':'none';
  sel('roster-right').innerHTML='<div class="no-data">Select a person to see their details</div>';
  renderRosterList();
}}
function toggleMgrGroup(el){{ el.classList.toggle('open'); el.nextElementSibling.classList.toggle('open'); }}

function certIndicator(p){{
  const isCert=p.PublicSector==='Yes';
  const color=isCert?cv('--green'):cv('--red');
  const bg=isCert?color+'33':'var(--surface2)';
  const border=isCert?color:'var(--border)';
  const txt=isCert?color:'var(--muted)';
  return `<div title="${{isCert?'Certified':'Not Certified'}}" style="width:16px;height:16px;border-radius:50%;flex-shrink:0;display:flex;align-items:center;justify-content:center;font-size:9px;font-weight:700;background:${{bg}};border:1.5px solid ${{border}};color:${{txt}}">${{isCert?'&#10003;':''}}</div>`;
}}

function renderRosterList(){{
  if(rosterView==='manager'){{
    const groups={{}};
    filtered.forEach(p=>{{
      const mgr=p.Manager||'No Manager';
      if(!groups[mgr]) groups[mgr]=[];
      groups[mgr].push(p);
    }});
    const mgrsSorted=Object.keys(groups).sort((a,b)=>groups[b].length-groups[a].length);
    sel('roster-left').innerHTML=mgrsSorted.map(mgr=>{{
      const team=groups[mgr];
      const cert=team.filter(p=>p.PublicSector==='Yes').length;
      const peopleHtml=team
        .sort((a,b)=>(a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName))
        .map(p=>{{
          const fullName=`${{p.FirstName}} ${{p.LastName}}`;
          const ts=p.PublicSector==='Yes'?`border-left:3px solid ${{cv('--green')}};padding-left:21px;`:'padding-left:24px;';
          return `<div class="roster-person" onclick="rosterSelect(this)" data-name="${{fullName}}" style="${{ts}}">
            <span class="roster-name">${{fullName}}</span>
            ${{certIndicator(p)}}
          </div>`;
        }}).join('');
      return `<div>
        <div class="mgr-group-hdr open" onclick="toggleMgrGroup(this)">
          <div>
            <div style="font-size:12px;font-weight:600;color:var(--text)">${{mgr}}</div>
            <div style="font-size:11px;color:var(--muted);margin-top:2px;">${{team.length}} rep${{team.length!==1?'s':''}} &middot; ${{cert}} certified</div>
          </div>
          <span class="mgr-chevron">&#9660;</span>
        </div>
        <div class="mgr-team open">${{peopleHtml}}</div>
      </div>`;
    }}).join('')||`<div class="no-data">No people match filters</div>`;
    filterRoster();
    return;
  }}

  const d=rosterSortDir==='desc'?-1:1;
  const sorted=[...filtered].sort((a,b)=>{{
    if(rosterSortField==='name') return d*(a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName);
    if(rosterSortField==='market'){{
      const rc=d*(a.Market||'').localeCompare(b.Market||'');
      return rc!==0?rc:(a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName);
    }}
    const sa=a.PublicSector==='Yes'?1:0, sb=b.PublicSector==='Yes'?1:0;
    if(sa!==sb) return d*(sa-sb);
    return (a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName);
  }});

  document.querySelectorAll('.sort-btn').forEach(btn=>{{
    if(!btn.dataset.sort) return;
    const active=btn.dataset.sort===rosterSortField;
    btn.classList.toggle('active',active);
    const arrow=active?(rosterSortDir==='desc'?' ↓':' ↑'):'';
    btn.textContent=btn.dataset.label+arrow;
  }});

  sel('roster-left').innerHTML=sorted.map(p=>{{
    const fullName=`${{p.FirstName}} ${{p.LastName}}`;
    const ts=p.PublicSector==='Yes'?`border-left:3px solid ${{cv('--green')}};padding-left:11px;`:'';
    return `<div class="roster-person" onclick="rosterSelect(this)" data-name="${{fullName}}" style="${{ts}}">
      <span class="roster-name">${{fullName}}</span>
      ${{certIndicator(p)}}
    </div>`;
  }}).join('')||`<div class="no-data">No people match filters</div>`;

  filterRoster();
  const first=sel('roster-left').querySelector('.roster-person:not([style*="none"])')||
               sel('roster-left').querySelector('.roster-person');
  if(first) rosterSelect(first);
}}

function rosterSelect(el){{
  document.querySelectorAll('.roster-person').forEach(p=>p.classList.remove('active'));
  el.classList.add('active');
  const name=el.dataset.name;
  const p=filtered.find(r=>`${{r.FirstName}} ${{r.LastName}}`===name);
  if(!p) return;
  const isCert=p.PublicSector==='Yes';
  function fmtDate(d){{
    if(!d) return '&#8212;';
    const [y,mo,dy]=d.split('-');
    const months=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
    return months[parseInt(mo)-1]+' '+parseInt(dy)+', '+y;
  }}
  sel('roster-right').innerHTML=`
    <div class="roster-right-header">
      <div style="font-size:15px;font-weight:700;margin-bottom:6px">${{p.FirstName}} ${{p.LastName}}</div>
      <span class="badge-status ${{isCert?'certified':'not-certified'}}">${{isCert?'Certified':'Not Yet Certified'}}</span>
      ${{isCert&&p.CertDate?`<span style="font-size:12px;color:var(--muted);margin-left:8px">${{fmtDate(p.CertDate)}}</span>`:''}}
    </div>
    <div class="detail-grid">
      <div><div class="detail-label">Job Title</div><div class="detail-value">${{p.JobTitle||'&#8212;'}}</div></div>
      <div><div class="detail-label">Market</div><div class="detail-value">${{p.Market||'&#8212;'}}</div></div>
      <div><div class="detail-label">Assigned</div><div class="detail-value">${{fmtDate(p.Date)}}</div></div>
      <div><div class="detail-label">Cert Date</div><div class="detail-value">${{fmtDate(p.CertDate)}}</div></div>
      <div><div class="detail-label">Cert Quarter</div><div class="detail-value">${{p.CertQtr||'&#8212;'}}</div></div>
      <div><div class="detail-label">Hire Date</div><div class="detail-value">${{fmtDate(p.HireDate)}}</div></div>
      <div style="grid-column:1/-1"><div class="detail-label">Email</div><div class="detail-value"><a href="mailto:${{p.Email}}" style="color:var(--accent);text-decoration:none">${{p.Email||'&#8212;'}}</a></div></div>
    </div>
    ${{p.Manager?`<div style="margin-top:16px;padding-top:14px;border-top:1px solid var(--border)">
      <div style="font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:10px">Manager</div>
      <div class="detail-grid">
        <div><div class="detail-label">Name</div><div class="detail-value">${{p.Manager}}</div></div>
        <div><div class="detail-label">Title</div><div class="detail-value">${{p.MgrTitle||'&#8212;'}}</div></div>
        <div><div class="detail-label">Email</div><div class="detail-value"><a href="mailto:${{p.MgrEmail}}" style="color:var(--accent);text-decoration:none">${{p.MgrEmail||'&#8212;'}}</a></div></div>
      </div>
    </div>`:''}}
  `;
}}

function filterRoster(){{
  const q=(sel('roster-search')?.value||'').toLowerCase();
  document.querySelectorAll('.roster-person').forEach(el=>{{
    el.style.display=(!q||el.dataset.name.toLowerCase().includes(q))?'':'none';
  }});
}}

function runExportXLSX(type){{
  sel('export-menu').classList.remove('open');
  const now=new Date().toLocaleDateString('en-US',{{year:'numeric',month:'long',day:'numeric'}});
  function makeSheet(rows,colWidths){{
    const ws=XLSX.utils.aoa_to_sheet(rows);
    ws['!cols']=colWidths.map(w=>({{wch:w}}));
    return ws;
  }}
  function dlXLSX(name,wb){{ XLSX.writeFile(wb,name+'.xlsx'); }}
  const wb=XLSX.utils.book_new();
  if(type==='full'){{
    const pTotal=filtered.length;
    const pCert=filtered.filter(p=>p.PublicSector==='Yes').length;
    const pNot=pTotal-pCert;
    const pRate=pTotal>0?Math.round(pCert/pTotal*100):0;
    const sumRows=[
      ['Public Sector Certification Report'],
      ['Generated:',now],
      [],
      ['SUMMARY'],
      ['Total Assigned',pTotal],
      ['Certified',pCert],
      ['Not Certified',pNot],
      ['Completion Rate',pRate+'%'],
    ];
    XLSX.utils.book_append_sheet(wb,makeSheet(sumRows,[30,18]),'Summary');
    const rRows=[['Name','Market','Job Title','Status','Cert Date','Manager','Email']];
    filtered.forEach(p=>{{
      const status=p.PublicSector==='Yes'?'Certified':'Not Certified';
      rRows.push([p.FirstName+' '+p.LastName,p.Market||'-',p.JobTitle||'-',status,p.CertDate||'-',p.Manager||'-',p.Email||'-']);
    }});
    XLSX.utils.book_append_sheet(wb,makeSheet(rRows,[28,18,28,14,14,28,32]),'Roster');
    dlXLSX('ps-full-report',wb);
  }} else if(type==='not-certified'){{
    const notCert=filtered.filter(p=>p.PublicSector!=='Yes').sort((a,b)=>(a.Manager||'').localeCompare(b.Manager||'')||(a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName));
    const rows=[['Name','Job Title','Market','Email','Manager','Manager Email']];
    notCert.forEach(p=>rows.push([p.FirstName+' '+p.LastName,p.JobTitle||'-',p.Market||'-',p.Email||'-',p.Manager||'-',p.MgrEmail||'-']));
    XLSX.utils.book_append_sheet(wb,makeSheet(rows,[28,28,18,32,28,32]),'Not Certified');
    dlXLSX('ps-not-certified',wb);
  }} else if(type==='manager-summary'){{
    const mgrMap={{}};
    filtered.forEach(p=>{{
      const k=p.Manager||'(No Manager)';
      if(!mgrMap[k]) mgrMap[k]={{name:k,email:p.MgrEmail||'-',total:0,cert:0}};
      mgrMap[k].total++;
      if(p.PublicSector==='Yes') mgrMap[k].cert++;
    }});
    const rows=[['Manager','Manager Email','Team Size','Certified','Not Certified','Completion %']];
    Object.values(mgrMap).sort((a,b)=>(b.cert/b.total)-(a.cert/a.total)).forEach(m=>{{
      rows.push([m.name,m.email,m.total,m.cert,m.total-m.cert,Math.round(m.cert/m.total*100)+'%']);
    }});
    XLSX.utils.book_append_sheet(wb,makeSheet(rows,[28,32,12,12,14,14]),'Manager Summary');
    dlXLSX('ps-manager-summary',wb);
  }}
}}

applyFilters();
(function(){{
  var n=0,t;
  var h=document.querySelector('.header h1');
  if(h) h.addEventListener('click',function(){{
    n++;clearTimeout(t);
    if(n>=3){{n=0;window.location.href='index.html';}}
    else t=setTimeout(function(){{n=0;}},1500);
  }});
}})();
</script>
<div class="info-popover" id="info-popover"></div>
<!-- Print-only elements -->
<div id="print-header" style="display:none;margin-bottom:20px;">
  <div style="font-size:20px;font-weight:700;margin-bottom:4px;" id="ph-title"></div>
  <div style="font-size:12px;color:#555;margin-bottom:2px;" id="ph-date"></div>
  <div style="font-size:12px;color:#555;" id="ph-filters"></div>
  <div id="ph-desc" style="display:none;margin-top:10px;padding-top:10px;border-top:1px solid #dde4f0;font-size:12px;color:#444;font-style:italic;"></div>
</div>
<div id="print-stats" style="display:none;gap:40px;flex-wrap:wrap;margin-bottom:18px;padding-bottom:16px;border-bottom:2px solid #dde4f0;"></div>
<div id="print-charts" style="display:none;gap:48px;flex-wrap:wrap;margin-bottom:18px;padding-bottom:16px;border-bottom:2px solid #dde4f0;"></div>
<div id="print-roster-wrap" style="display:none;">
  <table class="ptable" id="print-roster-table">
    <thead id="print-roster-head"><tr>
      <th>#</th><th>Name</th><th>Market</th><th>Job Title</th>
      <th>Status</th><th>Cert Date</th><th>Manager</th><th>Email</th>
    </tr></thead>
    <tbody id="print-roster-body"></tbody>
  </table>
</div>
</body>
</html>"""


def main():
    cert_dir = 'cert-data'
    if not os.path.isdir(cert_dir):
        print(f'No {cert_dir}/ directory found.')
        return

    files = sorted(f for f in os.listdir(cert_dir) if f.endswith('.xlsx'))
    if not files:
        print('No .xlsx files found in cert-data/')
        return

    generated = []
    hc_handled = False

    # ── New two-file HC format ────────────────────────────────────────────────
    HC_CERT_FILE  = 'Healthcare=Certification-Report-07.06.2026.xlsx'
    HC_LEARN_FILE = 'Healthcare-Certification-Foundations-Curricula-Report-07.06.2026.xlsx'
    hc_cert_path  = os.path.join(cert_dir, HC_CERT_FILE)
    hc_learn_path = os.path.join(cert_dir, HC_LEARN_FILE)

    if os.path.exists(hc_cert_path) and os.path.exists(hc_learn_path):
        print(f'\nhealthcare (Healthcare) — v2 two-file format:')
        print(f'  {HC_CERT_FILE}')
        print(f'  {HC_LEARN_FILE}')
        rows = load_rows_healthcare_v2(hc_cert_path, hc_learn_path)
        cert_count = sum(1 for r in rows if r['Certified'] == 'Yes')
        ip_count   = sum(1 for r in rows if r['overallPct'] > 0 and r['Certified'] != 'Yes')
        print(f'  → {len(rows)} people  ({cert_count} certified, {ip_count} in progress, {len(rows)-cert_count-ip_count} not started)')
        import datetime, time
        latest_mtime = max(os.path.getmtime(hc_cert_path), os.path.getmtime(hc_learn_path))
        hc_date_label = datetime.datetime.fromtimestamp(latest_mtime).strftime('%B %Y')
        html = generate_html_healthcare_v2('healthcare', 'Healthcare', rows, hc_date_label)
        out  = 'cert-healthcare.html'
        with open(out, 'w', encoding='utf-8') as fh:
            fh.write(html)
        print(f'  Written → {out}')
        generated.append(out)
        hc_handled = True

    # ── Original single-file loop (PS and other verticals) ───────────────────
    vert_files = {}
    for fname in files:
        slug = detect_vertical(fname)
        if not slug:
            print(f'  Skipping {fname} — could not detect vertical from filename')
            continue
        if slug == 'healthcare' and hc_handled:
            continue  # already handled above with v2 loader
        vert_files.setdefault(slug, []).append(fname)

    for slug in sorted(vert_files):
        # Sort files chronologically by YYYY-MM in filename
        fnames    = sorted(vert_files[slug], key=extract_file_date)
        vert_name = VERTICAL_MAP.get(slug, slug.title())
        print(f'\n{slug} ({vert_name}) — {len(fnames)} file(s):')

        # Pick loader based on vertical
        if slug == 'healthcare':
            loader = load_rows
        elif slug == 'publicsector':
            loader = load_rows_publicsector
        else:
            print(f'  No loader for vertical "{slug}" — skipping')
            continue

        # Load all files in order, tag each row with its file date
        all_rows = []
        for fname in fnames:
            file_date = extract_file_date(fname)
            filepath  = os.path.join(cert_dir, fname)
            print(f'  {fname}  [{file_date}]')
            rows = loader(filepath)
            for r in rows:
                r['_file_date'] = file_date
            all_rows.extend(rows)
            if slug == 'healthcare':
                print(f'    {len(rows)} rows  ({sum(1 for r in rows if r["Healthcare"]=="Yes")} HC certified, {sum(1 for r in rows if r["Complete"]=="Yes")} curriculum complete)')
            elif slug == 'publicsector':
                print(f'    {len(rows)} rows  ({sum(1 for r in rows if r["PublicSector"]=="Yes")} certified)')

        # Deduplicate: later files overwrite earlier ones for the same person
        all_rows.sort(key=lambda r: r['_file_date'])
        seen = {}
        for r in all_rows:
            seen[person_key(r)] = r
        deduped = list(seen.values())
        for r in deduped:
            del r['_file_date']

        if slug == 'healthcare':
            hc_cert = sum(1 for r in deduped if r['Healthcare'] == 'Yes')
            curr_complete = sum(1 for r in deduped if r['Complete'] == 'Yes')
            print(f'  → {len(deduped)} unique people  ({hc_cert} HC certified, {curr_complete} curriculum complete, {len(deduped)-hc_cert} not yet)')
            html = generate_html(slug, vert_name, deduped)
        elif slug == 'publicsector':
            ps_cert = sum(1 for r in deduped if r['PublicSector'] == 'Yes')
            print(f'  → {len(deduped)} unique people  ({ps_cert} certified, {len(deduped)-ps_cert} not yet)')
            ps_date_label = fmt_date_label(extract_file_date(fnames[-1]))
            html = generate_html_publicsector(slug, vert_name, deduped, ps_date_label)

        out  = f'cert-{slug}.html'
        with open(out, 'w', encoding='utf-8') as fh:
            fh.write(html)
        print(f'  Written → {out}')
        generated.append(out)

    print(f'\nDone. {len(generated)} dashboard(s) generated.')


def load_rows_healthcare_v2(cert_file, learning_file):
    """Load and join cert file + learning file for the new HC v2 data model.

    cert_file columns (0-based):
      2=First, 3=Last, 4=Email, 5=JobTitle, 7=Market,
      9=MgrFirst, 10=MgrLast, 11=MgrEmail, 12=MgrTitle,
      14=HireDate(datetime), 17=CertDate(datetime), 19=Certified("Yes"/"No")

    learning_file columns (0-based):
      4=Email, 16=CurriculumTitle, 20=ItemID, 24=ItemTitle,
      25=CompletionDate(datetime), 27=CompletionStatusDesc
      Skip rows where col 20 (ItemID) is None — those are curriculum-level rows.
    """

    HCF_ORDER = ['HCF_HBT','HC_PLAYBOOK','HCF_MFP','HCF_DPMS','HCF_ACS',
                 'HCF_HBMF','HCF_HIPAACS','HCF_DPFH','HCF_OW','HCF_HHSPD']
    LS_ORDER  = ['LS_ITM','DS','UA_ACCESSCONTROL','BIZHUB_BSMFP','MFPPT_BSBSNB',
                 'BREACH_IDADB','CSMWSG','LSSW','LSSB_HI','LSSB_NONPROFIT','LSSB_GOV']

    def _date(v):
        return v.strftime('%Y-%m-%d') if v and hasattr(v, 'strftime') else ''

    def _str(v):
        return str(v).strip() if v is not None else ''

    # ── Load cert file ────────────────────────────────────────────────────────
    cert_map = {}   # email.lower() -> person dict
    wb_c = openpyxl.load_workbook(cert_file, read_only=True, data_only=True)
    ws_c = wb_c.active
    for raw in ws_c.iter_rows(min_row=2, values_only=True):
        if raw[2] is None:
            continue
        email = _str(raw[4]).lower()
        mgr_first = _str(raw[9])
        mgr_last  = _str(raw[10])
        mgr_name  = (mgr_first + ' ' + mgr_last).strip()
        cert_date_raw = raw[17]
        cert_date = _date(cert_date_raw)
        cert_qtr  = _str(raw[18]) if raw[18] else ''
        certified = _str(raw[19]) if raw[19] else 'No'
        hire_date = _date(raw[14])
        cert_map[email] = {
            'FirstName': _str(raw[2]),
            'LastName':  _str(raw[3]),
            'Email':     _str(raw[4]),
            'JobTitle':  _str(raw[5]),
            'Market':    _str(raw[7]),
            'Manager':   mgr_name,
            'MgrEmail':  _str(raw[11]),
            'MgrTitle':  _str(raw[12]),
            'HireDate':  hire_date,
            'Certified': certified,
            'CertDate':  cert_date,
            'CertQtr':   cert_qtr,
        }
    wb_c.close()

    # ── Load learning file — build per-person item maps ───────────────────────
    # Structure: learning[email][curriculum_id][item_id] = {title, done, date}
    CURRIC_TITLE_TO_ID = {
        'healthcare foundations': 'HC_FOUNDATIONS',
        'layered security certification - direct sales': 'LSFDS',
    }
    learning = {}
    wb_l = openpyxl.load_workbook(learning_file, read_only=True, data_only=True)
    ws_l = wb_l.active
    for raw in ws_l.iter_rows(min_row=2, values_only=True):
        item_id = raw[20]
        if item_id is None:
            continue   # curriculum-level parent row — skip
        email = _str(raw[4]).lower()
        if not email:
            continue
        ctitle = _str(raw[16]).lower()
        curr_id = CURRIC_TITLE_TO_ID.get(ctitle)
        if not curr_id:
            continue   # unknown curriculum — skip
        item_id = _str(item_id).upper()
        title   = _str(raw[24])
        comp_date_raw = raw[25]
        comp_date = _date(comp_date_raw)
        status    = _str(raw[27])
        done = (status == 'Online-Complete')
        if email not in learning:
            learning[email] = {}
        if curr_id not in learning[email]:
            learning[email][curr_id] = {}
        learning[email][curr_id][item_id] = {
            'title': title,
            'done':  done,
            'date':  comp_date if done else '',
        }
    wb_l.close()

    # ── Build output rows — only people in cert file ──────────────────────────
    rows = []
    for email, person in cert_map.items():
        person_learning = learning.get(email, {})

        def _build_curriculum(order, curr_id):
            curr_items = person_learning.get(curr_id, {})
            items = []
            done_count = 0
            for iid in order:
                raw_item = curr_items.get(iid, {})
                is_done  = raw_item.get('done', False)
                # Attempt to find a title from the data; fall back to item ID
                title = raw_item.get('title', iid)
                if not title:
                    title = iid
                items.append({
                    'id':    iid,
                    'title': title,
                    'done':  is_done,
                    'date':  raw_item.get('date', '') if is_done else '',
                })
                if is_done:
                    done_count += 1
            total = len(order)
            pct = round(done_count / total * 100) if total else 0
            return {'done': done_count, 'total': total, 'pct': pct, 'items': items}

        hcf = _build_curriculum(HCF_ORDER, 'HC_FOUNDATIONS')
        ls  = _build_curriculum(LS_ORDER,  'LSFDS')

        overall_done = hcf['done'] + ls['done']
        overall_pct  = round(overall_done / 21 * 100)

        rows.append({
            'FirstName':   person['FirstName'],
            'LastName':    person['LastName'],
            'Email':       person['Email'],
            'JobTitle':    person['JobTitle'],
            'Market':      person['Market'],
            'Manager':     person['Manager'],
            'MgrEmail':    person['MgrEmail'],
            'MgrTitle':    person['MgrTitle'],
            'HireDate':    person['HireDate'],
            'Certified':   person['Certified'],
            'CertDate':    person['CertDate'],
            'CertQtr':     person['CertQtr'],
            'hcf':         hcf,
            'ls':          ls,
            'overallDone': overall_done,
            'overallPct':  overall_pct,
        })

    return rows


def generate_html_healthcare_v2(slug, name, rows, date_label=''):
    """Generate the Healthcare v2 certification dashboard HTML.

    Supports two curricula displayed as course-level progress:
      HC Foundations (10 items) and Layered Security (11 items).

    Uses double-brace escaping throughout for f-string safety.
    All JS string literals use double quotes to avoid apostrophe breakage.
    """
    raw_json = json.dumps(rows)
    tlg_json = json.dumps(sorted(TLG))

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{name} Certification Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<style>
  :root {{
    --bg:#0f1117; --surface:#1a1d27; --surface2:#22263a; --border:#2e3350;
    --accent:#4f8ef7; --accent2:#7c5cfc; --accent3:#f7c94f;
    --text:#e8ecf4; --muted:#7b82a0; --green:#3ecf8e; --red:#f76f6f;
    --teal:#2dd4bf; --green-subtle:#3ecf8e22; --red-subtle:#f76f6f22;
    --font:'Segoe UI',system-ui,sans-serif;
  }}
  body.light-mode {{
    --bg:#f4f6fb; --surface:#ffffff; --surface2:#eef1f7; --border:#d0d7e8;
    --accent:#2563eb; --accent2:#6d28d9; --accent3:#d97706;
    --text:#1a1d27; --muted:#475569; --green:#059669; --red:#dc2626;
    --teal:#0f766e; --green-subtle:#05966922; --red-subtle:#dc262622;
  }}
  body.light-mode select,body.light-mode input[type=date]{{color-scheme:light;}}
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{background:var(--bg);color:var(--text);font-family:var(--font);min-height:100vh;transition:background .2s,color .2s;}}

  /* ── Header ────────────────────────────────────────────────────────── */
  .header{{padding:20px 28px 16px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px;}}
  .header-left{{display:flex;align-items:center;gap:16px;}}
  .header h1{{font-size:18px;font-weight:700;letter-spacing:.3px;}}
  .header h1 span{{color:var(--muted);font-weight:400;}}
  .header-date{{font-size:11px;color:var(--muted);margin-top:2px;}}
  .btn-theme{{background:transparent;border:1px solid var(--border);color:var(--muted);border-radius:6px;padding:5px 12px;font-size:12px;cursor:pointer;transition:all .15s;}}
  .btn-theme:hover{{border-color:var(--accent);color:var(--text);}}

  /* ── Export dropdown ───────────────────────────────────────────────── */
  .btn-export{{background:var(--accent);border:1px solid var(--accent);color:#fff;border-radius:6px;padding:5px 12px;font-size:12px;cursor:pointer;transition:all .15s;font-weight:600;}}
  .btn-export:hover{{opacity:0.88;}}
  .export-drop{{position:relative;}}
  .export-menu{{position:absolute;top:calc(100% + 6px);right:0;background:var(--surface);border:1px solid var(--border);border-radius:8px;min-width:210px;box-shadow:0 4px 24px rgba(0,0,0,.28);display:none;z-index:200;overflow:visible;}}
  .export-menu.open{{display:block;}}
  .export-item{{display:block;width:100%;text-align:left;padding:10px 14px;font-size:13px;color:var(--text);background:transparent;border:none;cursor:pointer;transition:background .1s;font-family:inherit;}}
  .export-item:hover{{background:var(--surface2);}}
  .export-parent{{position:relative;display:flex;justify-content:space-between;align-items:center;padding:10px 14px;font-size:13px;color:var(--text);cursor:default;transition:background .1s;}}
  .export-parent:hover{{background:var(--surface2);}}
  .export-chevron{{font-size:11px;color:var(--muted);margin-left:10px;}}
  .export-submenu{{position:absolute;right:100%;top:0;background:var(--surface);border:1px solid var(--border);border-radius:8px;min-width:90px;box-shadow:0 4px 24px rgba(0,0,0,0.28);display:none;z-index:201;overflow:hidden;margin-right:4px;}}
  .export-parent:hover .export-submenu{{display:block;}}

  /* ── Info button / popover ─────────────────────────────────────────── */
  .info-btn{{display:inline-flex;align-items:center;justify-content:center;width:15px;height:15px;border-radius:50%;background:var(--surface2);border:1px solid var(--border);color:var(--muted);font-size:9px;font-weight:700;cursor:pointer;margin-left:5px;vertical-align:middle;flex-shrink:0;line-height:1;transition:border-color .15s,color .15s;}}
  .info-btn:hover{{border-color:var(--accent);color:var(--accent);}}
  .info-popover{{position:fixed;z-index:9999;background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:12px 14px;font-size:12px;color:var(--text);line-height:1.6;max-width:260px;box-shadow:0 4px 24px rgba(0,0,0,0.5);display:none;}}
  .info-popover.visible{{display:block;}}

  /* ── Filters ───────────────────────────────────────────────────────── */
  .filters{{padding:14px 28px;display:flex;gap:10px;flex-wrap:wrap;align-items:center;border-bottom:1px solid var(--border);background:var(--surface);}}
  .filter-label{{font-size:12px;color:var(--muted);text-transform:uppercase;letter-spacing:.6px;margin-right:4px;}}
  select,input[type=date],input[type=text]{{background:var(--surface2);border:1px solid var(--border);color:var(--text);border-radius:6px;padding:6px 10px;font-size:13px;cursor:pointer;outline:none;color-scheme:dark;}}
  select:focus,input:focus{{border-color:var(--accent);}}
  .btn-reset{{background:transparent;border:1px solid var(--border);color:var(--muted);border-radius:6px;padding:6px 14px;font-size:12px;cursor:pointer;transition:border-color .15s,color .15s;}}
  .btn-reset:hover{{border-color:var(--accent);color:var(--text);}}
  .result-count{{margin-left:auto;font-size:12px;color:var(--muted);}}

  /* ── Stat cards ────────────────────────────────────────────────────── */
  .stats{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;padding:20px 28px;}}
  .stat{{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:16px 18px;}}
  .stat-label{{font-size:11px;text-transform:uppercase;letter-spacing:.7px;color:var(--muted);margin-bottom:6px;}}
  .stat-value{{font-size:28px;font-weight:700;line-height:1;}}
  .stat-value.green{{color:var(--green);}}
  .stat-value.red{{color:var(--red);}}
  .stat-value.teal{{color:var(--teal);}}
  .stat-value.blue{{color:var(--accent);}}
  .stat-value.amber{{color:var(--accent3);}}
  .stat-sub{{font-size:11px;color:var(--muted);margin-top:4px;}}

  /* ── Charts ────────────────────────────────────────────────────────── */
  .charts{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px;padding:0 28px 16px;}}
  @media(max-width:1100px){{.charts{{grid-template-columns:1fr 1fr;}}}}
  @media(max-width:700px){{.charts{{grid-template-columns:1fr;}}}}
  @media(max-width:480px){{.chart-wrap{{height:180px;}}}}
  .chart-card{{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:18px;}}
  .chart-title{{font-size:13px;font-weight:600;margin-bottom:14px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;}}
  body.light-mode .chart-title{{color:var(--text);}}
  .chart-wrap{{position:relative;height:260px;}}

  /* ── Roster ────────────────────────────────────────────────────────── */
  .section{{padding:0 28px 32px;}}
  .section-header{{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;flex-wrap:wrap;gap:8px;}}
  .section-title{{font-size:13px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;}}
  body.light-mode .section-title{{color:var(--text);}}
  .section-hint{{font-size:11px;color:var(--muted);margin-top:3px;}}
  .roster-search{{width:200px;}}
  .roster-wrap{{display:flex;border:1px solid var(--border);border-radius:10px;overflow:hidden;}}
  .roster-left{{width:300px;flex-shrink:0;overflow-y:auto;max-height:820px;border-right:1px solid var(--border);}}
  .roster-right{{flex:1;overflow-y:auto;max-height:820px;padding:16px 20px;}}
  .roster-right-header{{margin-bottom:14px;padding-bottom:12px;border-bottom:1px solid var(--border);}}
  .no-data{{text-align:center;color:var(--muted);padding:40px;font-size:13px;}}
  @media(max-width:680px){{.roster-wrap{{flex-direction:column;}}.roster-left{{width:100%;max-height:220px;border-right:none;border-bottom:1px solid var(--border);}}}}

  /* Person cards */
  .roster-person{{display:flex;flex-direction:column;padding:10px 14px;cursor:pointer;border-bottom:1px solid var(--border);transition:background .1s;border-left:3px solid transparent;}}
  .roster-person:last-child{{border-bottom:none;}}
  .roster-person:hover{{background:var(--surface2);}}
  .roster-person.active{{background:#4f8ef711;}}
  .roster-person.stripe-green{{border-left-color:var(--green);}}
  .roster-person.stripe-yellow{{border-left-color:#f59e0b;}}
  .roster-person.stripe-red{{border-left-color:var(--red);}}
  .roster-top{{display:flex;justify-content:space-between;align-items:flex-start;gap:8px;}}
  .roster-name-block{{flex:1;min-width:0;}}
  .roster-name{{font-size:13px;font-weight:600;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}}
  .roster-title{{font-size:11px;color:var(--muted);margin-top:1px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}}
  .roster-bottom{{display:flex;justify-content:space-between;align-items:center;margin-top:7px;}}
  .roster-pills{{display:flex;gap:5px;flex-wrap:wrap;}}
  .pill{{font-size:8px;font-weight:700;border-radius:10px;padding:2px 7px;white-space:nowrap;letter-spacing:.03em;}}
  .pill.green{{color:var(--green);background:var(--green-subtle);border:1px solid var(--green)44;}}
  .pill.yellow{{color:#b45309;background:#fef3c7;border:1px solid #f59e0b44;}}
  .pill.red{{color:var(--red);background:var(--red-subtle);border:1px solid var(--red)44;}}
  .roster-pct{{font-size:12px;font-weight:700;color:var(--muted);flex-shrink:0;margin-left:8px;}}
  .roster-pct.pct-progress{{color:#f59e0b;}}
  .roster-pct.pct-done{{color:var(--green);}}

  /* Manager group view */
  .mgr-group-hdr{{display:flex;align-items:center;justify-content:space-between;padding:10px 14px;cursor:pointer;background:var(--surface2);border-bottom:1px solid var(--border);user-select:none;}}
  .mgr-group-hdr:hover{{background:var(--border);}}
  .mgr-group-hdr.open .mgr-chevron{{transform:rotate(180deg);}}
  .mgr-chevron{{transition:transform .15s;font-size:10px;color:var(--muted);flex-shrink:0;}}
  .mgr-team{{display:none;}}
  .mgr-team.open{{display:block;}}

  /* Roster key */
  .roster-key{{display:flex;gap:16px;align-items:center;padding:6px 12px;background:var(--surface2);border-radius:8px;margin-bottom:10px;font-size:11px;color:var(--muted);}}
  .key-stripe{{display:inline-block;width:10px;height:10px;border-radius:2px;vertical-align:middle;margin-right:5px;}}
  .key-stripe.green{{background:var(--green);}}
  .key-stripe.yellow{{background:#f59e0b;}}
  .key-stripe.red{{background:var(--red);}}

  /* Sort buttons */
  .sort-btn{{background:transparent;border:1px solid var(--border);color:var(--muted);border-radius:6px;padding:4px 10px;font-size:11px;cursor:pointer;transition:all .15s;white-space:nowrap;}}
  .sort-btn:hover{{border-color:var(--accent);color:var(--text);}}
  .sort-btn.active{{border-color:var(--accent);color:var(--accent);background:var(--accent)11;}}

  /* Detail panel */
  .detail-grid{{display:grid;grid-template-columns:1fr 1fr;gap:12px 28px;margin-top:14px;}}
  .detail-label{{font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:3px;}}
  .detail-value{{font-size:14px;font-weight:500;}}
  .badge-status{{display:inline-block;padding:3px 12px;border-radius:20px;font-size:12px;font-weight:700;}}
  .badge-status.certified{{background:var(--green-subtle);color:var(--green);}}
  .badge-status.in-progress{{background:#fef3c7;color:#b45309;border:1px solid #f59e0b44;}}
  .badge-status.not-certified{{background:var(--red-subtle);color:var(--red);border:1px solid var(--red)44;}}
  .roster-bottom .badge-status{{font-size:10px;padding:2px 8px;}}

  /* Progress bars */
  .prog-wrap{{background:var(--surface2);border-radius:4px;height:6px;width:100%;overflow:hidden;margin-top:4px;}}
  .prog-bar{{height:6px;border-radius:4px;transition:width .3s;}}
  .prog-bar.green{{background:var(--green);}}
  .prog-bar.blue{{background:var(--accent);}}

  /* Curriculum course list */
  .curriculum-section{{margin-top:18px;}}
  .curriculum-header{{display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;cursor:pointer;user-select:none;}}
  .curriculum-header:hover .curriculum-title{{color:var(--accent);}}
  .curriculum-title{{font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:var(--muted);transition:color .15s;}}
  body.light-mode .curriculum-title{{color:var(--text);}}
  .curriculum-count{{font-size:11px;color:var(--muted);}}
  .course-list{{margin-top:8px;display:none;}}
  .course-list.open{{display:block;}}
  .course-item{{display:flex;align-items:flex-start;gap:8px;padding:5px 0;border-bottom:1px solid var(--border);}}
  .course-item:last-child{{border-bottom:none;}}
  .course-icon{{flex-shrink:0;font-size:11px;margin-top:1px;}}
  .course-icon.done{{color:var(--green);}}
  .course-icon.todo{{color:var(--muted);}}
  .course-title{{flex:1;font-size:12px;line-height:1.4;}}
  .course-date{{font-size:10px;color:var(--muted);white-space:nowrap;}}

  /* Print */
  @page{{size:landscape;margin:.65in;}}
  @media print{{
    body{{background:#fff!important;color:#111!important;}}
    .header,.filters,.stats,.charts,.section,.print-hide{{display:none!important;}}
    #print-header{{display:block!important;}}
    #print-stats{{display:flex!important;gap:40px;flex-wrap:wrap;margin-bottom:18px;padding-bottom:16px;border-bottom:2px solid #dde4f0;}}
    body.print-no-summary #print-stats{{display:none!important;}}
    #print-roster-wrap{{display:block!important;}}
    .ptable{{width:100%;border-collapse:collapse;font-size:11px;}}
    .ptable th{{background:#f0f4ff;color:#111;font-weight:700;padding:5px 8px;border:1px solid #ccc;text-align:left;}}
    .ptable td{{padding:5px 8px;border:1px solid #ddd;vertical-align:middle;}}
    .ptable tr:nth-child(even) td{{background:#fafafa;}}
  }}
</style>
</head>
<body>

<!-- ── Header ──────────────────────────────────────────────────────────── -->
<div class="header">
  <div class="header-left">
    <div>
      <h1>{name} <span>Certification Dashboard</span></h1>
      <div class="header-date">Data as of {date_label}</div>
    </div>
  </div>
  <div style="display:flex;gap:8px;align-items:center;">
    <div class="export-drop print-hide" id="export-drop">
      <button class="btn-export" onclick="toggleExportDrop()">&#128438; Export &#9660;</button>
      <div class="export-menu" id="export-menu">
        <div class="export-parent">Full Report<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('full')">PDF</button><button class="export-item" onclick="runExportXLSX('full')">Excel</button></div></div>
        <div class="export-parent">Not Certified<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('not-certified')">PDF</button><button class="export-item" onclick="runExportXLSX('not-certified')">Excel</button></div></div>
        <div class="export-parent">Manager Summary<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('manager-summary')">PDF</button><button class="export-item" onclick="runExportXLSX('manager-summary')">Excel</button></div></div>
      </div>
    </div><span class="info-btn print-hide" onclick="showInfo(event,'export')">?</span>
    <a href="index.html?playbook=Healthcare" class="btn-theme print-hide" style="text-decoration:none;white-space:nowrap;">&#128200; Healthcare Playbook Metrics</a>
    <button class="btn-theme print-hide" id="btn-theme" onclick="toggleTheme()">&#9728; Light</button>
  </div>
</div>

<!-- ── Filters ─────────────────────────────────────────────────────────── -->
<div class="filters">
  <span class="filter-label">Status</span>
  <select id="f-status" onchange="applyFilters()">
    <option value="">All</option>
    <option value="Certified">Certified</option>
    <option value="In Progress">In Progress</option>
    <option value="Not Started">Not Started</option>
  </select>
  <span class="filter-label">Market</span>
  <select id="f-market" onchange="applyFilters()"><option value="">All Markets</option></select>
  <span class="filter-label" style="margin-right:2px">Cert Date From</span>
  <input type="date" id="f-date-from" onchange="applyFilters()">
  <span class="filter-label" style="margin:0 2px">To</span>
  <input type="date" id="f-date-to" onchange="applyFilters()">
  <button class="btn-reset" onclick="resetFilters()">Reset</button>
  <span class="result-count" id="result-count"></span>
</div>

<!-- ── Stat cards ──────────────────────────────────────────────────────── -->
<div class="stats">
  <div class="stat">
    <div class="stat-label">Total Enrolled <span class="info-btn" onclick="showInfo(event,'total-enrolled')">?</span></div>
    <div class="stat-value" id="s-total">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">Certified <span class="info-btn" onclick="showInfo(event,'certified')">?</span></div>
    <div class="stat-value green" id="s-certified">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">In Progress <span class="info-btn" onclick="showInfo(event,'in-progress')">?</span></div>
    <div class="stat-value blue" id="s-inprog">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">Not Started <span class="info-btn" onclick="showInfo(event,'not-started')">?</span></div>
    <div class="stat-value red" id="s-notstarted">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">Completion Rate <span class="info-btn" onclick="showInfo(event,'completion-rate')">?</span></div>
    <div class="stat-value teal" id="s-rate">&#8212;</div>
    <div class="stat-sub" id="s-rate-sub"></div>
  </div>
</div>

<!-- ── Charts ──────────────────────────────────────────────────────────── -->
<div class="charts">
  <div class="chart-card">
    <div class="chart-title">Certification Pipeline <span class="info-btn" onclick="showInfo(event,'pipeline-chart')">?</span></div>
    <div class="chart-wrap"><canvas id="pipelineChart"></canvas></div>
  </div>
  <div class="chart-card">
    <div class="chart-title">Learners by Market <span class="info-btn" onclick="showInfo(event,'market-chart')">?</span></div>
    <div class="chart-wrap"><canvas id="marketChart"></canvas></div>
  </div>
  <div class="chart-card">
    <div class="chart-title">Certifications Over Time <span class="info-btn" onclick="showInfo(event,'trend-chart')">?</span></div>
    <div class="chart-wrap" style="position:relative;"><canvas id="trendChart"></canvas><div id="trend-empty" style="display:none;position:absolute;inset:0;display:flex;align-items:center;justify-content:center;flex-direction:column;gap:6px;pointer-events:none;"><span style="font-size:28px;">📋</span><span style="font-size:12px;color:var(--muted);">No certifications yet</span></div></div>
  </div>
</div>

<!-- ── Roster ──────────────────────────────────────────────────────────── -->
<div class="section">
  <div class="section-header">
    <div>
      <div class="section-title">Certification Roster <span class="info-btn" onclick="showInfo(event,'roster')">?</span></div>
      <div class="section-hint">Click a person to see their course-level progress</div>
    </div>
    <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;">
      <span style="font-size:11px;color:var(--muted);margin-right:2px;">View:</span>
      <button class="sort-btn active" id="view-individual" onclick="setRosterView('individual')">Individual</button>
      <button class="sort-btn" id="view-manager" onclick="setRosterView('manager')">By Manager</button>
      <span id="sort-controls" style="display:flex;align-items:center;gap:6px;margin-left:6px;">
        <span style="font-size:11px;color:var(--muted);margin-right:2px;">Sort:</span>
        <button class="sort-btn" data-sort="name" onclick="setSort('name')">Name</button>
        <button class="sort-btn" data-sort="status" onclick="setSort('status')">Status</button>
        <button class="sort-btn active" data-sort="pct" onclick="setSort('pct')">Completion % &#9660;</button>
      </span>
    </div>
    <input type="text" id="f-search" class="roster-search" placeholder="Search by name&hellip;" oninput="applyFilters()" style="width:170px;">
  </div>
  <div class="roster-key">
    <span><span class="key-stripe green"></span>Certified</span>
    <span><span class="key-stripe yellow"></span>In Progress</span>
    <span><span class="key-stripe red"></span>Not Started</span>
  </div>
  <div class="roster-wrap">
    <div class="roster-left" id="roster-left"></div>
    <div class="roster-right" id="roster-right">
      <div class="no-data">Select a person to view details</div>
    </div>
  </div>
</div>

<!-- ── Info popover ────────────────────────────────────────────────────── -->
<div class="info-popover" id="info-popover"></div>

<!-- ── Print-only elements ────────────────────────────────────────────── -->
<div id="print-header" style="display:none;margin-bottom:20px;">
  <div style="font-size:20px;font-weight:700;margin-bottom:4px;" id="ph-title"></div>
  <div style="font-size:12px;color:#555;margin-bottom:2px;" id="ph-date"></div>
  <div style="font-size:12px;color:#555;" id="ph-filters"></div>
  <div id="ph-desc" style="display:none;margin-top:10px;padding-top:10px;border-top:1px solid #dde4f0;font-size:12px;color:#444;font-style:italic;"></div>
</div>
<div id="print-stats" style="display:none;gap:40px;flex-wrap:wrap;margin-bottom:18px;padding-bottom:16px;border-bottom:2px solid #dde4f0;"></div>
<div id="print-roster-wrap" style="display:none;">
  <table class="ptable" id="print-roster-table">
    <thead id="print-roster-head"></thead>
    <tbody id="print-roster-body"></tbody>
  </table>
</div>

<script>
const PEOPLE = {raw_json};
const TLG_SET = new Set({tlg_json});

let filtered = [];

let sortField = "pct";
let sortDir   = "desc";
let selectedEmail = null;
let rosterView = "individual";
let pipelineChart, trendChart, marketChart;

function sel(id) {{ return document.getElementById(id); }}
function cv(v)  {{ return getComputedStyle(document.body).getPropertyValue(v).trim(); }}

// ── Theme ──────────────────────────────────────────────────────────────────
(function(){{
  if(localStorage.getItem("pb-theme") === "light") document.body.classList.add("light-mode");
  sel("btn-theme").textContent = document.body.classList.contains("light-mode") ? "🌙 Dark" : "☀ Light";
}})();
function toggleTheme(){{
  var light = document.body.classList.toggle("light-mode");
  localStorage.setItem("pb-theme", light ? "light" : "dark");
  sel("btn-theme").textContent = light ? "🌙 Dark" : "☀ Light";
  applyFilters();
}}

document.addEventListener("click", function(e){{
  if(!e.target.classList.contains("info-btn")) sel("info-popover").classList.remove("visible");
}});

// ── Export dropdown ────────────────────────────────────────────────────────
function toggleExportDrop(){{
  sel("export-menu").classList.toggle("open");
}}
document.addEventListener("click", function(e){{
  var d = sel("export-drop");
  if(d && !d.contains(e.target)) sel("export-menu").classList.remove("open");
}});

// ── Info tooltip ───────────────────────────────────────────────────────────
var INFO_MSGS = {{
  "total-enrolled":  "The total number of people currently assigned to the Healthcare certification program. Use the Status and Market filters above to narrow this to a specific group.",
  "certified":       "People who have completed all required coursework and earned their Healthcare certification. Certification is confirmed by the LMS system.",
  "in-progress":     "People who have started the program and completed at least one course, but haven't finished everything yet.",
  "not-started":     "People who are assigned to the program but haven't completed any courses yet.",
  "completion-rate": "The percentage of assigned people who have earned full certification so far. For example, 25% means 1 in 4 people is certified. This number updates when you apply filters.",
  "pipeline-chart":  "A quick snapshot of where everyone stands: how many haven't started yet, how many are actively working through the courses, and how many have finished and are fully certified.",
  "trend-chart":     "Shows how many people earned their Healthcare certification in each quarter. KM's fiscal year runs April through March. Q1 is April to June, Q2 is July to September, Q3 is October to December, and Q4 is January to March.",
  "market-chart":    "Certification progress broken down by sales market. Each bar shows how many people in that market are Certified, In Progress, or Not Started. Hover over any bar to see the exact counts and total enrolled for that market. Updates when you apply filters.",
  "roster":          "The full list of people in the program. Each card shows their name and job title, their progress on Layered Security and HC Foundations (bottom left), their overall completion percentage (bottom right), and their current status (top right). Click any card to see a full course-by-course breakdown in the panel on the right.",
  "export":          "Download a report based on whoever is currently shown on screen. Apply filters first to scope the report to a specific group. Full Report includes everyone with all course progress columns. Not Certified is a contact list of people still working through the program, sorted by manager, useful for follow-up. Manager Summary shows each manager's team size and how many on their team have certified."
}};
function showInfo(e, key){{
  var pop = sel("info-popover");
  pop.textContent = INFO_MSGS[key] || "";
  pop.classList.add("visible");
  var r = e.target.getBoundingClientRect();
  pop.style.top  = (r.bottom + 6) + "px";
  pop.style.left = Math.min(r.left, window.innerWidth - 280) + "px";
  e.stopPropagation();
}}

// ── JS helpers ─────────────────────────────────────────────────────────────
function escHtml(s){{ return String(s).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;").replace(/"/g,"&quot;"); }}
function personStatus(p){{
  if(p.Certified === "Yes") return "Certified";
  if(p.overallPct > 0)      return "In Progress";
  return "Not Started";
}}
function pFiscalQtr(d){{
  if(!d) return null;
  var pts = d.split("-"), yr = +pts[0], mo = +pts[1];
  var fy = mo >= 4 ? yr + 1 : yr, q = mo >= 10 ? 3 : mo >= 7 ? 2 : mo >= 4 ? 1 : 4;
  return "Q" + q + " FY" + String(fy).slice(2);
}}
function fmtDate(d){{
  if(!d) return "-";
  var pts = d.split("-"), months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"];
  return months[parseInt(pts[1]) - 1] + " " + parseInt(pts[2]) + ", " + pts[0];
}}
function pillClass(pct, total){{
  if(pct === 100 || (total > 0 && pct >= 100)) return "green";
  if(pct > 0) return "yellow";
  return "red";
}}

// ── Populate market dropdown ───────────────────────────────────────────────
(function(){{
  var markets = [...new Set(PEOPLE.map(function(p){{ return p.Market; }}).filter(Boolean))].sort();
  markets.forEach(function(m){{
    sel("f-market").innerHTML += '<option value="' + m + '">' + m + "</option>";
  }});
}})();

function resetFilters(){{
  sel("f-market").value = "";
  sel("f-status").value = "";
  sel("f-date-from").value = "";
  sel("f-date-to").value = "";
  sel("f-search").value = "";
  applyFilters();
}}

// ── Roster view toggle ─────────────────────────────────────────────────────
function setRosterView(v){{
  rosterView = v;
  sel("view-individual").classList.toggle("active", v === "individual");
  sel("view-manager").classList.toggle("active", v === "manager");
  sel("sort-controls").style.display = v === "individual" ? "flex" : "none";
  selectedEmail = null;
  sel("roster-right").innerHTML = '<div class="no-data">Select a person to view details</div>';
  renderRoster();
}}
function toggleMgrGroup(el){{
  el.classList.toggle("open");
  el.nextElementSibling.classList.toggle("open");
}}

// ── Sort ───────────────────────────────────────────────────────────────────
function setSort(field){{
  if(sortField === field){{
    sortDir = sortDir === "desc" ? "asc" : "desc";
  }} else {{
    sortField = field;
    sortDir   = "desc";
  }}
  document.querySelectorAll(".sort-btn").forEach(function(btn){{
    var active = btn.dataset.sort === sortField;
    btn.classList.toggle("active", active);
    var arrow = active ? (sortDir === "desc" ? " ↓" : " ↑") : "";
    btn.textContent = btn.textContent.replace(/ [↑↓]$/, "") + arrow;
  }});
  renderRoster();
}}

// ── applyFilters ───────────────────────────────────────────────────────────
function applyFilters(){{
  var market = sel("f-market").value;
  var status = sel("f-status").value;
  var from   = sel("f-date-from").value;
  var to     = sel("f-date-to").value;
  var q      = (sel("f-search").value || "").toLowerCase();
  filtered = PEOPLE.filter(function(p){{
    if(TLG_SET.has(p.FirstName + " " + p.LastName)) return false;
    if(market && p.Market !== market) return false;
    if(status && personStatus(p) !== status) return false;
    if(from && p.CertDate && p.CertDate < from) return false;
    if(to && p.CertDate && p.CertDate > to) return false;
    if(q && !(p.FirstName + " " + p.LastName).toLowerCase().includes(q)) return false;
    return true;
  }});
  sel("result-count").textContent = filtered.length + " shown";
  renderStats();
  renderCharts();
  renderRoster();
}}

// ── renderStats ────────────────────────────────────────────────────────────
function renderStats(){{
  var total   = filtered.length;
  var cert    = filtered.filter(function(p){{ return p.Certified === "Yes"; }}).length;
  var inprog  = filtered.filter(function(p){{ return p.overallPct > 0 && p.Certified !== "Yes"; }}).length;
  var nostart = filtered.filter(function(p){{ return p.overallPct === 0 && p.Certified !== "Yes"; }}).length;
  var rate    = total > 0 ? Math.round(cert / total * 100) : 0;
  sel("s-total").textContent     = total;
  sel("s-certified").textContent = cert;
  sel("s-inprog").textContent    = inprog;
  sel("s-notstarted").textContent= nostart;
  sel("s-rate").textContent      = rate + "%";
  sel("s-rate-sub").textContent  = total > 0 ? (cert + " of " + total + " enrolled") : "";
}}

// ── renderCharts ───────────────────────────────────────────────────────────
function renderCharts(){{
  var isLight    = document.body.classList.contains("light-mode");
  var labelColor = isLight ? cv("--text") : cv("--muted");
  var gridColor  = cv("--border");

  // Chart 1: certification pipeline — Not Started, In Progress, Certified
  var pipelineNotStarted = filtered.filter(function(p){{ return personStatus(p) === "Not Started"; }}).length;
  var pipelineInProgress = filtered.filter(function(p){{ return personStatus(p) === "In Progress"; }}).length;
  var pipelineCertified  = filtered.filter(function(p){{ return personStatus(p) === "Certified";   }}).length;

  if(pipelineChart) pipelineChart.destroy();
  pipelineChart = new Chart(sel("pipelineChart"), {{
    type: "bar",
    data: {{
      labels: ["Not Started", "In Progress", "Certified"],
      datasets: [{{
        data: [pipelineNotStarted, pipelineInProgress, pipelineCertified],
        backgroundColor: [cv("--red") + "cc", "#f59e0b" + "cc", cv("--green") + "cc"],
        borderRadius: 4,
        borderSkipped: false
      }}]
    }},
    options: {{
      indexAxis: "y",
      responsive: true, maintainAspectRatio: false,
      plugins: {{
        legend: {{ display: false }},
        tooltip: {{ callbacks: {{ label: function(ctx){{ return " " + ctx.raw + " people"; }} }} }}
      }},
      scales: {{
        x: {{ grid: {{ color: gridColor }}, ticks: {{ color: labelColor, font: {{ size: 11 }}, stepSize: 1 }} }},
        y: {{ grid: {{ display: false }}, ticks: {{ color: labelColor, font: {{ size: 12 }} }} }}
      }}
    }}
  }});

  // Chart 2: certifications over time by fiscal quarter
  var qtrMap = {{}};
  filtered.forEach(function(p){{
    if(p.Certified === "Yes" && p.CertQtr){{
      qtrMap[p.CertQtr] = (qtrMap[p.CertQtr] || 0) + 1;
    }}
  }});
  function parseQtr(s){{ var m = s.match(/FY(\d+)\s+Q(\d)/); return m ? +m[1] * 10 + +m[2] : 0; }}
  var trendLabels = Object.keys(qtrMap).sort(function(a,b){{ return parseQtr(a) - parseQtr(b); }});
  var trendData   = trendLabels.map(function(q){{ return qtrMap[q]; }});
  var trendEmpty = sel("trend-empty");
  if(trendEmpty) trendEmpty.style.display = trendLabels.length === 0 ? "flex" : "none";

  if(trendChart) trendChart.destroy();
  trendChart = new Chart(sel("trendChart"), {{
    type: "bar",
    data: {{
      labels: trendLabels,
      datasets: [{{
        data: trendData,
        backgroundColor: cv("--green") + "bb",
        borderRadius: 3,
        borderSkipped: false
      }}]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      plugins: {{
        legend: {{ display: false }},
        tooltip: {{ callbacks: {{ label: function(ctx){{ return ctx.raw + " certified this quarter"; }} }} }}
      }},
      scales: {{
        x: {{ grid: {{ color: gridColor }}, ticks: {{ color: labelColor, font: {{ size: 10 }}, maxRotation: 45 }} }},
        y: {{ grid: {{ color: gridColor }}, ticks: {{ color: labelColor, font: {{ size: 11 }}, stepSize: 1 }} }}
      }}
    }}
  }});

  // Chart 3: completion by market — stacked horizontal bar
  var mktMap = {{}};
  filtered.forEach(function(p){{
    var m = p.Market || "Unknown";
    var s = personStatus(p);
    if(!mktMap[m]) mktMap[m] = {{ Certified: 0, "In Progress": 0, "Not Started": 0 }};
    mktMap[m][s] = (mktMap[m][s] || 0) + 1;
  }});
  var mktLabels = Object.keys(mktMap).sort(function(a,b){{
    var ta = mktMap[a].Certified + mktMap[a]["In Progress"] + mktMap[a]["Not Started"];
    var tb = mktMap[b].Certified + mktMap[b]["In Progress"] + mktMap[b]["Not Started"];
    return tb - ta;
  }});

  if(marketChart) marketChart.destroy();
  marketChart = new Chart(sel("marketChart"), {{
    type: "bar",
    data: {{
      labels: mktLabels,
      datasets: [
        {{ label: "Certified",    data: mktLabels.map(function(m){{ return mktMap[m].Certified;      }}), backgroundColor: cv("--green") + "cc", borderWidth: 0, borderRadius: 2 }},
        {{ label: "In Progress",  data: mktLabels.map(function(m){{ return mktMap[m]["In Progress"]; }}), backgroundColor: "#f59e0b"       + "cc", borderWidth: 0, borderRadius: 2 }},
        {{ label: "Not Started",  data: mktLabels.map(function(m){{ return mktMap[m]["Not Started"]; }}), backgroundColor: cv("--red")     + "cc", borderWidth: 0, borderRadius: 2 }}
      ]
    }},
    options: {{
      indexAxis: "y",
      responsive: true, maintainAspectRatio: false,
      plugins: {{
        legend: {{ display: true, position: "bottom", labels: {{ color: labelColor, font: {{ size: 11 }}, padding: 16, boxWidth: 12 }} }},
        tooltip: {{
          mode: "index",
          callbacks: {{
            title: function(items){{ return items[0].label; }},
            afterBody: function(items){{
              var total = items.reduce(function(s, i){{ return s + i.raw; }}, 0);
              return ["─────────────", "Enrolled: " + total];
            }}
          }}
        }},
        datalabels: {{ display: false }}
      }},
      scales: {{
        x: {{ stacked: true, grid: {{ color: gridColor }}, ticks: {{ color: labelColor, font: {{ size: 11 }}, stepSize: 1 }} }},
        y: {{ stacked: true, grid: {{ display: false }}, ticks: {{ color: labelColor, font: {{ size: 11 }} }} }}
      }}
    }}
  }});
}}

// ── buildPersonCard ────────────────────────────────────────────────────────
function buildPersonCard(p){{
  var fullName       = p.FirstName + " " + p.LastName;
  var status         = personStatus(p);
  var stripe         = status === "Certified" ? " stripe-green" : status === "In Progress" ? " stripe-yellow" : " stripe-red";
  var hcfClass       = pillClass(p.hcf.pct, p.hcf.total);
  var lsClass        = pillClass(p.ls.pct,  p.ls.total);
  var pctClass       = status === "Certified" ? " pct-done" : status === "In Progress" ? " pct-progress" : "";
  var badgeStatusCls = status === "Certified" ? "certified" : status === "In Progress" ? "in-progress" : "not-certified";
  var badgeStatusTxt = status === "Certified" ? "✓ Certified" : status === "In Progress" ? "In Progress" : "Not Started";
  var h = '<div class="roster-person' + stripe + '" data-email="' + escHtml(p.Email) + '" onclick="showDetail(this.dataset.email)">';
  h += '<div class="roster-top">';
  h += '<div class="roster-name-block">';
  h += '<div class="roster-name">' + fullName + "</div>";
  h += '<div class="roster-title">' + (p.JobTitle || "") + "</div>";
  h += "</div>";
  h += '<span class="badge-status ' + badgeStatusCls + '">' + badgeStatusTxt + "</span>";
  h += "</div>";
  h += '<div class="roster-bottom">';
  h += '<div class="roster-pills">';
  h += '<span class="pill ' + lsClass  + '">Layered Sec ' + p.ls.done + "/11</span>";
  h += '<span class="pill ' + hcfClass + '">HC Foundations ' + p.hcf.done + "/10</span>";
  h += "</div>";
  h += '<span class="roster-pct' + pctClass + '">' + p.overallPct + "%</span>";
  h += "</div></div>";
  return h;
}}

// ── renderRoster ───────────────────────────────────────────────────────────
function renderRoster(){{
  var html = "";

  if(rosterView === "manager"){{
    var groups = {{}};
    filtered.forEach(function(p){{
      var mgr = p.Manager || "No Manager";
      if(!groups[mgr]) groups[mgr] = [];
      groups[mgr].push(p);
    }});
    var mgrsSorted = Object.keys(groups).sort();
    mgrsSorted.forEach(function(mgr){{
      var team = groups[mgr];
      var certCount = team.filter(function(p){{ return p.Certified === "Yes"; }}).length;
      var teamCards = team.slice()
        .sort(function(a, b){{ return (a.LastName + a.FirstName).localeCompare(b.LastName + b.FirstName); }})
        .map(buildPersonCard).join("");
      html += "<div>";
      html += '<div class="mgr-group-hdr open" onclick="toggleMgrGroup(this)">';
      html += "<div>";
      html += '<div style="font-size:12px;font-weight:600;color:var(--text)">' + mgr + "</div>";
      html += '<div style="font-size:11px;color:var(--muted);margin-top:2px;">' + team.length + " rep" + (team.length !== 1 ? "s" : "") + " &middot; " + certCount + " certified</div>";
      html += "</div>";
      html += '<span class="mgr-chevron">&#9660;</span>';
      html += "</div>";
      html += '<div class="mgr-team open">' + teamCards + "</div>";
      html += "</div>";
    }});
  }} else {{
    var d = sortDir === "desc" ? -1 : 1;
    var sorted = filtered.slice().sort(function(a, b){{
      if(sortField === "name"){{
        return d * (a.LastName + a.FirstName).localeCompare(b.LastName + b.FirstName);
      }} else if(sortField === "status"){{
        var order = {{ "Certified": 2, "In Progress": 1, "Not Started": 0 }};
        var diff = (order[personStatus(a)] || 0) - (order[personStatus(b)] || 0);
        if(diff !== 0) return d * diff;
        return (a.LastName + a.FirstName).localeCompare(b.LastName + b.FirstName);
      }} else {{
        var diff2 = a.overallPct - b.overallPct;
        if(diff2 !== 0) return d * diff2;
        return (a.LastName + a.FirstName).localeCompare(b.LastName + b.FirstName);
      }}
    }});
    sorted.forEach(function(p){{ html += buildPersonCard(p); }});
  }}

  if(!html) {{
    var _df = sel("f-date-from").value, _dt = sel("f-date-to").value;
    if(_df || _dt) {{
      html = '<div class="no-data">No certifications found in this date range.<br><span style="font-size:0.85em;opacity:0.75">Try a different range or click Reset to clear all filters.</span></div>';
    }} else {{
      html = '<div class="no-data">No people match the selected filters.</div>';
    }}
  }}
  sel("roster-left").innerHTML = html;

  if(selectedEmail){{
    var el = sel("roster-left").querySelector('[data-email="' + selectedEmail + '"]');
    if(el) el.classList.add("active");
    else   sel("roster-right").innerHTML = '<div class="no-data">Select a person to view details</div>';
  }}
}}

// ── showDetail ─────────────────────────────────────────────────────────────
function showDetail(email){{
  selectedEmail = email;
  var p = PEOPLE.find(function(r){{ return r.Email === email; }});
  if(!p) return;
  document.querySelectorAll(".roster-person").forEach(function(el){{
    el.classList.toggle("active", el.dataset.email === email);
  }});
  var status  = personStatus(p);
  var isCert  = status === "Certified";
  var badgeClass = status === "Certified" ? "certified" : status === "In Progress" ? "in-progress" : "not-certified";
  var badgeText  = status === "Certified" ? ("✓ Certified" + (p.CertDate ? " · " + fmtDate(p.CertDate) : "")) : status === "In Progress" ? "In Progress" : "Not Started";

  function progBar(pct, cls){{
    return '<div class="prog-wrap"><div class="prog-bar ' + cls + '" style="width:' + pct + '%"></div></div>';
  }}
  function courseList(curriculum, id){{
    var items = curriculum.items;
    var html2 = '<div class="course-list open" id="cl-' + id + '">';
    items.forEach(function(item){{
      var iconClass = item.done ? "done" : "todo";
      var icon      = item.done ? "✓" : "○";
      var dateStr   = item.done && item.date ? '<span class="course-date">' + fmtDate(item.date) + "</span>" : "";
      html2 += '<div class="course-item">';
      html2 += '<span class="course-icon ' + iconClass + '">' + icon + "</span>";
      html2 += '<span class="course-title">' + item.title + "</span>";
      html2 += dateStr;
      html2 += "</div>";
    }});
    html2 += "</div>";
    return html2;
  }}
  function curriculumSection(label, curriculum, id){{
    var doneOf = curriculum.done + " / " + curriculum.total + " courses";
    var pct    = curriculum.pct;
    var barCls = pct >= 100 ? "green" : "blue";
    var html3  = '<div class="curriculum-section">';
    html3 += '<div class="curriculum-header" data-id="' + id + '" onclick="toggleCourseList(this.dataset.id)">';
    html3 += '<span class="curriculum-title">' + label + "</span>";
    html3 += '<span class="curriculum-count">' + doneOf + " ▼</span>";
    html3 += "</div>";
    html3 += progBar(pct, barCls);
    html3 += courseList(curriculum, id);
    html3 += "</div>";
    return html3;
  }}

  var detailHtml = "";
  detailHtml += '<div class="roster-right-header">';
  detailHtml += '<div style="font-size:16px;font-weight:700;margin-bottom:6px">' + p.FirstName + " " + p.LastName + "</div>";
  detailHtml += '<span class="badge-status ' + badgeClass + '">' + badgeText + "</span>";
  detailHtml += "</div>";
  detailHtml += '<div class="detail-grid">';
  detailHtml += '<div><div class="detail-label">Job Title</div><div class="detail-value">' + (p.JobTitle || "-") + "</div></div>";
  detailHtml += '<div><div class="detail-label">Market</div><div class="detail-value">' + (p.Market || "-") + "</div></div>";
  detailHtml += '<div><div class="detail-label">Hired</div><div class="detail-value">' + fmtDate(p.HireDate) + "</div></div>";
  detailHtml += '<div><div class="detail-label">Email</div><div class="detail-value"><a href="mailto:' + p.Email + '" style="color:var(--accent);text-decoration:none">' + (p.Email || "-") + "</a></div></div>";
  if(p.Manager){{
    detailHtml += '<div><div class="detail-label">Manager</div><div class="detail-value">' + p.Manager + "</div></div>";
    detailHtml += '<div><div class="detail-label">Manager Email</div><div class="detail-value"><a href="mailto:' + p.MgrEmail + '" style="color:var(--accent);text-decoration:none">' + (p.MgrEmail || "-") + "</a></div></div>";
  }}
  detailHtml += "</div>";
  detailHtml += '<hr style="border:none;border-top:1px solid var(--border);margin:16px 0;">';
  detailHtml += curriculumSection("Layered Security · " + p.ls.done + " / 11 courses", p.ls, "ls-" + email.replace(/[^a-z0-9]/gi, ""));
  detailHtml += curriculumSection("Healthcare Foundations · " + p.hcf.done + " / 10 courses", p.hcf, "hcf-" + email.replace(/[^a-z0-9]/gi, ""));
  detailHtml += '<div style="margin-top:14px;padding-top:12px;border-top:1px solid var(--border);font-size:12px;color:var(--muted);">';
  detailHtml += p.overallDone + " of 21 courses complete (" + p.overallPct + "%)";
  detailHtml += "</div>";
  sel("roster-right").innerHTML = detailHtml;
}}

function toggleCourseList(id){{
  var el = document.getElementById("cl-" + id);
  if(el) el.classList.toggle("open");
}}

// ── Print / Export ─────────────────────────────────────────────────────────
function setupPrintHeader(title, subtitle){{
  sel("ph-title").textContent   = title;
  sel("ph-date").textContent    = subtitle;
  var market = sel("f-market").value || "All Markets";
  var status = sel("f-status").options[sel("f-status").selectedIndex].text;
  var search = sel("f-search").value;
  var parts  = ["Status: " + status, "Market: " + market];
  if(search) parts.push("Search: " + search);
  sel("ph-filters").textContent = parts.join("  |  ");
}}
function pBox(n, l){{
  return '<div style="min-width:90px"><div style="font-size:30px;font-weight:700;color:#1a3a5c;line-height:1">' + n + "</div>"
       + '<div style="font-size:10px;color:#666;text-transform:uppercase;letter-spacing:.06em;margin-top:5px">' + l + "</div></div>";
}}
function tds(cells){{ return "<tr>" + cells.map(function(c){{ return "<td>" + c + "</td>"; }}).join("") + "</tr>"; }}
function thRow(labels){{ return "<tr>" + labels.map(function(l){{ return "<th>" + l + "</th>"; }}).join("") + "</tr>"; }}

function runExport(type){{
  sel("export-menu").classList.remove("open");
  var now = new Date().toLocaleDateString("en-US", {{year:"numeric",month:"long",day:"numeric"}});

  if(type === "full"){{
    setupPrintHeader("{name} Certification Report", "Generated: " + now + "  |  " + filtered.length + " People");
    var total  = filtered.length;
    var cert   = filtered.filter(function(p){{ return p.Certified === "Yes"; }}).length;
    var inprog = filtered.filter(function(p){{ return p.overallPct > 0 && p.Certified !== "Yes"; }}).length;
    var rate   = total > 0 ? Math.round(cert / total * 100) : 0;
    sel("print-stats").innerHTML = pBox(total,"Total Enrolled") + pBox(cert,"Certified") + pBox(inprog,"In Progress") + pBox(rate+"%","Completion Rate");
    sel("print-roster-head").innerHTML = thRow(["#","Name","Market","Job Title","Status","HC Foundations %","Layered Security %","Overall %","Cert Date","Manager"]);
    sel("print-roster-body").innerHTML = filtered.map(function(p,i){{
      return tds([i+1, "<b>"+p.FirstName+" "+p.LastName+"</b>", p.Market||"-", p.JobTitle||"-",
        personStatus(p), p.hcf.pct+"%", p.ls.pct+"%", p.overallPct+"%",
        p.CertDate||"-", p.Manager||"-"]);
    }}).join("");
    sel("ph-desc").style.display = "none";
    document.body.classList.remove("print-no-summary");
    window.print();

  }} else if(type === "not-certified"){{
    var notCert = filtered.filter(function(p){{ return p.Certified !== "Yes"; }});
    setupPrintHeader("Not Certified: {name}", "Generated: " + now + "  |  " + notCert.length + " Employees");
    sel("print-stats").innerHTML = "";
    sel("print-roster-head").innerHTML = thRow(["#","Name","Email","Market","HC Foundations %","Layered Security %","Manager","Manager Email"]);
    sel("print-roster-body").innerHTML = notCert.length
      ? notCert.slice().sort(function(a,b){{ return (a.Manager||"").localeCompare(b.Manager||"") || (a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName); }})
          .map(function(p,i){{ return tds([i+1,"<b>"+p.FirstName+" "+p.LastName+"</b>",p.Email||"-",p.Market||"-",p.hcf.pct+"%",p.ls.pct+"%",p.Manager||"-",p.MgrEmail||"-"]); }}).join("")
      : '<tr><td colspan="8" style="color:#999;font-style:italic;padding:10px">All enrolled people are certified.</td></tr>';
    sel("ph-desc").textContent = "Employees who have not yet earned Healthcare certification, sorted by manager.";
    sel("ph-desc").style.display = "block";
    document.body.classList.add("print-no-summary");
    window.print();
    document.body.classList.remove("print-no-summary");

  }} else if(type === "manager-summary"){{
    var mgrMap = {{}};
    filtered.forEach(function(p){{
      var k = p.Manager || "(No Manager)";
      if(!mgrMap[k]) mgrMap[k] = {{ name:k, email:p.MgrEmail||"-", total:0, cert:0, avgPct:0, sumPct:0 }};
      mgrMap[k].total++;
      if(p.Certified === "Yes") mgrMap[k].cert++;
      mgrMap[k].sumPct += p.overallPct;
    }});
    var mgrs = Object.values(mgrMap).map(function(m){{ m.avgPct = m.total > 0 ? Math.round(m.sumPct / m.total) : 0; return m; }})
      .sort(function(a,b){{ return (b.cert/b.total) - (a.cert/a.total); }});
    setupPrintHeader("Manager Summary: {name}", "Generated: " + now + "  |  " + mgrs.length + " Managers");
    sel("print-stats").innerHTML = "";
    sel("print-roster-head").innerHTML = thRow(["Manager","Manager Email","Team Size","Certified","Avg Overall %"]);
    sel("print-roster-body").innerHTML = mgrs.map(function(m){{ return tds(["<b>"+m.name+"</b>",m.email,m.total,m.cert,"<b>"+m.avgPct+"%</b>"]); }}).join("");
    sel("ph-desc").textContent = "Certification completion by manager, sorted from highest to lowest completion rate.";
    sel("ph-desc").style.display = "block";
    document.body.classList.add("print-no-summary");
    window.print();
    document.body.classList.remove("print-no-summary");
  }}
}}

// ── xlsx export ────────────────────────────────────────────────────────────
function runExportXLSX(type){{
  sel("export-menu").classList.remove("open");
  var now=new Date().toLocaleDateString("en-US",{{year:"numeric",month:"long",day:"numeric"}});
  function makeSheet(rows,colWidths){{
    var ws=XLSX.utils.aoa_to_sheet(rows);
    ws['!cols']=colWidths.map(function(w){{ return {{wch:w}}; }});
    return ws;
  }}
  function dlXLSX(name,wb){{ XLSX.writeFile(wb,name+'.xlsx'); }}
  var wb=XLSX.utils.book_new();
  if(type==="full"){{
    var total=filtered.length;
    var cert=filtered.filter(function(p){{ return p.Certified==="Yes"; }}).length;
    var rate=total>0?Math.round(cert/total*100):0;
    var inprogXL  = filtered.filter(function(p){{ return p.overallPct > 0 && p.Certified !== "Yes"; }}).length;
    var nostartXL = filtered.filter(function(p){{ return p.overallPct === 0 && p.Certified !== "Yes"; }}).length;
    var sumRows=[
      ["{name} Certification Report"],
      ["Generated:",now],
      [],
      ["SUMMARY"],
      ["Total Enrolled",total],
      ["Certified",cert],
      ["In Progress",inprogXL],
      ["Not Started",nostartXL],
      ["Completion Rate",rate+"%"],
    ];
    XLSX.utils.book_append_sheet(wb,makeSheet(sumRows,[30,18]),"Summary");
    var rRows=[["Name","Market","Job Title","Status","HC Foundations %","Layered Security %","Overall %","Cert Date","Manager"]];
    filtered.forEach(function(p){{
      rRows.push([p.FirstName+" "+p.LastName,p.Market||"-",p.JobTitle||"-",personStatus(p),p.hcf.pct+"%",p.ls.pct+"%",p.overallPct+"%",p.CertDate||"-",p.Manager||"-"]);
    }});
    XLSX.utils.book_append_sheet(wb,makeSheet(rRows,[28,18,28,14,16,16,12,14,28]),"Roster");
    dlXLSX("hc-full-report",wb);
  }} else if(type==="not-certified"){{
    var notCert=filtered.filter(function(p){{ return p.Certified!=="Yes"; }}).slice().sort(function(a,b){{
      return (a.Manager||"").localeCompare(b.Manager||"")||(a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName);
    }});
    var rows=[["Name","Email","Market","HC Foundations %","Layered Security %","Manager","Manager Email"]];
    notCert.forEach(function(p){{
      rows.push([p.FirstName+" "+p.LastName,p.Email||"-",p.Market||"-",p.hcf.pct+"%",p.ls.pct+"%",p.Manager||"-",p.MgrEmail||"-"]);
    }});
    XLSX.utils.book_append_sheet(wb,makeSheet(rows,[28,32,18,16,16,28,32]),"Not Certified");
    dlXLSX("hc-not-certified",wb);
  }} else if(type==="manager-summary"){{
    var mgrMap={{}};
    filtered.forEach(function(p){{
      var k=p.Manager||"(No Manager)";
      if(!mgrMap[k]) mgrMap[k]={{name:k,email:p.MgrEmail||"-",total:0,cert:0,sumPct:0}};
      mgrMap[k].total++;
      if(p.Certified==="Yes") mgrMap[k].cert++;
      mgrMap[k].sumPct+=p.overallPct;
    }});
    var rows=[["Manager","Manager Email","Team Size","Certified","Avg Overall %"]];
    Object.values(mgrMap).map(function(m){{ m.avgPct=m.total>0?Math.round(m.sumPct/m.total):0; return m; }})
      .sort(function(a,b){{ return (b.cert/b.total)-(a.cert/a.total); }})
      .forEach(function(m){{ rows.push([m.name,m.email,m.total,m.cert,m.avgPct+"%"]); }});
    XLSX.utils.book_append_sheet(wb,makeSheet(rows,[28,32,12,12,14]),"Manager Summary");
    dlXLSX("hc-manager-summary",wb);
  }}
}}

// ── init ───────────────────────────────────────────────────────────────────
applyFilters();
var firstCard = sel("roster-left").querySelector(".roster-person");
if(firstCard) showDetail(firstCard.dataset.email);
(function(){{
  var n=0,t;
  var h=document.querySelector(".header h1");
  if(h) h.addEventListener("click",function(){{
    n++;clearTimeout(t);
    if(n>=3){{n=0;window.location.href="index.html";}}
    else t=setTimeout(function(){{n=0;}},1500);
  }});
}})();
</script>
</body>
</html>"""


if __name__ == '__main__':
    main()
