import openpyxl, glob, os, re, json
from datetime import datetime

COL_FIRST     = 2
COL_LAST      = 3
COL_EMAIL     = 4
COL_JOBTITLE  = 5
COL_REGION    = 6
COL_MARKET    = 7
COL_MGR_FIRST = 9
COL_MGR_LAST  = 10
COL_MGR_EMAIL = 11
COL_MGR_TITLE = 12
COL_HIRE_DATE = 14
COL_CURRIC_ID = 17
COL_CURRIC_CMP = 19
COL_ASSIGN_DT = 20
COL_DAYS_REM  = 21
COL_ITEM_ID   = 22
COL_ITEM_TITLE = 26
COL_ITEM_DATE  = 27
COL_ITEM_STS   = 29   # "Online-Complete" = done

ITEM_ORDER = [
    "LSC_IAPO", "LS_ITM", "DS", "UA_ACCESSCONTROL", "BIZHUB_BSMFP",
    "MFPPT_BSBSNB", "BREACH_IDADB", "CSMWSG",
    "LSSW", "LSSB_HI", "LSSB_NONPROFIT", "LSSB_GOV"
]

ITEM_TITLES = {
    "LSC_IAPO":         "Layered Security Certification Introduction and Program Overview",
    "LS_ITM":           "Layered Security: Introducing the Model",
    "DS":               "Document Security",
    "UA_ACCESSCONTROL": "User Authentication and Access Control",
    "BIZHUB_BSMFP":     "bizhub Security at the MFP",
    "MFPPT_BSBSNB":     "MFP Protection Team: bizhub SECURE, bizhub SECURE Notifier, Bitdefender",
    "BREACH_IDADB":     "BreachAlert- Identify Document and Data Breaches at the Multi-Functional Device",
    "CSMWSG":           "Centralized Security Management with Shield Guard",
    "LSSW":             "Secure Workflow",
    "LSSB_HI":          "Layered Security Success Study - Healthcare Industry",
    "LSSB_NONPROFIT":   "Layered Security Success Study - Non-profit",
    "LSSB_GOV":         "Layered Security Success Study - Government",
}

TLG = {n.lower() for n in [
    "Jason Ackerman","Bianca Davis","James Parker","Resmie Biba",
    "Chris Curtis","Sara Thompson","Jeremy MacBean","Bradley Pierce",
    "Laura Sefcik","Samantha Maresca","Staci Musco","CJ Homer",
    "Rich Moore","Dale Kinsey","John Lechner","Resmie Nesimi",
    "Samantha D'Angelo","Bianca DiPasquale","Doug Falk"
]}


def _date(val):
    if val is None: return None
    if hasattr(val, 'strftime'): return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s in ('-', 'None'): return None
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m: return f"{int(m.group(3)):04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return s[:10] if len(s) >= 10 else s


def km_fiscal_quarter(date_str):
    if not date_str: return ''
    try:
        d = datetime.strptime(date_str[:10], '%Y-%m-%d')
    except ValueError:
        return ''
    m, y = d.month, d.year
    if m >= 4:
        fy = y + 1
        if   m <= 6:  q = 1
        elif m <= 9:  q = 2
        elif m <= 12: q = 3
        else:         q = 4
    else:
        fy = y
        q  = 4
    return f"Q{q} FY{str(fy)[-2:]}"


def _file_date_label(fname):
    base = os.path.basename(fname)
    # MM.DD.YYYY
    m = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', base)
    if m:
        months = ["January","February","March","April","May","June",
                  "July","August","September","October","November","December"]
        return f"{months[int(m.group(1))-1]} {int(m.group(2))}, {m.group(3)}"
    return base


def person_key(email, first, last):
    e = (email or '').strip().lower()
    if e:
        return e
    return f"{(first or '').strip().lower()} {(last or '').strip().lower()}"


CERT_FILE_GLOB  = 'cert-data/Layered-Security-Certification-Report-*.xlsx'
DEALS_FILE_GLOB = 'cert-data/Layered-Security-Certification-OPS-FY26-*.xlsx'
DEAL_MIN_AMOUNT = 5000


def _find_col(header, *substrings, exclude=None):
    """Find a column index by case-insensitive substring match on the header row."""
    for i, h in enumerate(header):
        hl = str(h or '').strip().lower()
        if exclude and exclude in hl:
            continue
        if any(s in hl for s in substrings):
            return i
    return None


def load_sales_cert():
    """Layered Security sales-certification report — one row per already-certified person
    (the report is scoped to certified people only, so presence in it = certified).
    Returns email(lower) -> {'certified': True, 'date': 'YYYY-MM-DD'}."""
    files = sorted(glob.glob(CERT_FILE_GLOB))
    if not files:
        return {}
    fpath = files[-1]
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    email_col = _find_col(header, 'email', exclude='manager')
    if email_col is None:
        print(f"WARNING: no Email column found in {os.path.basename(fpath)} — skipping sales certification data")
        return {}
    date_col = _find_col(header, 'completion date')
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        email = str(row[email_col] or '').strip().lower()
        if not email:
            continue
        d = _date(row[date_col]) if date_col is not None else None
        result[email] = {'certified': True, 'date': d or ''}
    return result


def load_sales_deals():
    """Layered Security OPS FY26 opportunity export — Closed Won deals of $5,000+ count toward
    the sales-proof piece of certification. Returns email(lower) -> {'closedWon': int,
    'amount': float, 'accountName': str, 'closeDate': str}."""
    files = sorted(glob.glob(DEALS_FILE_GLOB))
    if not files:
        return {}
    fpath = files[-1]
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    email_col = _find_col(header, 'email')
    stage_col = _find_col(header, 'stage')
    if email_col is None or stage_col is None:
        print(f"WARNING: missing Email or Stage column in {os.path.basename(fpath)} — skipping Closed Won deal data")
        return {}
    amount_col = _find_col(header, 'amount')
    acct_col   = _find_col(header, 'account name')
    close_col  = _find_col(header, 'close date')
    if amount_col is None:
        print(f"WARNING: no Amount column found in {os.path.basename(fpath)} — Closed Won $ will stay empty until one appears")
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        email = str(row[email_col] or '').strip().lower()
        stage = str(row[stage_col] or '').strip().lower()
        if not email:
            continue
        if email not in result:
            result[email] = {'closedWon': 0, 'amount': 0, 'accountName': '', 'closeDate': ''}
        if stage != 'closed won':
            continue
        amt = 0
        if amount_col is not None and row[amount_col] is not None:
            try:
                amt = float(re.sub(r'[^0-9.\-]', '', str(row[amount_col])))
            except ValueError:
                amt = 0
        if amt < DEAL_MIN_AMOUNT:
            continue
        result[email]['closedWon'] += 1
        result[email]['amount']    += amt
        if acct_col is not None and row[acct_col]:
            result[email]['accountName'] = str(row[acct_col])
        if close_col is not None and row[close_col]:
            result[email]['closeDate'] = _date(row[close_col]) or str(row[close_col])
    return result


def load_ls_data():
    files = sorted(glob.glob('cert-data/Layered-Security-Curricula-Report-*.xlsx'))
    if not files:
        print("No Layered Security data files found.")
        return [], "Unknown"

    date_label = _file_date_label(files[-1])
    seen = {}  # person_key -> row dict (latest file wins)

    for fpath in files:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb.active
        file_rows = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            email = str(row[COL_EMAIL] or '').strip()
            first = str(row[COL_FIRST] or '').strip()
            last  = str(row[COL_LAST]  or '').strip()
            key   = person_key(email, first, last)
            if not key: continue
            if key not in file_rows:
                file_rows[key] = {'meta': None, 'items': []}
            if row[COL_ITEM_ID]:
                file_rows[key]['items'].append(row)
            else:
                file_rows[key]['meta'] = row

        for key, data in file_rows.items():
            seen[key] = data   # later file overwrites earlier

    people = []
    for key, data in seen.items():
        meta = data['meta']
        if meta is None:
            continue  # shouldn't happen

        first = str(meta[COL_FIRST] or '').strip()
        last  = str(meta[COL_LAST]  or '').strip()
        title = str(meta[COL_JOBTITLE] or '').strip()
        market = str(meta[COL_MARKET] or '').strip()
        full_name = f"{first} {last}".strip()

        # Exclusions
        if full_name.lower() in TLG:
            continue
        if 'solutions consultant' in title.lower():
            continue
        if 'commercial print' in market.lower() or 'commercial print' in title.lower():
            continue

        email     = str(meta[COL_EMAIL] or '').strip()
        region    = str(meta[COL_REGION] or '').strip()
        mgr_first = str(meta[COL_MGR_FIRST] or '').strip()
        mgr_last  = str(meta[COL_MGR_LAST]  or '').strip()
        mgr_email = str(meta[COL_MGR_EMAIL] or '').strip()
        mgr_title = str(meta[COL_MGR_TITLE] or '').strip()
        hire_date  = _date(meta[COL_HIRE_DATE])
        assign_dt  = _date(meta[COL_ASSIGN_DT])
        days_rem   = meta[COL_DAYS_REM]
        if days_rem is not None:
            try:
                days_rem = int(days_rem)
            except (ValueError, TypeError):
                days_rem = None

        # Build items dict keyed by item ID
        item_map = {}
        for irow in data['items']:
            iid   = str(irow[COL_ITEM_ID]    or '').strip()
            ititle = str(irow[COL_ITEM_TITLE] or '').strip()
            idate  = _date(irow[COL_ITEM_DATE])
            idone  = str(irow[COL_ITEM_STS]  or '').strip() == 'Online-Complete'
            if iid:
                item_map[iid] = {'id': iid, 'title': ititle, 'done': idone, 'date': idate or ''}

        # Warn if file has items not in ITEM_ORDER (catches new courses on next data drop)
        for iid in item_map:
            if iid not in ITEM_ORDER:
                print(f"WARNING: unknown item ID '{iid}' ({item_map[iid]['title']}) — add to ITEM_ORDER")

        # Build ordered items list; fill in any missing items as not done
        items = []
        for iid in ITEM_ORDER:
            if iid in item_map:
                items.append(item_map[iid])
            else:
                items.append({'id': iid, 'title': ITEM_TITLES.get(iid, iid), 'done': False, 'date': ''})

        done_count = sum(1 for it in items if it['done'])
        total      = len(items)
        pct        = round(done_count / total * 100) if total else 0

        # Complete is derived from actual module progress, not the LMS's own
        # "Curriculum Complete" flag — that flag goes stale when a module is added
        # to the curriculum after someone was already marked complete under the
        # old requirement (e.g. stuck at 11/12 but still flagged "Yes" from when
        # the curriculum only had 11 modules).
        complete = 'Yes' if total and done_count >= total else 'No'
        if complete == 'Yes':
            days_rem = None

        # CompleteDate: max done item date when Complete=Yes
        complete_date = ''
        complete_qtr  = ''
        if complete == 'Yes':
            done_dates = [it['date'] for it in items if it['done'] and it['date']]
            if done_dates:
                complete_date = max(done_dates)
                complete_qtr  = km_fiscal_quarter(complete_date)

        person = {
            'FirstName':     first,
            'LastName':      last,
            'Email':         email,
            'JobTitle':      title,
            'Region':        region,
            'Market':        market,
            'Manager':       f"{mgr_first} {mgr_last}".strip(),
            'MgrEmail':      mgr_email,
            'MgrTitle':      mgr_title,
            'HireDate':      hire_date or '',
            'AssignDate':    assign_dt or '',
            'DaysRemaining': days_rem,
            'Complete':      complete,
            'CompleteDate':  complete_date,
            'CompleteQtr':   complete_qtr,
            'ls': {
                'done':  done_count,
                'total': total,
                'pct':   pct,
                'items': items
            },
            'overallDone': done_count,
            'overallPct':  pct,
        }
        people.append(person)

    people.sort(key=lambda p: (p['LastName'], p['FirstName']))
    return people, date_label


def generate_html(people, date_label, sales_cert, sales_deals):
    people_json      = json.dumps(people, separators=(',', ':'))
    sales_cert_json  = json.dumps(sales_cert, separators=(',', ':'))
    sales_deals_json = json.dumps(sales_deals, separators=(',', ':'))

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Layered Security Certification - Direct Sales</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<style>
  :root {{
    --bg:#0f1117; --surface:#1a1d27; --surface2:#22263a; --border:#2e3350;
    --accent:#4f8ef7; --accent2:#7c5cfc; --accent3:#f7c94f;
    --text:#e8ecf4; --muted:#7b82a0; --green:#3ecf8e; --red:#f76f6f;
    --teal:#2dd4bf; --green-subtle:#3ecf8e22; --red-subtle:#f76f6f22;
    --font:'Segoe UI',system-ui,sans-serif;
  }}
  body.light-mode {{
    --bg:#f4f6fb; --surface:#ffffff; --surface2:#eef1f7; --border:#d0d7e8;
    --accent:#2563eb; --accent2:#6d28d9; --accent3:#d97706;
    --text:#1a1d27; --muted:#475569; --green:#059669; --red:#dc2626;
    --teal:#0f766e; --green-subtle:#05966922; --red-subtle:#dc262622;
  }}
  body.light-mode select,body.light-mode input[type=date]{{color-scheme:light;}}
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{background:var(--bg);color:var(--text);font-family:var(--font);min-height:100vh;transition:background .2s,color .2s;}}

  /* ── Header ── */
  .header{{padding:20px 28px 16px;border-bottom:1px solid var(--border);display:grid;grid-template-columns:1fr auto 1fr;align-items:center;gap:12px;}}
  .header-center{{display:flex;justify-content:center;align-items:center;}}
  .header-right{{display:flex;justify-content:flex-end;align-items:center;gap:8px;flex-wrap:wrap;}}
  .header h1{{font-size:18px;font-weight:700;letter-spacing:.3px;cursor:pointer;user-select:none;}}
  .header h1 span{{color:var(--muted);font-weight:400;}}
  .header-date{{font-size:11px;color:var(--muted);margin-top:2px;}}
  .btn-theme{{background:transparent;border:1px solid var(--border);color:var(--muted);border-radius:6px;padding:5px 12px;font-size:12px;cursor:pointer;transition:all .15s;}}
  .btn-theme:hover{{border-color:var(--accent);color:var(--text);}}
  .kma-logo{{height:34px;width:auto;display:block;}}
  .kma-logo-light{{display:none;}}
  body.light-mode .kma-logo-dark{{display:none;}}
  body.light-mode .kma-logo-light{{display:block;}}

  /* ── Export dropdown ── */
  .btn-export{{background:var(--accent);border:1px solid var(--accent);color:#fff;border-radius:6px;padding:5px 12px;font-size:12px;cursor:pointer;transition:all .15s;font-weight:600;}}
  .btn-export:hover{{opacity:0.88;}}
  .export-drop{{position:relative;}}
  .export-menu{{position:absolute;top:calc(100% + 6px);right:0;background:var(--surface);border:1px solid var(--border);border-radius:8px;min-width:210px;box-shadow:0 4px 24px rgba(0,0,0,.28);display:none;z-index:200;overflow:visible;}}
  .export-menu.open{{display:block;}}
  .export-item{{display:block;width:100%;text-align:left;padding:10px 14px;font-size:13px;color:var(--text);background:transparent;border:none;cursor:pointer;transition:background .1s;font-family:inherit;}}
  .export-item:hover{{background:var(--surface2);}}
  .export-parent{{position:relative;display:flex;justify-content:space-between;align-items:center;padding:10px 14px;font-size:13px;color:var(--text);cursor:default;transition:background .1s;}}
  .export-parent:hover{{background:var(--surface2);}}
  .export-chevron{{font-size:11px;color:var(--muted);margin-left:10px;}}
  .export-submenu{{position:absolute;right:100%;top:0;background:var(--surface);border:1px solid var(--border);border-radius:8px;min-width:90px;box-shadow:0 4px 24px rgba(0,0,0,0.28);display:none;z-index:201;overflow:hidden;margin-right:4px;}}
  .export-parent:hover .export-submenu{{display:block;}}

  /* ── Info button / popover ── */
  .info-btn{{display:inline-flex;align-items:center;justify-content:center;width:15px;height:15px;border-radius:50%;background:var(--surface2);border:1px solid var(--border);color:var(--muted);font-size:9px;font-weight:700;cursor:pointer;margin-left:5px;vertical-align:middle;flex-shrink:0;line-height:1;transition:border-color .15s,color .15s;}}
  .info-btn:hover{{border-color:var(--accent);color:var(--accent);}}
  .info-popover{{position:fixed;z-index:9999;background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:12px 14px;font-size:12px;color:var(--text);line-height:1.6;max-width:260px;box-shadow:0 4px 24px rgba(0,0,0,0.5);display:none;}}
  .info-popover.visible{{display:block;}}

  /* ── Curriculum note ── */
  .curriculum-note{{padding:9px 28px;background:var(--surface2);border-bottom:1px solid var(--border);font-size:11.5px;color:var(--muted);line-height:1.5;}}
  .curriculum-note b{{color:var(--text);}}

  /* ── Filters ── */
  .filters{{padding:14px 28px;display:flex;gap:10px;flex-wrap:wrap;align-items:center;border-bottom:1px solid var(--border);background:var(--surface);}}
  .filter-label{{font-size:12px;color:var(--muted);text-transform:uppercase;letter-spacing:.6px;margin-right:4px;}}
  select,input[type=date],input[type=text]{{background:var(--surface2);border:1px solid var(--border);color:var(--text);border-radius:6px;padding:6px 10px;font-size:13px;cursor:pointer;outline:none;color-scheme:dark;}}
  select:focus,input:focus{{border-color:var(--accent);}}
  .btn-reset{{background:transparent;border:1px solid var(--border);color:var(--muted);border-radius:6px;padding:6px 14px;font-size:12px;cursor:pointer;transition:border-color .15s,color .15s;}}
  .btn-reset:hover{{border-color:var(--accent);color:var(--text);}}
  .result-count{{margin-left:auto;font-size:12px;color:var(--muted);}}

  /* ── Stat cards ── */
  .stats{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;padding:20px 28px;}}
  .stat{{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:16px 18px;}}
  .stat-label{{font-size:11px;text-transform:uppercase;letter-spacing:.7px;color:var(--muted);margin-bottom:6px;}}
  .stat-value{{font-size:28px;font-weight:700;line-height:1;}}
  .stat-value.green{{color:var(--green);}}
  .stat-value.red{{color:var(--red);}}
  .stat-value.teal{{color:var(--teal);}}
  .stat-value.blue{{color:var(--accent);}}
  .stat-value.amber{{color:var(--accent3);}}
  .stat-sub{{font-size:11px;color:var(--muted);margin-top:4px;}}

  /* ── Charts ── */
  .charts{{display:grid;grid-template-columns:1fr 1fr;gap:16px;padding:0 28px 16px;}}
  @media(max-width:700px){{.charts{{grid-template-columns:1fr;}}}}
  @media(max-width:480px){{.chart-wrap{{height:180px;}}}}
  .chart-card{{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:18px;}}
  .chart-title{{font-size:13px;font-weight:600;margin-bottom:14px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;}}
  body.light-mode .chart-title{{color:var(--text);}}
  .chart-wrap{{position:relative;height:260px;}}

  /* ── Roster ── */
  .section{{padding:0 28px 32px;}}
  .section-header{{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;flex-wrap:wrap;gap:8px;}}
  .section-title{{font-size:13px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;}}
  body.light-mode .section-title{{color:var(--text);}}
  .section-hint{{font-size:11px;color:var(--muted);margin-top:3px;}}
  .roster-search{{width:200px;}}
  .roster-wrap{{display:flex;border:1px solid var(--border);border-radius:10px;overflow:hidden;}}
  .roster-left{{flex:1.6;min-width:0;overflow:auto;max-height:820px;border-right:1px solid var(--border);}}
  .roster-right{{flex:1;overflow-y:auto;max-height:820px;padding:16px 20px;}}
  .roster-right-header{{margin-bottom:14px;padding-bottom:12px;border-bottom:1px solid var(--border);}}
  .no-data{{text-align:center;color:var(--muted);padding:40px;font-size:13px;}}
  @media(max-width:900px){{.roster-wrap{{flex-direction:column;}}.roster-left{{border-right:none;border-bottom:1px solid var(--border);max-height:420px;}}}}

  .pill{{font-size:8px;font-weight:700;border-radius:10px;padding:2px 7px;white-space:nowrap;letter-spacing:.03em;}}
  .pill.green{{color:var(--green);background:var(--green-subtle);border:1px solid var(--green)44;}}
  .pill.yellow{{color:#b45309;background:#fef3c7;border:1px solid #f59e0b44;}}
  .pill.red{{color:var(--red);background:var(--red-subtle);border:1px solid var(--red)44;}}
  .pill.gold{{color:#b45309;background:#fef3c722;border:1px solid var(--accent3)66;}}

  /* Progress Report table */
  .progress-table{{width:100%;border-collapse:collapse;font-size:13px;}}
  .progress-table thead th{{
    text-align:left;font-size:10.5px;text-transform:uppercase;letter-spacing:.05em;
    color:var(--muted);font-weight:600;padding:8px 12px;border-bottom:1px solid var(--border);
    background:var(--surface2);white-space:nowrap;
  }}
  .progress-table tbody td{{padding:9px 12px;border-bottom:1px solid var(--border);vertical-align:middle;}}
  .progress-table tbody tr.progress-row{{cursor:pointer;transition:background .1s;}}
  .progress-table tbody tr.progress-row:hover{{background:var(--surface2);}}
  .progress-table tbody tr.progress-row.active{{background:var(--accent)11;}}
  .progress-table .p-name{{font-weight:600;}}
  .progress-table .p-sub{{display:block;font-size:11px;color:var(--muted);margin-top:1px;font-weight:400;}}
  .mgr-group-row td{{background:var(--surface2);font-size:11.5px;font-weight:600;color:var(--text);padding:8px 12px;border-bottom:1px solid var(--border);}}
  .mgr-group-row .mgr-sub{{font-weight:400;color:var(--muted);margin-left:8px;font-size:11px;}}
  .deals-cell{{font-family:inherit;color:var(--accent2);font-weight:600;}}
  .deals-cell.empty{{color:var(--muted);font-weight:400;}}
  .roster-pct{{font-size:12px;font-weight:700;color:var(--muted);flex-shrink:0;}}
  .roster-pct.pct-progress{{color:#f59e0b;}}
  .roster-pct.pct-done{{color:var(--green);}}

  /* Roster key */
  .roster-key{{display:flex;gap:16px;align-items:center;padding:6px 12px;background:var(--surface2);border-radius:8px;margin-bottom:10px;font-size:11px;color:var(--muted);}}
  .key-stripe{{display:inline-block;width:10px;height:10px;border-radius:2px;vertical-align:middle;margin-right:5px;}}
  .key-stripe.green{{background:var(--green);}}
  .key-stripe.yellow{{background:#f59e0b;}}
  .key-stripe.red{{background:var(--red);}}

  /* Sort buttons */
  .sort-btn{{background:transparent;border:1px solid var(--border);color:var(--muted);border-radius:6px;padding:4px 10px;font-size:11px;cursor:pointer;transition:all .15s;white-space:nowrap;}}
  .sort-btn:hover{{border-color:var(--accent);color:var(--text);}}
  .sort-btn.active{{border-color:var(--accent);color:var(--accent);background:var(--accent)11;}}

  /* Detail panel */
  .detail-grid{{display:grid;grid-template-columns:1fr 1fr;gap:12px 28px;margin-top:14px;}}
  .detail-label{{font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:3px;}}
  .detail-value{{font-size:14px;font-weight:500;}}
  .badge-status{{display:inline-block;padding:3px 12px;border-radius:20px;font-size:12px;font-weight:700;}}
  .badge-status.complete{{background:var(--green-subtle);color:var(--green);}}
  .badge-status.in-progress{{background:#fef3c7;color:#b45309;border:1px solid #f59e0b44;}}
  .badge-status.not-complete{{background:var(--red-subtle);color:var(--red);border:1px solid var(--red)44;}}
  .roster-bottom .badge-status{{font-size:10px;padding:2px 8px;}}

  /* Progress bars */
  .prog-wrap{{background:var(--surface2);border-radius:4px;height:6px;width:100%;overflow:hidden;margin-top:4px;}}
  .prog-bar{{height:6px;border-radius:4px;transition:width .3s;}}
  .prog-bar.green{{background:var(--green);}}
  .prog-bar.blue{{background:var(--accent);}}

  /* Curriculum course list */
  .curriculum-section{{margin-top:18px;}}
  .curriculum-header{{display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;cursor:pointer;user-select:none;}}
  .curriculum-header:hover .curriculum-title{{color:var(--accent);}}
  .curriculum-title{{font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:var(--muted);transition:color .15s;}}
  body.light-mode .curriculum-title{{color:var(--text);}}
  .curriculum-count{{font-size:11px;color:var(--muted);}}
  .course-list{{margin-top:8px;display:none;}}
  .course-list.open{{display:block;}}
  .course-item{{display:flex;align-items:flex-start;gap:8px;padding:5px 0;border-bottom:1px solid var(--border);}}
  .course-item:last-child{{border-bottom:none;}}
  .course-icon{{flex-shrink:0;font-size:11px;margin-top:1px;}}
  .course-icon.done{{color:var(--green);}}
  .course-icon.todo{{color:var(--muted);}}
  .course-title{{flex:1;font-size:12px;line-height:1.4;}}
  .course-date{{font-size:10px;color:var(--muted);white-space:nowrap;}}

  /* Print */
  @page{{size:landscape;margin:.65in;}}
  @media print{{
    body{{background:#fff!important;color:#111!important;}}
    .header,.filters,.stats,.charts,.section,.print-hide{{display:none!important;}}
    #print-header{{display:block!important;}}
    #print-stats{{display:flex!important;gap:40px;flex-wrap:wrap;margin-bottom:18px;padding-bottom:16px;border-bottom:2px solid #dde4f0;}}
    body.print-no-summary #print-stats{{display:none!important;}}
    #print-roster-wrap{{display:block!important;}}
    .ptable{{width:100%;border-collapse:collapse;font-size:11px;}}
    .ptable th{{background:#f0f4ff;color:#111;font-weight:700;padding:5px 8px;border:1px solid #ccc;text-align:left;}}
    .ptable td{{padding:5px 8px;border:1px solid #ddd;vertical-align:middle;}}
    .ptable tr:nth-child(even) td{{background:#fafafa;}}
  }}
</style>
</head>
<body>

<!-- ── Header ── -->
<div class="header">
  <div>
    <h1 id="dash-title" title="Triple-click to return to the Analytics Hub">Layered Security <span>Certification &mdash; Direct Sales</span></h1>
    <div class="header-date">Data through {date_label}</div>
  </div>
  <div class="header-center">
    <img src="https://jasonackerman1.github.io/playbook-dashboard/KMA-wht.svg" class="kma-logo kma-logo-dark print-hide" alt="KM Academy">
    <img src="https://jasonackerman1.github.io/playbook-dashboard/KMA-drk.svg" class="kma-logo kma-logo-light print-hide" alt="KM Academy">
  </div>
  <div class="header-right print-hide">
    <div class="export-drop" id="export-drop">
      <button class="btn-export" onclick="toggleExportDrop()">&#128438; Export &#9660;</button>
      <div class="export-menu" id="export-menu">
        <div class="export-parent">Full Report<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('full')">PDF</button><button class="export-item" onclick="runExportXLSX('full')">Excel</button></div></div>
        <div class="export-parent">Not Complete<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('not-complete')">PDF</button><button class="export-item" onclick="runExportXLSX('not-complete')">Excel</button></div></div>
        <div class="export-parent">Manager Summary<span class="export-chevron">&#8249;</span><div class="export-submenu"><button class="export-item" onclick="runExport('manager-summary')">PDF</button><button class="export-item" onclick="runExportXLSX('manager-summary')">Excel</button></div></div>
      </div>
    </div><span class="info-btn" onclick="showInfo(event,'export')">?</span>
    <button class="btn-theme" id="btn-theme" onclick="toggleTheme()">&#9728; Light</button>
  </div>
</div>
<div class="curriculum-note print-hide">Earning the Certified Layered Security Specialist designation requires all four criteria: completing the full training curriculum shown below, engaging your SSE on a co-development opportunity tagged <b>"OPS FY26 Layered Security Certification"</b> in Salesforce, closing a real deal of <b>$5,000 or more</b>, and participating in a case study. This dashboard tracks curriculum completion and Closed Won sales progress toward those requirements &mdash; certification itself is confirmed separately once all four are met.</div>

<!-- ── Filters ── -->
<div class="filters">
  <span class="filter-label">Status</span>
  <select id="f-status" onchange="applyFilters()">
    <option value="">All</option>
    <option value="Complete">Complete</option>
    <option value="In Progress">In Progress</option>
    <option value="Not Started">Not Started</option>
  </select>
  <span class="filter-label">Market</span>
  <select id="f-market" onchange="applyFilters()"><option value="">All Markets</option></select>
  <button class="btn-reset" onclick="resetFilters()">Reset</button>
  <span class="result-count" id="result-count"></span>
</div>

<!-- ── Stat cards ── -->
<div class="stats">
  <div class="stat">
    <div class="stat-label">Total Enrolled <span class="info-btn" onclick="showInfo(event,'total-enrolled')">?</span></div>
    <div class="stat-value" id="s-total">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">Curriculum Complete <span class="info-btn" onclick="showInfo(event,'complete')">?</span></div>
    <div class="stat-value green" id="s-complete">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">In Progress <span class="info-btn" onclick="showInfo(event,'in-progress')">?</span></div>
    <div class="stat-value blue" id="s-inprog">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">Not Started <span class="info-btn" onclick="showInfo(event,'not-started')">?</span></div>
    <div class="stat-value red" id="s-notstarted">&#8212;</div>
  </div>
  <div class="stat">
    <div class="stat-label">Completion Rate <span class="info-btn" onclick="showInfo(event,'completion-rate')">?</span></div>
    <div class="stat-value teal" id="s-rate">&#8212;</div>
    <div class="stat-sub" id="s-rate-sub"></div>
  </div>
  <div class="stat" id="stat-sales-cert">
    <div class="stat-label">Certified <span class="info-btn" onclick="showInfo(event,'sales-certified')">?</span></div>
    <div class="stat-value" style="color:var(--accent2)" id="s-salescert">&#8212;</div>
    <div class="stat-sub" id="s-salescert-sub"></div>
  </div>
</div>

<!-- ── Charts ── -->
<div class="charts">
  <div class="chart-card">
    <div class="chart-title">Completion Pipeline <span class="info-btn" onclick="showInfo(event,'pipeline-chart')">?</span></div>
    <div class="chart-wrap"><canvas id="pipelineChart"></canvas></div>
  </div>
  <div class="chart-card">
    <div class="chart-title">Learners by Market <span class="info-btn" onclick="showInfo(event,'market-chart')">?</span></div>
    <div class="chart-wrap"><canvas id="marketChart"></canvas></div>
  </div>
</div>

<!-- ── Roster ── -->
<div class="section">
  <div class="section-header">
    <div>
      <div class="section-title">Progress Report <span class="info-btn" onclick="showInfo(event,'roster')">?</span></div>
      <div class="section-hint">Click any row to see full detail &mdash; course checklist &amp; closed won activity</div>
    </div>
    <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;">
      <span style="font-size:11px;color:var(--muted);margin-right:2px;">View:</span>
      <button class="sort-btn active" id="view-individual" onclick="setRosterView('individual')">Individual</button>
      <button class="sort-btn" id="view-manager" onclick="setRosterView('manager')">By Manager</button>
      <span id="sort-controls" style="display:flex;align-items:center;gap:6px;margin-left:6px;">
        <span style="font-size:11px;color:var(--muted);margin-right:2px;">Sort:</span>
        <button class="sort-btn" data-sort="name" onclick="setSort('name')">Name</button>
        <button class="sort-btn" data-sort="status" onclick="setSort('status')">Status</button>
        <button class="sort-btn active" data-sort="pct" onclick="setSort('pct')">Completion % &#9660;</button>
      </span>
    </div>
    <input type="text" id="f-search" class="roster-search" placeholder="Search by name&hellip;" oninput="applyFilters()" style="width:170px;">
  </div>
  <div class="roster-key">
    <span><span class="key-stripe green"></span>Complete</span>
    <span><span class="key-stripe yellow"></span>In Progress</span>
    <span><span class="key-stripe red"></span>Not Started</span>
  </div>
  <div class="roster-wrap">
    <div class="roster-left" id="roster-left"></div>
    <div class="roster-right" id="roster-right">
      <div class="no-data">Select a person to view details</div>
    </div>
  </div>
</div>

<!-- ── Info popover ── -->
<div class="info-popover" id="info-popover"></div>

<!-- ── Print-only elements ── -->
<div id="print-header" style="display:none;margin-bottom:20px;">
  <div style="font-size:20px;font-weight:700;margin-bottom:4px;" id="ph-title"></div>
  <div style="font-size:12px;color:#555;margin-bottom:2px;" id="ph-date"></div>
  <div style="font-size:12px;color:#555;" id="ph-filters"></div>
  <div id="ph-desc" style="display:none;margin-top:10px;padding-top:10px;border-top:1px solid #dde4f0;font-size:12px;color:#444;font-style:italic;"></div>
</div>
<div id="print-stats" style="display:none;gap:40px;flex-wrap:wrap;margin-bottom:18px;padding-bottom:16px;border-bottom:2px solid #dde4f0;"></div>
<div id="print-roster-wrap" style="display:none;">
  <table class="ptable" id="print-roster-table">
    <thead id="print-roster-head"></thead>
    <tbody id="print-roster-body"></tbody>
  </table>
</div>

<script>
const PEOPLE      = {people_json};
const SALES_CERT  = {sales_cert_json};
const SALES_DEALS = {sales_deals_json};
</script>
<script>
let filtered = [];
let sortField = "pct";
let sortDir   = "desc";
let selectedEmail = null;
let rosterView = "individual";
let pipelineChart, marketChart;

// ── Sales Certification (baked in from the Certification Report at build time) ──
function isSalesCertified(email){{
  var rec = SALES_CERT[(email || "").toLowerCase()];
  return !!(rec && rec.certified);
}}
function salesCertDate(email){{
  var rec = SALES_CERT[(email || "").toLowerCase()];
  return rec && rec.date ? rec.date : "";
}}

// ── Closed Won deals (baked in from the OPS FY26 opportunity export at build time) ──
function closedWonCount(email){{
  var rec = SALES_DEALS[(email || "").toLowerCase()];
  return rec && typeof rec.closedWon === "number" ? rec.closedWon : 0;
}}
function closedWonAmount(email){{
  var rec = SALES_DEALS[(email || "").toLowerCase()];
  return rec && typeof rec.amount === "number" ? rec.amount : 0;
}}
function closedWonAccount(email){{
  var rec = SALES_DEALS[(email || "").toLowerCase()];
  return rec && rec.accountName ? rec.accountName : "";
}}
function closedWonCloseDate(email){{
  var rec = SALES_DEALS[(email || "").toLowerCase()];
  return rec && rec.closeDate ? rec.closeDate : "";
}}
function fmtMoney(n){{
  return "$" + (n || 0).toLocaleString(undefined, {{maximumFractionDigits: 0}});
}}

function sel(id) {{ return document.getElementById(id); }}
function cv(v)   {{ return getComputedStyle(document.body).getPropertyValue(v).trim(); }}

// ── Theme ──
(function(){{
  if(localStorage.getItem("pb-theme") !== "dark") document.body.classList.add("light-mode");
  sel("btn-theme").textContent = document.body.classList.contains("light-mode") ? "🌙 Dark" : "☀ Light";
}})();
function toggleTheme(){{
  var light = document.body.classList.toggle("light-mode");
  localStorage.setItem("pb-theme", light ? "light" : "dark");
  sel("btn-theme").textContent = light ? "🌙 Dark" : "☀ Light";
  applyFilters();
}}

// ── Triple-click on title → Analytics Hub ──
(function(){{
  var n = 0, t;
  sel("dash-title").addEventListener("click", function(){{
    n++; clearTimeout(t);
    t = setTimeout(function(){{ n = 0; }}, 500);
    if(n >= 3){{ n = 0; window.location.href = "index.html"; }}
  }});
}})();

document.addEventListener("click", function(e){{
  if(!e.target.classList.contains("info-btn")) sel("info-popover").classList.remove("visible");
}});

// ── Export dropdown ──
function toggleExportDrop(){{
  sel("export-menu").classList.toggle("open");
}}
document.addEventListener("click", function(e){{
  var d = sel("export-drop");
  if(d && !d.contains(e.target)) sel("export-menu").classList.remove("open");
}});

// ── Info tooltip ──
var INFO_MSGS = {{
  "total-enrolled":  "The total number of people currently assigned to the Layered Security curriculum (Direct Sales). This stays fixed regardless of the filters above — use the ‘shown’ count next to Reset to see how many match your current filters.",
  "complete":        "People who have completed all 12 required modules in the Layered Security curriculum. This reflects curriculum completion only — actual certification also requires $5,000 in qualifying sales, tracked in an external system not shown here.",
  "in-progress":     "People who have started the curriculum and completed at least one module, but haven't finished everything yet.",
  "not-started":     "People who are assigned to the curriculum but haven't completed any modules yet.",
  "past-due":        "People who have not finished the curriculum and have passed their required completion date (negative days remaining in the LMS export).",
  "sales-certified": "People confirmed as certified in the Sales Certification report — curriculum completion plus the $5,000 sale, SSE engagement, and case study confirmed externally. Anyone not listed there is treated as not yet certified.",
  "completion-rate": "The percentage of assigned people who have finished all 12 curriculum modules so far. Updates when you apply filters.",
  "pipeline-chart":  "A quick snapshot of where everyone stands: how many haven't started yet, how many are actively working through the modules, and how many have finished all 12.",
  "market-chart":    "Curriculum progress broken down by sales market. Each bar shows how many people in that market are Complete, In Progress, or Not Started. Hover for exact counts. Updates when you apply filters.",
  "roster":          "The full list of people in the curriculum. Each row shows their manager, Layered Security status, overall completion %, and Closed Won total. Click any row to see a full course-by-course breakdown, plus certification and deal detail, in the panel on the right.",
  "export":          "Download a report based on whoever is currently shown. Apply filters first to scope the report. Full Report includes everyone with all module progress columns. Not Complete is a contact list for follow-up, sorted by manager. Manager Summary shows each manager's team size and completion counts."
}};
function showInfo(e, key){{
  var pop = sel("info-popover");
  pop.textContent = INFO_MSGS[key] || "";
  pop.classList.add("visible");
  var r = e.target.getBoundingClientRect();
  pop.style.top  = (r.bottom + 6) + "px";
  pop.style.left = Math.min(r.left, window.innerWidth - 280) + "px";
  e.stopPropagation();
}}

// ── JS helpers ──
function escHtml(s){{ return String(s).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;").replace(/"/g,"&quot;"); }}
function personStatus(p){{
  // Derive completion from actual module progress rather than the LMS's
  // Complete flag: that flag can go stale when a module is added to the
  // curriculum after someone was already marked complete under the old
  // requirement, leaving them flagged "Yes" despite not finishing all
  // current modules (e.g. 11/12 instead of 12/12).
  if(p.ls && typeof p.ls.done === "number" && typeof p.ls.total === "number" && p.ls.total > 0){{
    if(p.ls.done >= p.ls.total) return "Complete";
    if(p.ls.done > 0) return "In Progress";
    return "Not Started";
  }}
  if(p.overallPct >= 100) return "Complete";
  if(p.overallPct > 0)     return "In Progress";
  return "Not Started";
}}
function isPastDue(p){{
  return personStatus(p) !== "Complete" && typeof p.DaysRemaining === "number" && p.DaysRemaining < 0;
}}
function fmtDate(d){{
  if(!d) return "-";
  var pts = d.split("-"), months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"];
  return months[parseInt(pts[1]) - 1] + " " + parseInt(pts[2]) + ", " + pts[0];
}}
// ── Populate market dropdown ──
(function(){{
  var markets = [...new Set(PEOPLE.map(function(p){{ return p.Market; }}).filter(Boolean))].sort();
  markets.forEach(function(m){{
    sel("f-market").innerHTML += '<option value="' + m + '">' + m + "</option>";
  }});
}})();

function resetFilters(){{
  sel("f-market").value = "";
  sel("f-status").value = "";
  sel("f-search").value = "";
  applyFilters();
}}

// ── Roster view toggle ──
function setRosterView(v){{
  rosterView = v;
  sel("view-individual").classList.toggle("active", v === "individual");
  sel("view-manager").classList.toggle("active", v === "manager");
  sel("sort-controls").style.display = v === "individual" ? "flex" : "none";
  selectedEmail = null;
  sel("roster-right").innerHTML = '<div class="no-data">Select a person to view details</div>';
  renderRoster();
}}

// ── Sort ──
function setSort(field){{
  if(sortField === field){{
    sortDir = sortDir === "desc" ? "asc" : "desc";
  }} else {{
    sortField = field;
    sortDir   = "desc";
  }}
  document.querySelectorAll(".sort-btn").forEach(function(btn){{
    var active = btn.dataset.sort === sortField;
    btn.classList.toggle("active", active);
    var arrow = active ? (sortDir === "desc" ? " ↓" : " ↑") : "";
    btn.textContent = btn.textContent.replace(/ [↑↓]$/, "") + arrow;
  }});
  renderRoster();
}}

// ── applyFilters ──
function applyFilters(){{
  var market = sel("f-market").value;
  var status = sel("f-status").value;
  var q      = (sel("f-search").value || "").toLowerCase();
  filtered = PEOPLE.filter(function(p){{
    if(market && p.Market !== market) return false;
    if(status && personStatus(p) !== status) return false;
    if(q && !(p.FirstName + " " + p.LastName).toLowerCase().includes(q)) return false;
    return true;
  }});
  sel("result-count").textContent = filtered.length + " shown";
  renderStats();
  renderCharts();
  renderRoster();
}}

// ── renderStats ──
function renderStats(){{
  var total   = filtered.length;
  var cert    = filtered.filter(function(p){{ return personStatus(p) === "Complete"; }}).length;
  var inprog  = filtered.filter(function(p){{ return personStatus(p) === "In Progress"; }}).length;
  var nostart = filtered.filter(function(p){{ return personStatus(p) === "Not Started"; }}).length;
  var rate    = total > 0 ? Math.round(cert / total * 100) : 0;
  sel("s-total").textContent      = PEOPLE.length;
  sel("s-complete").textContent   = cert;
  sel("s-inprog").textContent     = inprog;
  sel("s-notstarted").textContent = nostart;
  sel("s-rate").textContent       = rate + "%";
  sel("s-rate-sub").textContent   = total > 0 ? (cert + " of " + total + " enrolled") : "";
  var salesCert = filtered.filter(function(p){{ return isSalesCertified(p.Email); }}).length;
  sel("s-salescert").textContent     = salesCert;
  sel("s-salescert-sub").textContent = total > 0 ? (salesCert + " of " + total + " confirmed") : "";
}}

// ── renderCharts ──
function renderCharts(){{
  var labelColor = document.body.classList.contains("light-mode") ? cv("--text") : cv("--muted");
  var gridColor  = cv("--border");

  var pNotStarted = filtered.filter(function(p){{ return personStatus(p) === "Not Started"; }}).length;
  var pInProgress = filtered.filter(function(p){{ return personStatus(p) === "In Progress"; }}).length;
  var pComplete   = filtered.filter(function(p){{ return personStatus(p) === "Complete";    }}).length;

  if(pipelineChart) pipelineChart.destroy();
  pipelineChart = new Chart(sel("pipelineChart"), {{
    type: "bar",
    data: {{
      labels: ["Not Started", "In Progress", "Complete"],
      datasets: [{{
        data: [pNotStarted, pInProgress, pComplete],
        backgroundColor: [cv("--red") + "cc", "#f59e0b" + "cc", cv("--green") + "cc"],
        borderRadius: 4,
        borderSkipped: false
      }}]
    }},
    options: {{
      indexAxis: "y",
      responsive: true, maintainAspectRatio: false,
      plugins: {{
        legend: {{ display: false }},
        tooltip: {{ callbacks: {{ label: function(ctx){{ return " " + ctx.raw + " people"; }} }} }}
      }},
      scales: {{
        x: {{ grid: {{ color: gridColor }}, ticks: {{ color: labelColor, font: {{ size: 11 }}, stepSize: 1 }} }},
        y: {{ grid: {{ display: false }}, ticks: {{ color: labelColor, font: {{ size: 12 }} }} }}
      }}
    }}
  }});

  var mktMap = {{}};
  filtered.forEach(function(p){{
    var m = p.Market || "Unknown";
    var s = personStatus(p);
    if(!mktMap[m]) mktMap[m] = {{ Complete: 0, "In Progress": 0, "Not Started": 0 }};
    mktMap[m][s] = (mktMap[m][s] || 0) + 1;
  }});
  var mktLabels = Object.keys(mktMap).sort(function(a,b){{
    var ta = mktMap[a].Complete + mktMap[a]["In Progress"] + mktMap[a]["Not Started"];
    var tb = mktMap[b].Complete + mktMap[b]["In Progress"] + mktMap[b]["Not Started"];
    return tb - ta;
  }});

  if(marketChart) marketChart.destroy();
  marketChart = new Chart(sel("marketChart"), {{
    type: "bar",
    data: {{
      labels: mktLabels,
      datasets: [
        {{ label: "Complete",    data: mktLabels.map(function(m){{ return mktMap[m].Complete;       }}), backgroundColor: cv("--green") + "cc", borderWidth: 0, borderRadius: 2 }},
        {{ label: "In Progress", data: mktLabels.map(function(m){{ return mktMap[m]["In Progress"]; }}), backgroundColor: "#f59e0b"       + "cc", borderWidth: 0, borderRadius: 2 }},
        {{ label: "Not Started", data: mktLabels.map(function(m){{ return mktMap[m]["Not Started"]; }}), backgroundColor: cv("--red")     + "cc", borderWidth: 0, borderRadius: 2 }}
      ]
    }},
    options: {{
      indexAxis: "y",
      responsive: true, maintainAspectRatio: false,
      plugins: {{
        legend: {{ display: true, position: "bottom", labels: {{ color: labelColor, font: {{ size: 11 }}, padding: 16, boxWidth: 12 }} }},
        tooltip: {{
          mode: "index",
          callbacks: {{
            title: function(items){{ return items[0].label; }},
            afterBody: function(items){{
              var total = items.reduce(function(s, i){{ return s + i.raw; }}, 0);
              return ["─────────", "Enrolled: " + total];
            }}
          }}
        }}
      }},
      scales: {{
        x: {{ stacked: true, grid: {{ color: gridColor }}, ticks: {{ color: labelColor, font: {{ size: 11 }}, stepSize: 1 }} }},
        y: {{ stacked: true, grid: {{ display: false }}, ticks: {{ color: labelColor, font: {{ size: 11 }} }} }}
      }}
    }}
  }});
}}

// ── buildPersonRow ──
function buildPersonRow(p){{
  var fullName  = p.FirstName + " " + p.LastName;
  var status    = personStatus(p);
  var pctClass  = status === "Complete" ? " pct-done" : status === "In Progress" ? " pct-progress" : "";
  var badgeCls  = status === "Complete" ? "complete" : status === "In Progress" ? "in-progress" : "not-complete";
  var badgeTxt  = status === "Complete" ? "&#10003; Complete" : status === "In Progress" ? "In Progress" : "Not Started";
  var overdueTag = isPastDue(p) ? '<span class="pill red" style="margin-left:6px;">PAST DUE</span>' : "";
  var certTag   = isSalesCertified(p.Email) ? '<span class="pill gold" style="margin-top:4px;display:inline-block;">Certified Layered Security Specialist</span>' : "";
  var amt       = closedWonAmount(p.Email);
  var dealsHtml = amt > 0
    ? '<span class="deals-cell">' + fmtMoney(amt) + "</span>" + (closedWonCount(p.Email) > 1 ? ' <span style="font-size:10px;color:var(--muted)">(' + closedWonCount(p.Email) + ")</span>" : "")
    : '<span class="deals-cell empty">&mdash;</span>';
  var h = '<tr class="progress-row" data-email="' + escHtml(p.Email) + '" onclick="showDetail(this.dataset.email)">';
  h += '<td><span class="p-name">' + escHtml(fullName) + overdueTag + '</span><span class="p-sub">' + escHtml(p.JobTitle || "") + "</span>" + certTag + "</td>";
  h += "<td>" + escHtml(p.Manager || "-") + "</td>";
  h += '<td><span class="badge-status ' + badgeCls + '">' + badgeTxt + "</span></td>";
  h += '<td><span class="roster-pct' + pctClass + '">' + p.overallPct + "%</span></td>";
  h += "<td>" + dealsHtml + "</td>";
  h += "</tr>";
  return h;
}}

// ── renderRoster ──
function renderRoster(){{
  var html = '<table class="progress-table"><thead><tr>' +
    "<th>Learner</th><th>Manager</th><th>Layered Security</th><th>Overall %</th><th>Closed Won</th>" +
    "</tr></thead><tbody>";

  if(rosterView === "manager"){{
    var groups = {{}};
    filtered.forEach(function(p){{
      var mgr = p.Manager || "No Manager";
      if(!groups[mgr]) groups[mgr] = [];
      groups[mgr].push(p);
    }});
    var mgrsSorted = Object.keys(groups).sort();
    mgrsSorted.forEach(function(mgr){{
      var team = groups[mgr];
      var certCount = team.filter(function(p){{ return personStatus(p) === "Complete"; }}).length;
      html += '<tr class="mgr-group-row"><td colspan="5">' + escHtml(mgr) +
        '<span class="mgr-sub">' + team.length + " rep" + (team.length !== 1 ? "s" : "") + " &middot; " + certCount + " complete</span></td></tr>";
      team.slice()
        .sort(function(a, b){{ return (a.LastName + a.FirstName).localeCompare(b.LastName + b.FirstName); }})
        .forEach(function(p){{ html += buildPersonRow(p); }});
    }});
  }} else {{
    var d = sortDir === "desc" ? -1 : 1;
    var sorted = filtered.slice().sort(function(a, b){{
      if(sortField === "name"){{
        return d * (a.LastName + a.FirstName).localeCompare(b.LastName + b.FirstName);
      }} else if(sortField === "status"){{
        var order = {{ "Complete": 2, "In Progress": 1, "Not Started": 0 }};
        var diff = (order[personStatus(a)] || 0) - (order[personStatus(b)] || 0);
        if(diff !== 0) return d * diff;
        return (a.LastName + a.FirstName).localeCompare(b.LastName + b.FirstName);
      }} else {{
        var diff2 = a.overallPct - b.overallPct;
        if(diff2 !== 0) return d * diff2;
        return (a.LastName + a.FirstName).localeCompare(b.LastName + b.FirstName);
      }}
    }});
    sorted.forEach(function(p){{ html += buildPersonRow(p); }});
  }}

  html += "</tbody></table>";
  if(filtered.length === 0) html = '<div class="no-data">No people match the selected filters.</div>';
  sel("roster-left").innerHTML = html;
  if(selectedEmail){{
    var el = sel("roster-left").querySelector('[data-email="' + selectedEmail + '"]');
    if(el) el.classList.add("active");
    else   sel("roster-right").innerHTML = '<div class="no-data">Select a person to view details</div>';
  }}
}}

// ── showDetail ──
function showDetail(email){{
  selectedEmail = email;
  var p = PEOPLE.find(function(r){{ return r.Email === email; }});
  if(!p) return;
  document.querySelectorAll(".progress-row").forEach(function(el){{
    el.classList.toggle("active", el.dataset.email === email);
  }});
  var status     = personStatus(p);
  var badgeClass = status === "Complete" ? "complete" : status === "In Progress" ? "in-progress" : "not-complete";
  var badgeText  = status === "Complete" ? ("&#10003; Complete" + (p.CompleteDate ? " &middot; " + fmtDate(p.CompleteDate) : "")) : status === "In Progress" ? "In Progress" : "Not Started";

  function progBar(pct, cls){{
    return '<div class="prog-wrap"><div class="prog-bar ' + cls + '" style="width:' + pct + '%"></div></div>';
  }}
  function courseList(curriculum, id){{
    var items = curriculum.items;
    var html2 = '<div class="course-list open" id="cl-' + id + '">';
    items.forEach(function(item){{
      var iconClass = item.done ? "done" : "todo";
      var icon      = item.done ? "&#10003;" : "&#9675;";
      var dateStr   = item.done && item.date ? '<span class="course-date">' + fmtDate(item.date) + "</span>" : "";
      html2 += '<div class="course-item">';
      html2 += '<span class="course-icon ' + iconClass + '">' + icon + "</span>";
      html2 += '<span class="course-title">' + escHtml(item.title) + "</span>";
      html2 += dateStr;
      html2 += "</div>";
    }});
    html2 += "</div>";
    return html2;
  }}
  function curriculumSection(label, curriculum, id){{
    var doneOf = curriculum.done + " / " + curriculum.total + " modules";
    var pct    = curriculum.pct;
    var barCls = pct >= 100 ? "green" : "blue";
    var html3  = '<div class="curriculum-section">';
    html3 += '<div class="curriculum-header" data-id="' + id + '" onclick="toggleCourseList(this.dataset.id)">';
    html3 += '<span class="curriculum-title">' + escHtml(label) + "</span>";
    html3 += '<span class="curriculum-count">' + doneOf + " &#9660;</span>";
    html3 += "</div>";
    html3 += progBar(pct, barCls);
    html3 += courseList(curriculum, id);
    html3 += "</div>";
    return html3;
  }}

  var detailHtml = "";
  detailHtml += '<div class="roster-right-header">';
  detailHtml += '<div style="font-size:16px;font-weight:700;margin-bottom:6px">' + escHtml(p.FirstName + " " + p.LastName) + "</div>";
  detailHtml += '<span class="badge-status ' + badgeClass + '">' + badgeText + "</span>";
  if(isPastDue(p)) detailHtml += ' <span class="badge-status not-complete" style="margin-left:6px">Past Due</span>';
  if(isSalesCertified(p.Email)){{
    detailHtml += ' <span class="badge-status" style="margin-left:6px;background:#fef3c722;color:#b45309;border:1px solid var(--accent3)66">&#9733; Certified' + (salesCertDate(p.Email) ? " &middot; " + salesCertDate(p.Email) : "") + '</span>';
  }} else {{
    detailHtml += ' <span class="badge-status not-complete" style="margin-left:6px">Not Certified</span>';
  }}
  detailHtml += "</div>";
  detailHtml += '<div class="detail-grid">';
  detailHtml += '<div><div class="detail-label">Job Title</div><div class="detail-value">' + escHtml(p.JobTitle || "-") + "</div></div>";
  detailHtml += '<div><div class="detail-label">Market</div><div class="detail-value">' + escHtml(p.Market || "-") + "</div></div>";
  detailHtml += '<div><div class="detail-label">Hired</div><div class="detail-value">' + fmtDate(p.HireDate) + "</div></div>";
  detailHtml += '<div><div class="detail-label">Email</div><div class="detail-value"><a href="mailto:' + escHtml(p.Email) + '" style="color:var(--accent);text-decoration:none">' + escHtml(p.Email || "-") + "</a></div></div>";
  detailHtml += '<div><div class="detail-label">Assigned</div><div class="detail-value">' + fmtDate(p.AssignDate) + "</div></div>";
  detailHtml += '<div><div class="detail-label">Closed Won</div><div class="detail-value" style="color:var(--accent2)">' +
    (closedWonAmount(p.Email) > 0
      ? fmtMoney(closedWonAmount(p.Email)) + (closedWonCount(p.Email) > 1 ? " across " + closedWonCount(p.Email) + " deals" : (closedWonAccount(p.Email) ? " &middot; " + escHtml(closedWonAccount(p.Email)) : ""))
      : "&mdash;") + "</div></div>";
  if(closedWonCloseDate(p.Email)){{
    detailHtml += '<div><div class="detail-label">Close Date</div><div class="detail-value">' + escHtml(closedWonCloseDate(p.Email)) + "</div></div>";
  }}
  if(p.Manager){{
    detailHtml += '<div><div class="detail-label">Manager</div><div class="detail-value">' + escHtml(p.Manager) + "</div></div>";
    detailHtml += '<div><div class="detail-label">Manager Email</div><div class="detail-value"><a href="mailto:' + escHtml(p.MgrEmail) + '" style="color:var(--accent);text-decoration:none">' + escHtml(p.MgrEmail || "-") + "</a></div></div>";
  }}
  detailHtml += "</div>";
  detailHtml += '<hr style="border:none;border-top:1px solid var(--border);margin:16px 0;">';
  detailHtml += curriculumSection("Layered Security &middot; " + p.ls.done + " / " + p.ls.total + " modules", p.ls, "ls-" + email.replace(/[^a-z0-9]/gi, ""));
  detailHtml += '<div style="margin-top:14px;padding-top:12px;border-top:1px solid var(--border);font-size:12px;color:var(--muted);">';
  detailHtml += p.overallDone + " of " + p.ls.total + " modules complete (" + p.overallPct + "%)";
  detailHtml += "</div>";
  sel("roster-right").innerHTML = detailHtml;
}}

function toggleCourseList(id){{
  var el = document.getElementById("cl-" + id);
  if(el) el.classList.toggle("open");
}}

// ── Print / Export ──
function setupPrintHeader(title, subtitle){{
  sel("ph-title").textContent  = title;
  sel("ph-date").textContent   = subtitle;
  var market = sel("f-market").value || "All Markets";
  var status = sel("f-status").options[sel("f-status").selectedIndex].text;
  var search = sel("f-search").value;
  var parts  = ["Status: " + status, "Market: " + market];
  if(search) parts.push("Search: " + search);
  sel("ph-filters").textContent = parts.join("  |  ");
}}
function pBox(n, l){{
  return '<div style="min-width:90px"><div style="font-size:30px;font-weight:700;color:#1a3a5c;line-height:1">' + n + "</div>"
       + '<div style="font-size:10px;color:#666;text-transform:uppercase;letter-spacing:.06em;margin-top:5px">' + l + "</div></div>";
}}
function tds(cells){{ return "<tr>" + cells.map(function(c){{ return "<td>" + c + "</td>"; }}).join("") + "</tr>"; }}
function thRow(labels){{ return "<tr>" + labels.map(function(l){{ return "<th>" + l + "</th>"; }}).join("") + "</tr>"; }}

function runExport(type){{
  sel("export-menu").classList.remove("open");
  var now = new Date().toLocaleDateString("en-US", {{year:"numeric",month:"long",day:"numeric"}});
  if(type === "full"){{
    setupPrintHeader("Layered Security Curriculum Completion Report — Direct Sales", "Generated: " + now + "  |  " + filtered.length + " People");
    var total  = filtered.length;
    var cert   = filtered.filter(function(p){{ return personStatus(p) === "Complete"; }}).length;
    var inprog = filtered.filter(function(p){{ return personStatus(p) === "In Progress"; }}).length;
    var rate   = total > 0 ? Math.round(cert / total * 100) : 0;
    sel("print-stats").innerHTML = pBox(total,"Total Enrolled") + pBox(cert,"Complete") + pBox(inprog,"In Progress") + pBox(rate+"%","Completion Rate");
    sel("print-roster-head").innerHTML = thRow(["#","Name","Market","Job Title","Status","Layered Security %","Completion Date","Closed Won","Manager"]);
    sel("print-roster-body").innerHTML = filtered.map(function(p,i){{
      return tds([i+1,"<b>"+escHtml(p.FirstName+" "+p.LastName)+"</b>",escHtml(p.Market||"-"),escHtml(p.JobTitle||"-"),
        personStatus(p), p.ls.pct+"%", p.CompleteDate?fmtDate(p.CompleteDate):"-", closedWonAmount(p.Email) > 0 ? fmtMoney(closedWonAmount(p.Email)) : "-", escHtml(p.Manager||"-")]);
    }}).join("");
    sel("ph-desc").style.display = "none";
    document.body.classList.remove("print-no-summary");
    window.print();
  }} else if(type === "not-complete"){{
    var notCert = filtered.filter(function(p){{ return personStatus(p) !== "Complete"; }});
    setupPrintHeader("Not Complete: Layered Security", "Generated: " + now + "  |  " + notCert.length + " Employees");
    sel("print-stats").innerHTML = "";
    sel("print-roster-head").innerHTML = thRow(["#","Name","Email","Market","Layered Security %","Manager","Manager Email"]);
    sel("print-roster-body").innerHTML = notCert.length
      ? notCert.slice().sort(function(a,b){{ return (a.Manager||"").localeCompare(b.Manager||"")||(a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName); }})
          .map(function(p,i){{ return tds([i+1,"<b>"+escHtml(p.FirstName+" "+p.LastName)+"</b>",escHtml(p.Email||"-"),escHtml(p.Market||"-"),p.ls.pct+"%",escHtml(p.Manager||"-"),escHtml(p.MgrEmail||"-")]); }}).join("")
      : '<tr><td colspan="7" style="color:#999;font-style:italic;padding:10px">All enrolled people have completed the curriculum.</td></tr>';
    sel("ph-desc").textContent = "Employees who have not yet completed the Layered Security curriculum, sorted by manager.";
    sel("ph-desc").style.display = "block";
    document.body.classList.add("print-no-summary");
    window.print();
    document.body.classList.remove("print-no-summary");
  }} else if(type === "manager-summary"){{
    var mgrMap = {{}};
    filtered.forEach(function(p){{
      var k = p.Manager || "(No Manager)";
      if(!mgrMap[k]) mgrMap[k] = {{ name:k, email:p.MgrEmail||"-", total:0, cert:0, sumPct:0 }};
      mgrMap[k].total++;
      if(personStatus(p) === "Complete") mgrMap[k].cert++;
      mgrMap[k].sumPct += p.overallPct;
    }});
    var mgrs = Object.values(mgrMap).map(function(m){{ m.avgPct = m.total > 0 ? Math.round(m.sumPct / m.total) : 0; return m; }})
      .sort(function(a,b){{ return (b.cert/b.total) - (a.cert/a.total); }});
    setupPrintHeader("Manager Summary: Layered Security", "Generated: " + now + "  |  " + mgrs.length + " Managers");
    sel("print-stats").innerHTML = "";
    sel("print-roster-head").innerHTML = thRow(["Manager","Manager Email","Team Size","Complete","Avg Overall %"]);
    sel("print-roster-body").innerHTML = mgrs.map(function(m){{ return tds(["<b>"+escHtml(m.name)+"</b>",escHtml(m.email),m.total,m.cert,"<b>"+m.avgPct+"%</b>"]); }}).join("");
    sel("ph-desc").textContent = "Curriculum completion by manager, sorted from highest to lowest completion rate.";
    sel("ph-desc").style.display = "block";
    document.body.classList.add("print-no-summary");
    window.print();
    document.body.classList.remove("print-no-summary");
  }}
}}

// ── xlsx export ──
function runExportXLSX(type){{
  sel("export-menu").classList.remove("open");
  var now=new Date().toLocaleDateString("en-US",{{year:"numeric",month:"long",day:"numeric"}});
  function makeSheet(rows,colWidths){{
    var ws=XLSX.utils.aoa_to_sheet(rows);
    ws['!cols']=colWidths.map(function(w){{ return {{wch:w}}; }});
    return ws;
  }}
  function dlXLSX(name,wb){{ XLSX.writeFile(wb,name+'.xlsx'); }}
  var wb=XLSX.utils.book_new();
  if(type==="full"){{
    var total=filtered.length;
    var cert=filtered.filter(function(p){{ return personStatus(p)==="Complete"; }}).length;
    var rate=total>0?Math.round(cert/total*100):0;
    var inprogXL  = filtered.filter(function(p){{ return personStatus(p) === "In Progress"; }}).length;
    var nostartXL = filtered.filter(function(p){{ return personStatus(p) === "Not Started"; }}).length;
    var sumRows=[
      ["Layered Security Curriculum Completion Report — Direct Sales"],
      ["Generated:",now],
      [],
      ["SUMMARY"],
      ["Total Enrolled",total],
      ["Complete",cert],
      ["In Progress",inprogXL],
      ["Not Started",nostartXL],
      ["Completion Rate",rate+"%"],
    ];
    XLSX.utils.book_append_sheet(wb,makeSheet(sumRows,[30,18]),"Summary");
    var rRows=[["Name","Market","Job Title","Status","Layered Security %","Overall %","Cert Date","Closed Won $","Manager"]];
    filtered.forEach(function(p){{
      rRows.push([p.FirstName+" "+p.LastName,p.Market||"-",p.JobTitle||"-",personStatus(p),p.ls.pct+"%",p.overallPct+"%",p.CompleteDate||"-",closedWonAmount(p.Email),p.Manager||"-"]);
    }});
    XLSX.utils.book_append_sheet(wb,makeSheet(rRows,[28,18,28,14,16,12,14,12,28]),"Roster");
    dlXLSX("ls-full-report",wb);
  }} else if(type==="not-complete"){{
    var notCert=filtered.filter(function(p){{ return personStatus(p)!=="Complete"; }}).slice().sort(function(a,b){{
      return (a.Manager||"").localeCompare(b.Manager||"")||(a.LastName+a.FirstName).localeCompare(b.LastName+b.FirstName);
    }});
    var rows=[["Name","Email","Market","Layered Security %","Manager","Manager Email"]];
    notCert.forEach(function(p){{
      rows.push([p.FirstName+" "+p.LastName,p.Email||"-",p.Market||"-",p.ls.pct+"%",p.Manager||"-",p.MgrEmail||"-"]);
    }});
    XLSX.utils.book_append_sheet(wb,makeSheet(rows,[28,32,18,16,28,32]),"Not Complete");
    dlXLSX("ls-not-complete",wb);
  }} else if(type==="manager-summary"){{
    var mgrMap={{}};
    filtered.forEach(function(p){{
      var k=p.Manager||"(No Manager)";
      if(!mgrMap[k]) mgrMap[k]={{name:k,email:p.MgrEmail||"-",total:0,cert:0,sumPct:0}};
      mgrMap[k].total++;
      if(personStatus(p)==="Complete") mgrMap[k].cert++;
      mgrMap[k].sumPct+=p.overallPct;
    }});
    var rows=[["Manager","Manager Email","Team Size","Complete","Avg Overall %"]];
    Object.values(mgrMap).map(function(m){{ m.avgPct=m.total>0?Math.round(m.sumPct/m.total):0; return m; }})
      .sort(function(a,b){{ return (b.cert/b.total)-(a.cert/a.total); }})
      .forEach(function(m){{ rows.push([m.name,m.email,m.total,m.cert,m.avgPct+"%"]); }});
    XLSX.utils.book_append_sheet(wb,makeSheet(rows,[28,32,12,12,14]),"Manager Summary");
    dlXLSX("ls-manager-summary",wb);
  }}
}}

// ── init ──
applyFilters();
var firstRow = sel("roster-left").querySelector(".progress-row");
if(firstRow) showDetail(firstRow.dataset.email);
</script>
</body>
</html>"""


def main():
    people, date_label = load_ls_data()
    if not people:
        return
    sales_cert  = load_sales_cert()
    sales_deals = load_sales_deals()
    html = generate_html(people, date_label, sales_cert, sales_deals)
    with open('cert-layered-security.html', 'w', encoding='utf-8') as f:
        f.write(html)
    complete_n = sum(1 for p in people if p['Complete'] == 'Yes')
    print(f"Generated cert-layered-security.html — {len(people)} learners, {complete_n} complete")


if __name__ == '__main__':
    main()
